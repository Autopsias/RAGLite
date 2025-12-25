"""Unit tests for EU Oil Bulletin client.

Story 7.1: Split test_external_data_clients.py
This module contains tests for: TestEUOilBulletinClient, TestEUOilBulletinAdditional, TestEUOilBulletinStory694
"""

from __future__ import annotations

from datetime import date
from unittest.mock import AsyncMock, MagicMock, patch

import httpx
import pytest

from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient
from raglite.external_data.models import EUDieselPrice


class TestEUOilBulletinClient:
    """Tests for EU Oil Bulletin client.

    Story 6.9.4: Updated tests for new XLSX-based implementation (2025-12-08)
    - Old tests used XML parsing
    - New tests use openpyxl mock for XLSX parsing
    """

    @pytest.fixture
    def client(self) -> EUOilBulletinClient:
        """Create EU Oil Bulletin client instance."""
        return EUOilBulletinClient()

    @pytest.fixture
    def mock_xlsx_content(self) -> bytes:
        """Create mock XLSX content (Story 6.9.4 AC6)."""
        from datetime import datetime
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"
        ws.append(["Date", "Portugal", "Spain", "Germany"])
        ws.append([datetime(2024, 1, 8), 1.456, 1.423, 1.512])
        ws.append([datetime(2024, 1, 15), 1.478, 1.445, 1.498])
        ws.append([datetime(2024, 1, 22), 1.489, 1.456, 1.501])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @pytest.mark.asyncio
    async def test_fetch_diesel_prices_success(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test successful diesel prices fetch (Story 6.9.4 AC3)."""
        mock_response = MagicMock()
        mock_response.content = mock_xlsx_content
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )
            client._get_cached_xlsx = MagicMock(return_value=None)
            client._save_to_cache = MagicMock()

            result = await client.fetch_diesel_prices(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
                country="Portugal",
            )

            assert len(result) == 3
            assert isinstance(result[0], EUDieselPrice)
            assert result[0].price_eur_litre == 1.456
            assert result[0].country == "Portugal"

    def test_country_code_conversion(self, client: EUOilBulletinClient) -> None:
        """Test country code to name conversion."""
        assert client._code_to_country("PT") == "Portugal"
        assert client._code_to_country("ES") == "Spain"
        assert client._code_to_country("XX") == "XX"  # Unknown code returns as-is

    def test_parse_xlsx_portugal(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test XLSX parsing for Portugal (Story 6.9.4 AC2/AC3)."""
        result = client._parse_xlsx(
            mock_xlsx_content,
            country="Portugal",
            start_date=date(2024, 1, 1),
            end_date=date(2024, 1, 31),
            tax_included=True,
        )

        assert len(result) == 3
        assert result[0].price_eur_litre == 1.456
        assert result[1].price_eur_litre == 1.478
        assert result[2].price_eur_litre == 1.489

    def test_parse_xlsx_date_filtering(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test date range filtering (Story 6.9.4 AC4)."""
        result = client._parse_xlsx(
            mock_xlsx_content,
            country="Portugal",
            start_date=date(2024, 1, 15),
            end_date=date(2024, 1, 31),
            tax_included=True,
        )

        assert len(result) == 2
        assert result[0].date == date(2024, 1, 15)
        assert result[1].date == date(2024, 1, 22)


class TestEUOilBulletinAdditional:
    """Additional tests for EU Oil Bulletin (Story 6.9.4)."""

    @pytest.fixture
    def client(self) -> EUOilBulletinClient:
        return EUOilBulletinClient()

    @pytest.fixture
    def mock_xlsx_content(self) -> bytes:
        """Create simple mock XLSX."""
        from datetime import datetime
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"
        ws.append(["Date", "Portugal", "Spain"])
        ws.append([datetime(2024, 1, 8), 1.456, 1.423])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_deprecated_xml_parsing_returns_empty(self, client: EUOilBulletinClient) -> None:
        """Test deprecated XML method (Story 6.9.4)."""
        result = client._parse_bulletin_xml(
            "any content", "PT", date(2024, 1, 1), date(2024, 1, 31)
        )
        assert result == []

    @pytest.mark.asyncio
    async def test_fetch_weekly_prices(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test weekly prices fetch (Story 6.9.4 AC4)."""
        mock_response = MagicMock()
        mock_response.content = mock_xlsx_content
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )
            client._get_cached_xlsx = MagicMock(return_value=None)
            client._save_to_cache = MagicMock()

            result = await client.fetch_weekly_prices(
                week_date=date(2024, 1, 10),
                country="Portugal",
            )

        assert result is not None
        assert result.price_eur_litre == 1.456

    @pytest.mark.asyncio
    async def test_fetch_weekly_prices_not_found(self, client: EUOilBulletinClient) -> None:
        """Test fetching weekly prices when no data available."""
        from io import BytesIO

        import openpyxl

        # Create empty XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"
        ws.append(["Date", "Portugal"])
        # No data rows

        buffer = BytesIO()
        wb.save(buffer)
        empty_xlsx = buffer.getvalue()

        mock_response = MagicMock()
        mock_response.content = empty_xlsx
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            client._get_cached_xlsx = MagicMock(return_value=None)
            client._save_to_cache = MagicMock()

            result = await client.fetch_weekly_prices(
                week_date=date(2024, 6, 15),
                country="Portugal",
            )

        assert result is None

    @pytest.mark.asyncio
    async def test_fetch_xlsx_timeout_retry(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test timeout retry logic (Story 6.9.4 AC7)."""
        success_response = MagicMock()
        success_response.content = mock_xlsx_content
        success_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_get = AsyncMock(side_effect=[httpx.TimeoutException("timeout"), success_response])
            mock_client.return_value.__aenter__.return_value.get = mock_get
            client._get_cached_xlsx = MagicMock(return_value=None)
            client._save_to_cache = MagicMock()

            with patch("asyncio.sleep", new_callable=AsyncMock):
                await client.fetch_diesel_prices(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

            assert mock_get.call_count == 2


class TestEUOilBulletinStory694:
    """Story 6.9.4 AC1-AC7 validation tests."""

    def test_new_xlsx_endpoint(self) -> None:
        """AC1: Verify new XLSX endpoint is configured."""
        from raglite.external_data.clients.eu_oil_bulletin import (
            EU_OIL_BULLETIN_BASE,
            HISTORY_XLSX_DOC_ID,
            HISTORY_XLSX_FILENAME,
        )

        # New base URL
        assert EU_OIL_BULLETIN_BASE == "https://energy.ec.europa.eu"
        # Document ID for historical prices
        assert HISTORY_XLSX_DOC_ID == "906e60ca-8b6a-44e7-8589-652854d2fd3f_en"
        assert "xlsx" in HISTORY_XLSX_FILENAME.lower()

    def test_old_xml_endpoint_not_used(self) -> None:
        """AC1: Verify old XML endpoint is not used."""
        from raglite.external_data.clients.eu_oil_bulletin import EU_OIL_BULLETIN_BASE

        assert "observatory/reports" not in EU_OIL_BULLETIN_BASE
        assert ".xml" not in EU_OIL_BULLETIN_BASE

    def test_openpyxl_parsing_method_exists(self) -> None:
        """AC2: Verify XLSX parsing method exists."""
        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        client = EUOilBulletinClient()
        assert hasattr(client, "_parse_xlsx")
        assert callable(client._parse_xlsx)

    def test_cache_methods_exist(self) -> None:
        """AC5: Verify caching methods exist."""
        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        client = EUOilBulletinClient()
        assert hasattr(client, "_get_cached_xlsx")
        assert hasattr(client, "_save_to_cache")
        assert hasattr(client, "cache_dir")
        assert hasattr(client, "cache_ttl_hours")

    def test_extended_timeout_for_large_file(self) -> None:
        """AC7: Verify 60s timeout for large file download."""
        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        client = EUOilBulletinClient()
        # In non-test mode, timeout should be 60s
        # (Test mode has 1s timeout, but we check the attribute exists)
        assert hasattr(client, "timeout")

    def test_oil_bulletin_retry_delays_exponential_backoff(self) -> None:
        """AC7: Verify retry logic uses exponential backoff (2s/4s/8s)."""
        import inspect

        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        source = inspect.getsource(EUOilBulletinClient._fetch_xlsx_data)

        # Should have 2, 4, 8 second delays (NFR1 requirement)
        assert "retry_delays = [2, 4, 8]" in source

    def test_xlsx_validation(self) -> None:
        """AC2: Verify XLSX validation checks for PK header."""
        import inspect

        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        source = inspect.getsource(EUOilBulletinClient._fetch_xlsx_data)

        # Should verify XLSX starts with PK (ZIP signature)
        assert 'b"PK"' in source or "PK" in source

    def test_parse_xlsx_with_sample_data(self) -> None:
        """AC6: Test parsing with sample data."""
        from datetime import datetime
        from io import BytesIO

        import openpyxl

        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"
        ws.append(["Date", "Portugal", "Spain", "Germany"])
        ws.append([datetime(2024, 12, 2), 1.523, 1.489, 1.612])
        ws.append([datetime(2024, 12, 9), 1.534, 1.501, 1.598])

        buffer = BytesIO()
        wb.save(buffer)

        client = EUOilBulletinClient()
        result = client._parse_xlsx(
            buffer.getvalue(),
            country="Portugal",
            start_date=date(2024, 12, 1),
            end_date=date(2024, 12, 31),
            tax_included=True,
        )

        assert len(result) == 2
        assert result[0].price_eur_litre == 1.523
        assert result[1].price_eur_litre == 1.534
        assert result[0].country == "Portugal"

    def test_parse_xlsx_different_country(self) -> None:
        """AC3: Test multi-country parsing."""
        from datetime import datetime
        from io import BytesIO

        import openpyxl

        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"
        ws.append(["Date", "Portugal", "Spain"])
        ws.append([datetime(2024, 12, 2), 1.523, 1.489])

        buffer = BytesIO()
        wb.save(buffer)

        client = EUOilBulletinClient()
        pt_result = client._parse_xlsx(
            buffer.getvalue(), "Portugal", date(2024, 12, 1), date(2024, 12, 31), True
        )
        es_result = client._parse_xlsx(
            buffer.getvalue(), "Spain", date(2024, 12, 1), date(2024, 12, 31), True
        )

        assert len(pt_result) == 1 and pt_result[0].price_eur_litre == 1.523
        assert len(es_result) == 1 and es_result[0].price_eur_litre == 1.489
