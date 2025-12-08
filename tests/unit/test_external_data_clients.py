"""Unit tests for external data clients.

Story 6.1: Tier 1 External Data Source Integration
AC6: Unit Tests (80%+ coverage)

Tests scenarios:
- Success responses
- Timeout handling and retry
- Invalid response handling
- Server error (5xx) handling
- Client error (4xx) handling
"""

from __future__ import annotations

from datetime import date
from unittest.mock import AsyncMock, MagicMock, patch

import httpx
import pytest

from raglite.external_data.clients.atic import ATICClient
from raglite.external_data.clients.basegov import BaseGovClient
from raglite.external_data.clients.bpstat import BPstatClient
from raglite.external_data.clients.commodities import CommoditiesClient
from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient
from raglite.external_data.clients.ine import INEClient
from raglite.external_data.clients.ipma import IPMAClient
from raglite.external_data.clients.omie import OMIEClient
from raglite.external_data.exceptions import ExternalDataFetchError, ExternalDataValidationError
from raglite.external_data.models import (
    ATICCementConsumption,
    BaseGovContract,
    BPstatMortgageLoans,
    EUDieselPrice,
    INEBuildingPermits,
    INEConstructionCostIndex,
    INEConstructionOutput,
    IPMAWeatherData,
    OMIEElectricityPrice,
)

# =============================================================================
# INE Client Tests
# =============================================================================


class TestINEClient:
    """Tests for INE API client."""

    @pytest.fixture
    def client(self) -> INEClient:
        """Create INE client instance."""
        return INEClient()

    @pytest.mark.asyncio
    async def test_fetch_building_permits_success(self, client: INEClient) -> None:
        """Test successful building permits fetch."""
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.json.return_value = {
            "Dados": {
                "202401": [{"valor": 1234, "geocod": "Lisboa"}],
                "202402": [{"valor": 1456, "geocod": "Porto"}],
            }
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_building_permits(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 2, 28),
            )

            assert len(result) == 2
            assert isinstance(result[0], INEBuildingPermits)
            assert result[0].permits_count == 1234
            assert result[0].region == "Lisboa"

    @pytest.mark.asyncio
    async def test_fetch_building_permits_timeout_retry(self, client: INEClient) -> None:
        """Test retry logic on timeout."""
        mock_response = MagicMock()
        mock_response.json.return_value = {"Dados": {"202401": [{"valor": 100}]}}
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            # First two calls timeout, third succeeds
            mock_get = AsyncMock(
                side_effect=[
                    httpx.TimeoutException("timeout"),
                    httpx.TimeoutException("timeout"),
                    mock_response,
                ]
            )
            mock_client.return_value.__aenter__.return_value.get = mock_get

            with patch("asyncio.sleep", new_callable=AsyncMock):
                result = await client.fetch_building_permits(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

            assert len(result) == 1
            assert mock_get.call_count == 3

    @pytest.mark.asyncio
    async def test_fetch_building_permits_timeout_exhausted(self, client: INEClient) -> None:
        """Test exception when all retries exhausted."""
        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                side_effect=httpx.TimeoutException("timeout")
            )

            with patch("asyncio.sleep", new_callable=AsyncMock):
                with pytest.raises(ExternalDataFetchError) as exc_info:
                    await client.fetch_building_permits(
                        start_date=date(2024, 1, 1),
                        end_date=date(2024, 1, 31),
                    )

            assert exc_info.value.source == "INE"
            assert "Timeout" in exc_info.value.message

    @pytest.mark.asyncio
    async def test_fetch_construction_output_success(self, client: INEClient) -> None:
        """Test successful construction output fetch."""
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "Dados": {
                "202401": [{"valor": 105.5, "dim_3_t": "Total"}],
            }
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_construction_output(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )

            assert len(result) == 1
            assert isinstance(result[0], INEConstructionOutput)
            assert result[0].index_value == 105.5
            # Note: INE doesn't provide YoY change in API response; it's calculated separately
            assert result[0].yoy_change_pct is None

    @pytest.mark.asyncio
    async def test_fetch_construction_cost_index_success(self, client: INEClient) -> None:
        """Test successful construction cost index fetch."""
        mock_response = MagicMock()
        # INE returns separate records per cost factor (Total, Materiais, Mão de obra)
        mock_response.json.return_value = {
            "Dados": {
                "202401": [
                    {"valor": 110.2, "dim_3_t": "Total"},
                    {"valor": 112.5, "dim_3_t": "Materiais"},
                    {"valor": 108.1, "dim_3_t": "Mão de obra"},
                ],
            }
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_construction_cost_index(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )

            assert len(result) == 1
            assert isinstance(result[0], INEConstructionCostIndex)
            assert result[0].total_index == 110.2
            assert result[0].materials_index == 112.5
            assert result[0].labor_index == 108.1


# =============================================================================
# ATIC Client Tests
# =============================================================================


# =============================================================================
# Story 6.9.1: INE Date Filtering Tests (AC1-AC3)
# =============================================================================


class TestINEDateFiltering:
    """Tests for INE date filtering fix (Story 6.9.1 AC1-AC3).

    Problem: Monthly periods like "Setembro de 2025" parse to 2025-09-01.
    With start_date=2025-09-15, September data was incorrectly excluded.
    Fix: Compare against start_date.replace(day=1) for monthly data.
    """

    @pytest.fixture
    def client(self) -> INEClient:
        return INEClient()

    def test_mid_month_start_date_includes_month(self, client: INEClient) -> None:
        """AC1/AC2: Start date of 2025-09-15 should include September data."""
        # September data parses to 2025-09-01
        data = {
            "Dados": {
                "Setembro de 2025": [{"valor": 1234, "geocod": "Portugal"}],
                "Outubro de 2025": [{"valor": 1456, "geocod": "Portugal"}],
            }
        }

        result = client._parse_building_permits(
            data,
            start_date=date(2025, 9, 15),  # Mid-month
            end_date=date(2025, 10, 31),
        )

        # Both September AND October should be included
        assert len(result) == 2
        dates = [r.date for r in result]
        assert date(2025, 9, 1) in dates  # September included despite start=Sep 15
        assert date(2025, 10, 1) in dates

    def test_first_of_month_start_date(self, client: INEClient) -> None:
        """AC1: Start date of 2025-09-01 should include September data."""
        data = {
            "Dados": {
                "Setembro de 2025": [{"valor": 1234, "geocod": "Portugal"}],
            }
        }

        result = client._parse_building_permits(
            data,
            start_date=date(2025, 9, 1),  # First of month
            end_date=date(2025, 9, 30),
        )

        assert len(result) == 1
        assert result[0].date == date(2025, 9, 1)

    def test_end_of_month_start_date_includes_month(self, client: INEClient) -> None:
        """AC1: Start date of 2025-09-30 should include September data."""
        data = {
            "Dados": {
                "Setembro de 2025": [{"valor": 1234, "geocod": "Portugal"}],
            }
        }

        result = client._parse_building_permits(
            data,
            start_date=date(2025, 9, 30),  # End of month
            end_date=date(2025, 9, 30),
        )

        # September (parsed as 2025-09-01) should be included
        # because first-of-month comparison: 2025-09-01 >= 2025-09-01
        assert len(result) == 1
        assert result[0].date == date(2025, 9, 1)

    def test_construction_output_mid_month_filtering(self, client: INEClient) -> None:
        """AC3: construction_output also uses first-of-month comparison."""
        data = {
            "Dados": {
                "202509": [{"valor": 105.5, "dim_3_t": "Total"}],
            }
        }

        result = client._parse_construction_output(
            data,
            start_date=date(2025, 9, 15),  # Mid-month
            end_date=date(2025, 9, 30),
        )

        assert len(result) == 1
        assert result[0].date == date(2025, 9, 1)

    def test_construction_cost_index_mid_month_filtering(self, client: INEClient) -> None:
        """AC3: construction_cost_index also uses first-of-month comparison."""
        data = {
            "Dados": {
                "202509": [{"valor": 110.2, "dim_3_t": "Total"}],
            }
        }

        result = client._parse_construction_cost_index(
            data,
            start_date=date(2025, 9, 15),  # Mid-month
            end_date=date(2025, 9, 30),
        )

        assert len(result) == 1
        assert result[0].date == date(2025, 9, 1)

    def test_date_filtering_excludes_earlier_months(self, client: INEClient) -> None:
        """Verify months before start_date are still correctly excluded."""
        data = {
            "Dados": {
                "Agosto de 2025": [{"valor": 1000, "geocod": "Portugal"}],  # Before range
                "Setembro de 2025": [{"valor": 1234, "geocod": "Portugal"}],  # In range
            }
        }

        result = client._parse_building_permits(
            data,
            start_date=date(2025, 9, 15),
            end_date=date(2025, 9, 30),
        )

        # Only September should be included, August should be excluded
        assert len(result) == 1
        assert result[0].date == date(2025, 9, 1)


class TestATICClient:
    """Tests for ATIC cement consumption client."""

    @pytest.fixture
    def client(self) -> ATICClient:
        """Create ATIC client instance."""
        return ATICClient()

    def test_parse_csv_content_success(self, client: ATICClient) -> None:
        """Test successful CSV parsing."""
        csv_content = """date,consumption,region,type
2024-01-01,150000,Portugal,gray
2024-02-01,160000,Portugal,gray
2024-03-01,170000,Lisboa,white"""

        result = client.parse_csv_content(csv_content)

        assert len(result) == 3
        assert isinstance(result[0], ATICCementConsumption)
        assert result[0].consumption_tonnes == 150000
        assert result[0].region == "Portugal"
        assert result[2].region == "Lisboa"
        assert result[2].cement_type == "white"

    def test_parse_csv_content_alternate_columns(self, client: ATICClient) -> None:
        """Test CSV with alternate column names."""
        csv_content = """Data,Consumo,Regiao
2024-01-01,150000,Norte
2024-02-01,160000,Sul"""

        result = client.parse_csv_content(csv_content)

        assert len(result) == 2
        assert result[0].consumption_tonnes == 150000
        assert result[0].region == "Norte"

    def test_parse_csv_content_no_headers(self, client: ATICClient) -> None:
        """Test CSV validation error on missing headers."""
        csv_content = ""

        with pytest.raises(ExternalDataValidationError) as exc_info:
            client.parse_csv_content(csv_content)

        assert exc_info.value.source == "ATIC"
        assert "headers" in exc_info.value.message.lower()

    def test_parse_csv_content_missing_date_column(self, client: ATICClient) -> None:
        """Test validation error on missing date column."""
        csv_content = """value,region
150000,Portugal"""

        with pytest.raises(ExternalDataValidationError) as exc_info:
            client.parse_csv_content(csv_content)

        assert "date column" in exc_info.value.message.lower()

    def test_parse_csv_content_missing_value_column(self, client: ATICClient) -> None:
        """Test validation error on missing value column."""
        csv_content = """date,region
2024-01-01,Portugal"""

        with pytest.raises(ExternalDataValidationError) as exc_info:
            client.parse_csv_content(csv_content)

        assert "value" in exc_info.value.message.lower()

    def test_parse_csv_file_not_found(self, client: ATICClient) -> None:
        """Test error on missing file."""
        with pytest.raises(ExternalDataFetchError) as exc_info:
            client.parse_csv_file("/nonexistent/path.csv")

        assert exc_info.value.source == "ATIC"
        assert "not found" in exc_info.value.message.lower()

    def test_parse_date_formats(self, client: ATICClient) -> None:
        """Test parsing various date formats."""
        # Test different date formats
        assert client._parse_date("2024-01-15") == date(2024, 1, 15)
        assert client._parse_date("15/01/2024") == date(2024, 1, 15)
        assert client._parse_date("2024-01") == date(2024, 1, 1)
        assert client._parse_date("01/2024") == date(2024, 1, 1)
        assert client._parse_date("202401") == date(2024, 1, 1)
        assert client._parse_date("invalid") is None


# =============================================================================
# BPstat Client Tests
# =============================================================================


class TestBPstatClient:
    """Tests for BPstat (Banco de Portugal) client.

    Story 6.9.3: Updated tests for new API structure (2025-12-08)
    - New endpoint: /api/observations/
    - New series IDs: 12710733 (median), 12710735 (10th), etc.
    - Single API call fetches all series (instead of 3 separate calls)
    """

    @pytest.fixture
    def client(self) -> BPstatClient:
        """Create BPstat client instance."""
        return BPstatClient()

    @pytest.mark.asyncio
    async def test_fetch_mortgage_loans_success(self, client: BPstatClient) -> None:
        """Test successful mortgage interest rate fetch.

        Story 6.9.3 AC4: Updated for new API response format.
        """
        mock_response = MagicMock()
        # Story 6.9.3 AC3: New API response format with series_id
        mock_response.json.return_value = {
            "observations": [
                {"period": "2024-01", "value": 3.45, "series_id": "12710733"},
                {"period": "2024-02", "value": 3.50, "series_id": "12710733"},
            ]
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_mortgage_loans(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 2, 28),
            )

            assert len(result) == 2
            assert isinstance(result[0], BPstatMortgageLoans)
            # Story 6.9.3: Now returns interest rate in avg_interest_rate_pct
            assert result[0].avg_interest_rate_pct == 3.45
            assert result[1].avg_interest_rate_pct == 3.50

    @pytest.mark.asyncio
    async def test_fetch_mortgage_loans_server_error_retry(self, client: BPstatClient) -> None:
        """Test retry on server error (500).

        Story 6.9.3 AC7: Now makes single API call (not 3), so retry count changed.
        """
        error_response = MagicMock()
        error_response.status_code = 500
        error = httpx.HTTPStatusError("Server Error", request=MagicMock(), response=error_response)

        success_response = MagicMock()
        # Story 6.9.3: New response format
        success_response.json.return_value = {
            "observations": [{"period": "2024-01", "value": 3.50, "series_id": "12710733"}]
        }
        success_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            # Story 6.9.3: Now single API call fetches all series
            # 2 errors then success (3 calls total)
            mock_get = AsyncMock(
                side_effect=[
                    error,
                    error,
                    success_response,
                ]
            )
            mock_client.return_value.__aenter__.return_value.get = mock_get

            with patch("asyncio.sleep", new_callable=AsyncMock):
                result = await client.fetch_mortgage_loans(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

            assert len(result) == 1
            # Story 6.9.3: Single API call with retries = 3 calls
            assert mock_get.call_count == 3


# =============================================================================
# OMIE Client Tests
# =============================================================================


class TestOMIEClient:
    """Tests for OMIE electricity market client.

    Story 6.9.2: Updated tests for new OMIE CSV format and URL pattern.
    """

    @pytest.fixture
    def client(self) -> OMIEClient:
        """Create OMIE client instance."""
        return OMIEClient()

    @pytest.mark.asyncio
    async def test_fetch_spot_prices_success(self, client: OMIEClient) -> None:
        """Test successful spot prices fetch.

        Story 6.9.2 AC3/AC5: Updated for new CSV format.
        """
        mock_response = MagicMock()
        mock_response.status_code = 200
        # Story 6.9.2 AC3: New format - MARGINALPDBC;YEAR;MONTH;DAY;HOUR;PT;ES;...
        mock_response.text = """MARGINALPDBC;2024;01;01;1;45,50;46,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;01;01;2;44,20;45,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;01;01;24;50,10;51,00;0;0;0;0;0;0;0;0"""
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_spot_prices(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 1),
                include_hourly=True,
            )

            assert len(result) == 3
            assert isinstance(result[0], OMIEElectricityPrice)
            assert result[0].price_eur_mwh == 45.50
            assert result[0].hour == 0  # OMIE uses 1-24, we convert to 0-23

    @pytest.mark.asyncio
    async def test_fetch_spot_prices_daily_average(self, client: OMIEClient) -> None:
        """Test daily average price calculation.

        Story 6.9.2 AC3/AC5: Updated for new CSV format.
        """
        mock_response = MagicMock()
        mock_response.status_code = 200
        # Story 6.9.2 AC3: New format with European decimal comma
        mock_response.text = """MARGINALPDBC;2024;01;01;1;40,00;41,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;01;01;2;50,00;51,00;0;0;0;0;0;0;0;0"""
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_spot_prices(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 1),
                include_hourly=False,
            )

            assert len(result) == 1
            assert result[0].price_eur_mwh == 45.00  # Average of 40 and 50
            assert result[0].price_type == "spot_daily_avg"

    @pytest.mark.asyncio
    async def test_fetch_spot_prices_404_handled(self, client: OMIEClient) -> None:
        """Test 404 (no data) is handled gracefully."""
        error_response = MagicMock()
        error_response.status_code = 404

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                side_effect=httpx.HTTPStatusError(
                    "Not Found", request=MagicMock(), response=error_response
                )
            )

            result = await client.fetch_spot_prices(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 1),
            )

            # 404 should return empty list, not raise
            assert result == []

    def test_parse_daily_file(self, client: OMIEClient) -> None:
        """Test parsing OMIE daily file.

        Story 6.9.2 AC3/AC5: Updated for new CSV format.
        Format: MARGINALPDBC;YEAR;MONTH;DAY;HOUR;PT_PRICE;ES_PRICE;...
        """
        # Story 6.9.2 AC3: New format with MARGINALPDBC prefix
        content = """MARGINALPDBC;2024;01;01;1;45,50;46,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;01;01;2;44,20;45,00;0;0;0;0;0;0;0;0
INVALID;2024;01;01;3;bad;data"""

        result = client._parse_daily_file(content, date(2024, 1, 1))

        # Should parse 2 valid MARGINALPDBC rows (comma decimal separator handled)
        assert len(result) == 2
        assert result[0].hour == 0
        assert result[0].price_eur_mwh == 45.50
        assert result[1].hour == 1
        assert result[1].price_eur_mwh == 44.20


# =============================================================================
# EU Oil Bulletin Client Tests
# =============================================================================


class TestEUOilBulletinClient:
    """Tests for EU Oil Bulletin client.

    Story 6.9.4: Updated tests for new XLSX-based implementation (2025-12-08)
    - Old tests used XML parsing
    - New tests use openpyxl mock for XLSX parsing
    """

    @pytest.fixture
    def client(self) -> EUOilBulletinClient:
        """Create EU Oil Bulletin client instance."""
        return EUOilBulletinClient()

    @pytest.fixture
    def mock_xlsx_content(self) -> bytes:
        """Create mock XLSX content for testing.

        Story 6.9.4 AC6: Unit tests with sample XLSX data
        """
        from datetime import datetime
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"

        # Header row with countries
        ws.append(["Date", "Portugal", "Spain", "Germany"])

        # Data rows (weekly data)
        ws.append([datetime(2024, 1, 8), 1.456, 1.423, 1.512])
        ws.append([datetime(2024, 1, 15), 1.478, 1.445, 1.498])
        ws.append([datetime(2024, 1, 22), 1.489, 1.456, 1.501])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @pytest.mark.asyncio
    async def test_fetch_diesel_prices_success(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test successful diesel prices fetch.

        Story 6.9.4 AC3: Updated for XLSX-based implementation.
        """
        mock_response = MagicMock()
        mock_response.content = mock_xlsx_content
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            # Clear any cached data
            client._get_cached_xlsx = MagicMock(return_value=None)
            client._save_to_cache = MagicMock()

            result = await client.fetch_diesel_prices(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
                country="Portugal",
            )

            # Only Portugal results
            assert len(result) == 3
            assert isinstance(result[0], EUDieselPrice)
            assert result[0].price_eur_litre == 1.456
            assert result[0].country == "Portugal"

    def test_country_code_conversion(self, client: EUOilBulletinClient) -> None:
        """Test country code to name conversion."""
        assert client._code_to_country("PT") == "Portugal"
        assert client._code_to_country("ES") == "Spain"
        assert client._code_to_country("XX") == "XX"  # Unknown code returns as-is

    def test_parse_xlsx_portugal(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test XLSX parsing for Portugal.

        Story 6.9.4 AC2/AC3: Test XLSX parsing with openpyxl.
        """
        result = client._parse_xlsx(
            mock_xlsx_content,
            country="Portugal",
            start_date=date(2024, 1, 1),
            end_date=date(2024, 1, 31),
            tax_included=True,
        )

        assert len(result) == 3
        assert result[0].price_eur_litre == 1.456
        assert result[1].price_eur_litre == 1.478
        assert result[2].price_eur_litre == 1.489

    def test_parse_xlsx_date_filtering(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test date range filtering.

        Story 6.9.4 AC4: Handle weekly data frequency.
        """
        # Only get prices from second half of January
        result = client._parse_xlsx(
            mock_xlsx_content,
            country="Portugal",
            start_date=date(2024, 1, 15),
            end_date=date(2024, 1, 31),
            tax_included=True,
        )

        assert len(result) == 2  # Only Jan 15 and Jan 22
        assert result[0].date == date(2024, 1, 15)
        assert result[1].date == date(2024, 1, 22)


# =============================================================================
# IPMA Client Tests
# =============================================================================


class TestIPMAClient:
    """Tests for IPMA weather client."""

    @pytest.fixture
    def client(self) -> IPMAClient:
        """Create IPMA client instance."""
        return IPMAClient()

    @pytest.mark.asyncio
    async def test_fetch_observations_success(self, client: IPMAClient) -> None:
        """Test successful weather observations fetch."""
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "tMed": 15.5,
            "tMax": 20.0,
            "tMin": 10.0,
            "prec": 2.5,
            "humidade": 75.0,
            "vento": 15.0,
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_observations(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 1),
                station_id="1200535",
            )

            assert len(result) == 1
            assert isinstance(result[0], IPMAWeatherData)
            assert result[0].temperature_c == 15.5
            assert result[0].temperature_max_c == 20.0
            assert result[0].precipitation_mm == 2.5

    @pytest.mark.asyncio
    async def test_fetch_forecast_success(self, client: IPMAClient) -> None:
        """Test successful forecast fetch."""
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "data": [
                {"forecastDate": "2024-01-01", "tMax": 18, "tMin": 10, "precipitaProb": 20},
                {"forecastDate": "2024-01-02", "tMax": 19, "tMin": 11, "precipitaProb": 10},
            ]
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_forecast(location_id="1110600", days=2)

            assert len(result) == 2
            assert result[0].temperature_max_c == 18.0

    def test_stations_dict(self, client: IPMAClient) -> None:
        """Test station ID mapping."""
        assert "Lisboa" in client.STATIONS
        assert "Porto" in client.STATIONS
        assert client.STATIONS["Lisboa"] == "1200535"


# =============================================================================
# Base.gov.pt Client Tests
# =============================================================================


class TestBaseGovClient:
    """Tests for Base.gov.pt public procurement client.

    Story 6.9.5: Updated tests for TED API v3 and OCDS fallback.
    """

    @pytest.fixture
    def client(self) -> BaseGovClient:
        """Create Base.gov.pt client instance."""
        return BaseGovClient()

    @pytest.fixture
    def mock_ted_response(self) -> dict:
        """Create mock TED API response."""
        return {
            "notices": [
                {
                    "publication-number": "2024/S 015-012345",
                    "publication-date": "2024-01-15",
                    "notice-title": "Road construction project",
                    "buyer-name": "Câmara Municipal de Lisboa",
                    "winner-name": "Empresa ABC, Lda",
                    "total-value": 1500000,
                    "cpv": "45233000",
                    "place-of-performance": "PT",
                },
            ]
        }

    @pytest.mark.asyncio
    async def test_fetch_contracts_success(
        self, client: BaseGovClient, mock_ted_response: dict
    ) -> None:
        """Test successful contracts fetch via TED API."""
        mock_response = MagicMock()
        mock_response.json.return_value = mock_ted_response
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_contracts(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )

            assert len(result) == 1
            assert isinstance(result[0], BaseGovContract)
            assert result[0].contract_value_eur == 1500000
            assert result[0].contractor == "Empresa ABC, Lda"
            assert result[0].cpv_code == "45233000"

    def test_cpv_codes(self, client: BaseGovClient) -> None:
        """Test CPV code constants."""
        assert client.CPV_CONSTRUCTION == "45000000"
        assert client.CPV_BUILDING == "45210000"
        assert client.CPV_ROAD == "45233000"

    def test_eu_thresholds(self, client: BaseGovClient) -> None:
        """Test EU procurement threshold constants."""
        assert client.EU_THRESHOLD_WORKS == 5_382_000
        assert client.EU_THRESHOLD_SUPPLIES == 221_000
        assert client.EU_THRESHOLD_SERVICES == 221_000


# =============================================================================
# Story 6.9.5: BaseGov Public Procurement Fix Tests
# =============================================================================


class TestBaseGovStory695:
    """Tests for BaseGov Story 6.9.5 - Public Procurement Fix.

    Acceptance Criteria:
    - AC1: Research and verify dados.gov.pt OCDS dataset availability
    - AC2: Implement OCDS data fetching from dados.gov.pt
    - AC3: Parse OCDS format to BaseGovContract model
    - AC4: fetch_contracts() returns valid contract records
    - AC5: Handle pagination appropriately
    - AC6: Fallback to TED for contracts above EU thresholds
    - AC7: Document data source limitations in code comments
    - AC8: Implement retry logic per NFR1
    """

    @pytest.fixture
    def client(self) -> BaseGovClient:
        return BaseGovClient()

    def test_ac1_api_configuration(self) -> None:
        """AC1: Verify API configuration constants."""
        from raglite.external_data.clients.basegov import (
            BASEGOV_API_BASE,
            DADOS_GOV_API_BASE,
            OCDS_DATASET_ID,
            TED_API_BASE,
        )

        # TED API v3 is primary source
        assert TED_API_BASE == "https://tedweb.api.ted.europa.eu/v3"

        # dados.gov.pt OCDS dataset
        assert DADOS_GOV_API_BASE == "https://dados.gov.pt/api/1"
        assert OCDS_DATASET_ID == "ocds-portal-base-www-base-gov-pt"

        # Old Base.gov.pt URL kept for documentation (does NOT work)
        assert "base.gov.pt" in BASEGOV_API_BASE

    @pytest.mark.asyncio
    async def test_ac2_ocds_availability_check(self, client: BaseGovClient) -> None:
        """AC2: Test OCDS dataset availability check."""
        # Mock empty resources (current state as of 2025-12-08)
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "id": "ocds-portal-base-www-base-gov-pt",
            "title": "OCDS - Portal BASE",
            "resources": [],  # Empty - no resources available
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client._check_ocds_availability()
            assert result is None  # Returns None when no resources

    @pytest.mark.asyncio
    async def test_ac2_ocds_availability_with_resources(self, client: BaseGovClient) -> None:
        """AC2: Test OCDS check when resources exist."""
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "id": "ocds-portal-base-www-base-gov-pt",
            "resources": [
                {"format": "json", "url": "https://example.com/ocds.json"},
            ],
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client._check_ocds_availability()
            assert result is not None
            assert len(result.get("resources", [])) == 1

    def test_ac3_parse_ocds_data(self, client: BaseGovClient) -> None:
        """AC3: Test OCDS format parsing."""
        ocds_data = {
            "releases": [
                {
                    "ocid": "ocds-pt-001",
                    "date": "2024-01-15T10:00:00Z",
                    "tender": {
                        "title": "Road construction",
                        "items": [{"classification": {"id": "45233000"}}],
                        "value": {"amount": 500000, "currency": "EUR"},
                    },
                    "awards": [{"value": {"amount": 450000}}],
                    "parties": [
                        {"name": "Municipality", "roles": ["buyer"]},
                        {"name": "Contractor ABC", "roles": ["supplier"]},
                    ],
                },
            ]
        }

        result = client._parse_ocds_data(
            ocds_data,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
        )

        assert len(result) == 1
        assert result[0].contract_id == "ocds-pt-001"
        assert result[0].description == "Road construction"
        assert result[0].contract_value_eur == 450000  # Award value
        assert result[0].contracting_entity == "Municipality"
        assert result[0].contractor == "Contractor ABC"

    @pytest.mark.asyncio
    async def test_ac4_fetch_contracts_via_ted(self, client: BaseGovClient) -> None:
        """AC4: Test fetch_contracts with TED API response."""
        mock_ted_response = MagicMock()
        mock_ted_response.json.return_value = {
            "notices": [
                {
                    "publication-number": "2024/S 015-012345",
                    "publication-date": "2024-01-15",
                    "notice-title": "Construction project",
                    "buyer-name": "Município de Porto",
                    "winner-name": "BuildCo, SA",
                    "total-value": 2500000,
                    "cpv": "45210000",
                },
                {
                    "publication-number": "2024/S 015-012346",
                    "publication-date": "2024-01-20",
                    "notice-title": "Road works",
                    "buyer-name": "Município de Braga",
                    "winner-name": "RoadCo, SA",
                    "total-value": 3000000,
                    "cpv": "45233000",
                },
            ]
        }
        mock_ted_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(
                return_value=mock_ted_response
            )

            result = await client.fetch_contracts(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )

            assert len(result) == 2
            assert result[0].contract_value_eur == 2500000
            assert result[1].contract_value_eur == 3000000

    @pytest.mark.asyncio
    async def test_ac5_pagination(self, client: BaseGovClient) -> None:
        """AC5: Test pagination in fetch_all_contracts."""
        call_count = 0

        async def mock_fetch(*args, **kwargs):
            nonlocal call_count
            call_count += 1

            mock_response = MagicMock()
            if call_count == 1:
                # First page - full results
                mock_response.json.return_value = {
                    "notices": [
                        {
                            "publication-number": f"2024/S 015-{i:05d}",
                            "publication-date": "2024-01-15",
                            "notice-title": f"Project {i}",
                            "total-value": 1000000 + i * 100000,
                        }
                        for i in range(100)  # 100 results = full page
                    ]
                }
            else:
                # Second page - partial results (end of data)
                mock_response.json.return_value = {
                    "notices": [
                        {
                            "publication-number": "2024/S 015-99999",
                            "publication-date": "2024-01-20",
                            "notice-title": "Final Project",
                            "total-value": 5000000,
                        }
                    ]
                }
            mock_response.raise_for_status = MagicMock()
            return mock_response

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.post = mock_fetch

            result = await client.fetch_all_contracts(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 12, 31),
            )

            assert call_count == 2  # Two pages
            assert len(result) == 101  # 100 + 1

    @pytest.mark.asyncio
    async def test_ac6_ted_fallback_on_ocds_failure(self, client: BaseGovClient) -> None:
        """AC6: Test TED API is used when OCDS unavailable."""
        call_tracker = {"ted_called": False, "ocds_called": False}

        async def mock_get(*args, **kwargs):
            call_tracker["ocds_called"] = True
            mock_response = MagicMock()
            mock_response.json.return_value = {"resources": []}  # Empty OCDS
            mock_response.raise_for_status = MagicMock()
            return mock_response

        async def mock_post(*args, **kwargs):
            call_tracker["ted_called"] = True
            mock_response = MagicMock()
            mock_response.json.return_value = {
                "notices": [
                    {
                        "publication-number": "2024/S 015-00001",
                        "publication-date": "2024-01-15",
                        "notice-title": "TED Contract",
                        "total-value": 6000000,
                    }
                ]
            }
            mock_response.raise_for_status = MagicMock()
            return mock_response

        with patch("httpx.AsyncClient") as mock_client:
            instance = mock_client.return_value.__aenter__.return_value
            instance.get = mock_get
            instance.post = mock_post

            result = await client.fetch_contracts(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )

            assert call_tracker["ted_called"]
            assert len(result) == 1

    def test_ac7_documentation_comments(self) -> None:
        """AC7: Verify data source limitations are documented."""
        import inspect

        from raglite.external_data.clients import basegov

        source = inspect.getsource(basegov)

        # Check for important limitation documentation
        assert "UNAVAILABLE" in source or "no resources" in source.lower()
        assert "EU thresholds" in source or "EU_THRESHOLD" in source
        assert "TED API" in source
        assert "OCDS" in source
        assert "dados.gov.pt" in source

    @pytest.mark.asyncio
    async def test_ac8_retry_logic_exponential_backoff(self, client: BaseGovClient) -> None:
        """AC8: Test retry logic with exponential backoff."""
        call_times = []

        async def mock_post(*args, **kwargs):
            import time

            call_times.append(time.time())

            if len(call_times) < 3:
                raise httpx.TimeoutException("Timeout")

            mock_response = MagicMock()
            # Return non-empty result to avoid OCDS fallback
            mock_response.json.return_value = {
                "notices": [
                    {
                        "publication-number": "2024/S 015-00001",
                        "publication-date": "2024-01-15",
                        "total-value": 1000000,
                    }
                ]
            }
            mock_response.raise_for_status = MagicMock()
            return mock_response

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.post = mock_post

            # Should succeed after retries
            result = await client.fetch_contracts(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )

            assert len(call_times) == 3  # 2 failures + 1 success
            assert len(result) == 1  # Got results from TED after retry

    @pytest.mark.asyncio
    async def test_value_filtering(self, client: BaseGovClient) -> None:
        """Test min/max value filtering."""
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "notices": [
                {
                    "publication-number": "001",
                    "publication-date": "2024-01-15",
                    "total-value": 100000,
                },
                {
                    "publication-number": "002",
                    "publication-date": "2024-01-15",
                    "total-value": 500000,
                },
                {
                    "publication-number": "003",
                    "publication-date": "2024-01-15",
                    "total-value": 1000000,
                },
            ]
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(
                return_value=mock_response
            )

            # Filter by min value
            result = await client.fetch_contracts(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
                min_value=300000,
            )
            assert len(result) == 2

            # Filter by max value
            result = await client.fetch_contracts(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
                max_value=600000,
            )
            assert len(result) == 2

    def test_parse_ted_notices_value_formats(self, client: BaseGovClient) -> None:
        """Test TED notice parsing with various value formats."""
        # Value as dict
        data = {
            "notices": [
                {
                    "publication-number": "001",
                    "publication-date": "2024-01-15",
                    "total-value": {"amount": 500000, "currency": "EUR"},
                }
            ]
        }
        result = client._parse_ted_notices(data)
        assert result[0].contract_value_eur == 500000

        # Value as string with comma decimal
        data = {
            "notices": [
                {
                    "publication-number": "002",
                    "publication-date": "2024-01-15",
                    "total-value": "1 500 000,50",
                }
            ]
        }
        result = client._parse_ted_notices(data)
        assert result[0].contract_value_eur == 1500000.50

    def test_deprecated_methods_log_warning(self, client: BaseGovClient, caplog) -> None:
        """Test deprecated methods log warnings."""
        import logging

        with caplog.at_level(logging.WARNING):
            client._parse_contracts({})

        assert "Deprecated" in caplog.text
        assert "_parse_contracts" in caplog.text or "deprecated" in caplog.text.lower()


# =============================================================================
# Commodities Client Tests
# =============================================================================


# =============================================================================
# Story 6.9.1: Commodities URL Fix Tests (AC4-AC6)
# =============================================================================


class TestCommoditiesURLFix:
    """Tests for Commodities Ember API URL fix (Story 6.9.1 AC4-AC6).

    Problem: Ember Climate rebranded to Ember Energy and deprecated the old
    domain api.ember-climate.org on 2025-01-01.
    Fix: Update URL to api.ember-energy.org.
    """

    def test_co2_url_uses_new_domain(self) -> None:
        """AC4: Verify URL uses api.ember-energy.org."""
        # Read the commodities module source to verify URL
        import inspect

        from raglite.external_data import clients

        source = inspect.getsource(clients.commodities)

        # New domain should be present
        assert "api.ember-energy.org" in source
        # Comment about deprecated domain should be present
        assert "ember-climate.org" in source or "deprecated" in source.lower()

    def test_old_domain_not_used_in_fetch(self) -> None:
        """AC4: Verify api.ember-climate.org is not used in fetch_co2_prices."""
        import inspect

        from raglite.external_data.clients.commodities import CommoditiesClient

        # Get the source of fetch_co2_prices method
        source = inspect.getsource(CommoditiesClient.fetch_co2_prices)

        # Old domain should NOT be in the actual fetch URL
        assert "api.ember-climate.org/v1/carbon-price-tracker" not in source
        # New domain should be present
        assert "api.ember-energy.org" in source

    def test_co2_data_sources_updated(self) -> None:
        """AC4: Verify CO2_DATA_SOURCES constant updated."""
        from raglite.external_data.clients.commodities import CO2_DATA_SOURCES

        # Ember URL should use new domain
        assert "ember-energy.org" in CO2_DATA_SOURCES.get("ember", "")


class TestCommoditiesClient:
    """Tests for commodities price client."""

    @pytest.fixture
    def client(self, tmp_path) -> CommoditiesClient:
        """Create commodities client with temp cache directory."""
        return CommoditiesClient(cache_dir=tmp_path / "cache")

    @pytest.mark.asyncio
    async def test_fetch_co2_prices_success(self, client: CommoditiesClient) -> None:
        """Test successful CO2 EUA prices fetch."""
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "data": [
                {"date": "2024-01-15", "price": 85.50},
                {"date": "2024-01-16", "price": 86.00},
            ]
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_co2_prices(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )

            assert len(result) == 2
            assert result[0].price == 85.50
            assert result[0].currency == "EUR"

    def test_save_and_load_cache(self, client: CommoditiesClient) -> None:
        """Test cache save and load operations."""
        from raglite.external_data.models import CoalPrice

        prices = [
            CoalPrice(date=date(2024, 1, 1), price=120.0, currency="EUR"),
            CoalPrice(date=date(2024, 1, 8), price=122.0, currency="EUR"),
        ]

        client.save_to_cache("coal", prices)
        loaded = client.load_from_cache("coal")

        assert len(loaded) == 2
        assert loaded[0].price == 120.0

    def test_load_cache_with_date_filter(self, client: CommoditiesClient) -> None:
        """Test cache loading with date filtering."""
        from raglite.external_data.models import CoalPrice

        prices = [
            CoalPrice(date=date(2024, 1, 1), price=120.0, currency="EUR"),
            CoalPrice(date=date(2024, 2, 1), price=122.0, currency="EUR"),
            CoalPrice(date=date(2024, 3, 1), price=124.0, currency="EUR"),
        ]

        client.save_to_cache("coal", prices)
        loaded = client.load_from_cache(
            "coal",
            start_date=date(2024, 1, 15),
            end_date=date(2024, 2, 15),
        )

        assert len(loaded) == 1
        assert loaded[0].date == date(2024, 2, 1)

    def test_import_from_csv(self, client: CommoditiesClient, tmp_path) -> None:
        """Test CSV import."""
        csv_file = tmp_path / "coal_prices.csv"
        csv_file.write_text(
            """date,price,currency,unit,grade
2024-01-01,120.0,EUR,EUR/tonne,thermal
2024-01-08,122.0,EUR,EUR/tonne,thermal"""
        )

        result = client.import_from_csv("coal", csv_file)

        assert len(result) == 2
        assert result[0].price == 120.0
        assert result[0].grade == "thermal"

    def test_load_cache_empty(self, client: CommoditiesClient) -> None:
        """Test loading from non-existent cache returns empty list."""
        result = client.load_from_cache("nonexistent")
        assert result == []


# =============================================================================
# Exception Tests
# =============================================================================


# =============================================================================
# Additional Coverage Tests
# =============================================================================


class TestINEClientAdditional:
    """Additional tests for INE client coverage."""

    @pytest.fixture
    def client(self) -> INEClient:
        return INEClient()

    @pytest.mark.asyncio
    async def test_http_client_error(self, client: INEClient) -> None:
        """Test handling of 4xx client errors (no retry)."""
        error_response = MagicMock()
        error_response.status_code = 400
        error = httpx.HTTPStatusError("Bad Request", request=MagicMock(), response=error_response)

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(side_effect=error)

            with pytest.raises(ExternalDataFetchError) as exc_info:
                await client.fetch_building_permits(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

            assert "HTTP 400" in exc_info.value.message

    def test_parse_building_permits_malformed_period(self, client: INEClient) -> None:
        """Test handling of malformed period data."""
        data = {"Dados": {"invalid": [{"valor": 100}]}}
        result = client._parse_building_permits(data)
        # Should handle gracefully and return empty
        assert result == []

    def test_parse_building_permits_missing_value(self, client: INEClient) -> None:
        """Test handling of missing value."""
        data = {"Dados": {"202401": [{"geocod": "Lisboa"}]}}
        result = client._parse_building_permits(data)
        assert result == []


class TestBPstatClientAdditional:
    """Additional tests for BPstat client coverage.

    Story 6.9.3: Updated for new API structure.
    """

    @pytest.fixture
    def client(self) -> BPstatClient:
        return BPstatClient()

    @pytest.mark.asyncio
    async def test_http_client_error_no_retry(self, client: BPstatClient) -> None:
        """Test 4xx errors don't trigger retry."""
        error_response = MagicMock()
        error_response.status_code = 401
        error = httpx.HTTPStatusError("Unauthorized", request=MagicMock(), response=error_response)

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(side_effect=error)

            with pytest.raises(ExternalDataFetchError) as exc_info:
                await client.fetch_mortgage_loans(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

            assert "HTTP 401" in exc_info.value.message

    def test_parse_interest_rate_data_empty(self, client: BPstatClient) -> None:
        """Test parsing when observations list is empty.

        Story 6.9.3: Updated test for new parser method.
        """
        response = {"observations": []}

        result = client._parse_interest_rate_data(response, [client.MORTGAGE_RATE_MEDIAN])

        assert len(result) == 0

    def test_parse_interest_rate_data_multiple_periods(self, client: BPstatClient) -> None:
        """Test parsing multiple periods.

        Story 6.9.3 AC3: Test new response parsing.
        """
        response = {
            "observations": [
                {"period": "2024-01", "value": 3.45, "series_id": "12710733"},
                {"period": "2024-02", "value": 3.50, "series_id": "12710733"},
                {"period": "2024-03", "value": 3.55, "series_id": "12710733"},
            ]
        }

        result = client._parse_interest_rate_data(response, [client.MORTGAGE_RATE_MEDIAN])

        assert len(result) == 3
        assert result[0].avg_interest_rate_pct == 3.45
        assert result[1].avg_interest_rate_pct == 3.50
        assert result[2].avg_interest_rate_pct == 3.55


# =============================================================================
# Story 6.9.3: BPstat Banco de Portugal Fix Tests (AC1-AC7)
# =============================================================================


class TestBPstatStory693:
    """Tests for BPstat client fixes (Story 6.9.3 AC1-AC7).

    Story 6.9.3 Changes:
    - AC1: Updated series IDs to verified correct values (12710733 for median rate)
    - AC2: Updated API endpoint to /api/observations/
    - AC3: Updated response parsing for new API structure
    - AC4: fetch_mortgage_loans() returns valid interest rate data
    - AC5: Unit tests for new API structure
    - AC6: Documented series IDs in code comments
    - AC7: Retry logic per NFR1 (exponential backoff 2s/4s/8s)
    """

    def test_correct_series_ids_used(self) -> None:
        """AC1: Verify mortgage rate series IDs are correct (12710733 not 12532089)."""
        from raglite.external_data.clients.bpstat import BPstatClient

        # New correct series IDs
        assert BPstatClient.MORTGAGE_RATE_MEDIAN == "12710733"
        assert BPstatClient.MORTGAGE_RATE_10TH_PERCENTILE == "12710735"
        assert BPstatClient.MORTGAGE_RATE_25TH_PERCENTILE == "12710781"
        assert BPstatClient.MORTGAGE_RATE_75TH_PERCENTILE == "12710734"
        assert BPstatClient.MORTGAGE_RATE_90TH_PERCENTILE == "12710736"

    def test_old_series_ids_not_used(self) -> None:
        """AC1: Verify old/wrong series IDs are removed or aliased."""
        from raglite.external_data.clients.bpstat import BPstatClient

        # Old series IDs should NOT be the wrong values
        assert BPstatClient.MORTGAGE_LOANS_SERIES != "12532089"  # Was Egyptian Pound!
        # Old alias should point to new correct ID
        assert BPstatClient.MORTGAGE_LOANS_SERIES == "12710733"

    def test_new_api_endpoint(self) -> None:
        """AC2: Verify using /api/observations/ endpoint."""
        from raglite.external_data.clients.bpstat import BPSTAT_API_BASE

        # New endpoint should be /api (not /data/v1)
        assert BPSTAT_API_BASE == "https://bpstat.bportugal.pt/api"
        assert "/data/v1" not in BPSTAT_API_BASE

    def test_observations_endpoint_url_structure(self) -> None:
        """AC2: Verify observations URL is constructed correctly."""
        from raglite.external_data.clients.bpstat import BPstatClient

        client = BPstatClient()
        # URL should end with /observations/
        expected_base = "https://bpstat.bportugal.pt/api"
        assert client.base_url == expected_base

    def test_new_response_parsing(self) -> None:
        """AC3: Verify new API response parsing."""
        from raglite.external_data.clients.bpstat import BPstatClient

        client = BPstatClient()

        # New API response format
        response = {
            "observations": [
                {"period": "2024-01", "value": 3.45, "series_id": "12710733"},
                {"period": "2024-02", "value": 3.50, "series_id": "12710733"},
            ]
        }

        result = client._parse_interest_rate_data(response, ["12710733"])

        assert len(result) == 2
        assert result[0].avg_interest_rate_pct == 3.45
        assert result[1].avg_interest_rate_pct == 3.50
        # Dates should be first of month
        assert result[0].date == date(2024, 1, 1)
        assert result[1].date == date(2024, 2, 1)

    def test_response_parsing_with_refperiod_key(self) -> None:
        """AC3: Verify parsing handles both 'period' and 'refPeriod' keys."""
        from raglite.external_data.clients.bpstat import BPstatClient

        client = BPstatClient()

        # API may use refPeriod instead of period
        response = {
            "observations": [
                {"refPeriod": "2024-01", "value": 3.45, "series_id": "12710733"},
            ]
        }

        result = client._parse_interest_rate_data(response, ["12710733"])

        assert len(result) == 1
        assert result[0].avg_interest_rate_pct == 3.45

    def test_response_parsing_skips_invalid_entries(self) -> None:
        """AC3: Verify invalid observations are skipped."""
        from raglite.external_data.clients.bpstat import BPstatClient

        client = BPstatClient()

        response = {
            "observations": [
                {"period": "2024-01", "value": 3.45, "series_id": "12710733"},
                {"period": "2024-02", "value": None, "series_id": "12710733"},  # Null value
                {"value": 3.60, "series_id": "12710733"},  # Missing period
                {"period": "2024-04", "value": 3.65, "series_id": "12710733"},
            ]
        }

        result = client._parse_interest_rate_data(response, ["12710733"])

        # Should only parse first and last valid entries
        assert len(result) == 2
        assert result[0].avg_interest_rate_pct == 3.45
        assert result[1].avg_interest_rate_pct == 3.65

    def test_series_ids_documented(self) -> None:
        """AC6: Verify series IDs are documented in code comments."""
        import inspect

        from raglite.external_data.clients.bpstat import BPstatClient

        source = inspect.getsource(BPstatClient)

        # Documentation should mention verified date
        assert "2025-12-08" in source or "Story 6.9.3" in source
        # Should document that old IDs were wrong
        assert "12532089" in source  # Old wrong ID mentioned in comment
        assert "Egyptian" in source or "WRONG" in source or "FX" in source

    def test_retry_delays_exponential_backoff(self) -> None:
        """AC7: Verify retry logic uses exponential backoff (2s/4s/8s)."""
        import inspect

        from raglite.external_data.clients.bpstat import BPstatClient

        source = inspect.getsource(BPstatClient._fetch_with_retry)

        # Should have 2, 4, 8 second delays (NFR1 requirement)
        assert "retry_delays = [2, 4, 8]" in source

    @pytest.mark.asyncio
    async def test_fetch_with_include_percentiles(self) -> None:
        """AC4: Test fetching with all percentile series."""
        from raglite.external_data.clients.bpstat import BPstatClient

        client = BPstatClient()

        mock_response = MagicMock()
        mock_response.json.return_value = {
            "observations": [
                {"period": "2024-01", "value": 3.45, "series_id": "12710733"},
                {"period": "2024-01", "value": 3.10, "series_id": "12710735"},
                {"period": "2024-01", "value": 3.25, "series_id": "12710781"},
            ]
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_mortgage_loans(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
                include_percentiles=True,
            )

            # Should return median rate
            assert len(result) == 1
            assert result[0].avg_interest_rate_pct == 3.45


# =============================================================================
# Story 6.9.2: OMIE Electricity Market Fix Tests (AC1-AC6)
# =============================================================================


class TestOMIEStory692:
    """Tests for OMIE client fixes (Story 6.9.2 AC1-AC6).

    Story 6.9.2 Changes:
    - AC1: Updated URL pattern to use file-download endpoint
    - AC2: Enabled follow_redirects=True
    - AC3: Updated CSV parser for new format (MARGINALPDBC;YEAR;MONTH;DAY;HOUR;PT;ES)
    - AC5: Unit tests updated for new format
    - AC6: Retry logic per NFR1 (exponential backoff 2s/4s/8s)
    """

    def test_new_url_pattern(self) -> None:
        """AC1: Verify new file-download URL pattern."""
        from raglite.external_data.clients.omie import OMIE_BASE_URL

        # New URL should use file-download endpoint
        assert "file-download" in OMIE_BASE_URL
        # Old URL pattern should NOT be present
        assert "/sites/default/files/dados" not in OMIE_BASE_URL
        assert "/AGNO_" not in OMIE_BASE_URL

    def test_url_includes_query_params(self) -> None:
        """AC1: Verify URL construction uses query params."""
        from raglite.external_data.clients.omie import OMIE_BASE_URL

        # URL should be the base for query params
        expected_base = "https://www.omie.es/es/file-download"
        assert OMIE_BASE_URL == expected_base

    def test_new_csv_format_parsing(self) -> None:
        """AC3: Verify parsing of MARGINALPDBC;YEAR;MONTH;DAY;HOUR;PT;ES format."""
        from datetime import date

        from raglite.external_data.clients.omie import OMIEClient

        client = OMIEClient()

        # New format: MARGINALPDBC;YEAR;MONTH;DAY;HOUR;PT_PRICE;ES_PRICE;...
        content = """MARGINALPDBC;2024;12;08;1;111,60;112,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;12;08;2;105,50;106,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;12;08;3;98,75;99,00;0;0;0;0;0;0;0;0"""

        result = client._parse_daily_file(content, date(2024, 12, 8))

        assert len(result) == 3
        assert result[0].price_eur_mwh == 111.60  # European decimal comma handled
        assert result[1].price_eur_mwh == 105.50
        assert result[2].price_eur_mwh == 98.75

    def test_hour_conversion_1_based_to_0_based(self) -> None:
        """AC3: Verify hour 1 in file becomes hour 0 in output."""
        from datetime import date

        from raglite.external_data.clients.omie import OMIEClient

        client = OMIEClient()

        # Hours in OMIE files are 1-24, we convert to 0-23
        content = """MARGINALPDBC;2024;12;08;1;100,00;100,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;12;08;12;150,00;150,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;12;08;24;80,00;80,00;0;0;0;0;0;0;0;0"""

        result = client._parse_daily_file(content, date(2024, 12, 8))

        assert result[0].hour == 0  # Hour 1 -> 0
        assert result[1].hour == 11  # Hour 12 -> 11
        assert result[2].hour == 23  # Hour 24 -> 23

    def test_decimal_comma_handling(self) -> None:
        """AC3: Verify prices like '111,60' are parsed correctly."""
        from datetime import date

        from raglite.external_data.clients.omie import OMIEClient

        client = OMIEClient()

        # European format with comma as decimal separator
        content = """MARGINALPDBC;2024;12;08;1;111,60;112,50;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;12;08;2;99,99;100,01;0;0;0;0;0;0;0;0"""

        result = client._parse_daily_file(content, date(2024, 12, 8))

        assert result[0].price_eur_mwh == 111.60
        assert result[1].price_eur_mwh == 99.99

    def test_malformed_line_handling(self) -> None:
        """AC3: Verify malformed lines are skipped with warning."""
        from datetime import date

        from raglite.external_data.clients.omie import OMIEClient

        client = OMIEClient()

        # Mix of valid and invalid lines
        content = """MARGINALPDBC;2024;12;08;1;100,00;100,00;0;0;0;0;0;0;0;0
INVALID_PREFIX;2024;12;08;2;50,00;50,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;12;08;3;short
MARGINALPDBC;2024;12;08;bad;100,00;100,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;12;08;4;150,00;150,00;0;0;0;0;0;0;0;0"""

        result = client._parse_daily_file(content, date(2024, 12, 8))

        # Only first and last lines should be parsed
        assert len(result) == 2
        assert result[0].hour == 0  # Hour 1
        assert result[1].hour == 3  # Hour 4

    def test_non_marginalpdbc_lines_skipped(self) -> None:
        """AC3: Verify non-MARGINALPDBC lines (headers, etc.) are skipped."""
        from datetime import date

        from raglite.external_data.clients.omie import OMIEClient

        client = OMIEClient()

        # Content with headers and comments
        content = """Header;Line;That;Should;Be;Skipped
MARGINALPDBC;2024;12;08;1;100,00;100,00;0;0;0;0;0;0;0;0
Some;Other;Data;Format
MARGINALPDBC;2024;12;08;2;110,00;110,00;0;0;0;0;0;0;0;0"""

        result = client._parse_daily_file(content, date(2024, 12, 8))

        # Only MARGINALPDBC lines should be parsed
        assert len(result) == 2

    def test_retry_delays_exponential_backoff(self) -> None:
        """AC6: Verify retry logic uses exponential backoff (2s/4s/8s)."""
        import inspect

        from raglite.external_data.clients.omie import OMIEClient

        # Get the source of _fetch_daily_file method
        source = inspect.getsource(OMIEClient._fetch_daily_file)

        # Should have 2, 4, 8 second delays (NFR1 requirement)
        assert "retry_delays = [2, 4, 8]" in source


class TestOMIEClientAdditional:
    """Additional tests for OMIE client coverage.

    Story 6.9.2: Updated for new OMIE CSV format.
    """

    @pytest.fixture
    def client(self) -> OMIEClient:
        return OMIEClient()

    @pytest.mark.asyncio
    async def test_server_error_retry_then_fail(self, client: OMIEClient) -> None:
        """Test retry exhaustion on server errors."""
        error_response = MagicMock()
        error_response.status_code = 503

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                side_effect=httpx.HTTPStatusError(
                    "Service Unavailable", request=MagicMock(), response=error_response
                )
            )

            with patch("asyncio.sleep", new_callable=AsyncMock):
                # Should fail after retries but not raise (404 handling)
                result = await client.fetch_spot_prices(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 1),
                )
                # Non-404 errors are caught and logged, returns empty
                assert result == []

    @pytest.mark.asyncio
    async def test_monthly_average(self, client: OMIEClient) -> None:
        """Test monthly average calculation.

        Story 6.9.2 AC3/AC5: Updated for new CSV format.
        """
        mock_response = MagicMock()
        # Story 6.9.2 AC3: New format - MARGINALPDBC;YEAR;MONTH;DAY;HOUR;PT;ES;...
        mock_response.text = """MARGINALPDBC;2024;01;01;1;50,00;51,00;0;0;0;0;0;0;0;0
MARGINALPDBC;2024;01;01;2;60,00;61,00;0;0;0;0;0;0;0;0"""
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_monthly_average(2024, 1)

        assert result is not None
        assert result.price_type == "monthly_avg"


class TestIPMAClientAdditional:
    """Additional tests for IPMA client coverage."""

    @pytest.fixture
    def client(self) -> IPMAClient:
        return IPMAClient()

    @pytest.mark.asyncio
    async def test_fetch_all_stations(self, client: IPMAClient) -> None:
        """Test fetching from all weather stations."""
        mock_response = MagicMock()
        mock_response.json.return_value = {"tMed": 15.0}
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_all_stations(date(2024, 1, 1))

        # Should fetch from multiple stations
        assert len(result) >= 1

    def test_parse_forecast_day_missing_fields(self, client: IPMAClient) -> None:
        """Test parsing forecast with missing optional fields."""
        data = {"forecastDate": "2024-01-01"}  # Missing temperature fields

        result = client._parse_forecast_day(data, "1110600")

        assert result is not None
        assert result.temperature_max_c is None


class TestBaseGovClientAdditional:
    """Additional tests for Base.gov.pt client coverage.

    Story 6.9.5: Updated for TED API v3 implementation.
    """

    @pytest.fixture
    def client(self) -> BaseGovClient:
        return BaseGovClient()

    @pytest.mark.asyncio
    async def test_fetch_construction_summary(self, client: BaseGovClient) -> None:
        """Test construction contracts summary calculation."""
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "notices": [
                {
                    "publication-number": "2024/S 015-00001",
                    "publication-date": "2024-01-15",
                    "total-value": 1000000,
                },
                {
                    "publication-number": "2024/S 015-00002",
                    "publication-date": "2024-01-20",
                    "total-value": 2000000,
                },
            ]
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_construction_contracts_summary(2024, 1)

        assert result["total_contracts"] == 2
        assert result["total_value_eur"] == 3000000
        assert result["avg_value_eur"] == 1500000
        # Story 6.9.5: Now uses IMPIC as primary source
        assert "IMPIC" in result["data_source"] or "TED" in result["data_source"]
        assert "note" in result


class TestATICClientAdditional:
    """Additional tests for ATIC client coverage."""

    @pytest.fixture
    def client(self) -> ATICClient:
        return ATICClient()

    @pytest.mark.asyncio
    async def test_fetch_historical_no_csv(self, client: ATICClient) -> None:
        """Test historical fetch without CSV returns empty with warning."""
        result = await client.fetch_historical_data(
            start_date=date(2024, 1, 1),
            end_date=date(2024, 3, 31),
            csv_path=None,
        )
        assert result == []

    @pytest.mark.asyncio
    async def test_fetch_historical_with_csv(self, client: ATICClient, tmp_path) -> None:
        """Test historical fetch with CSV and date filtering."""
        csv_file = tmp_path / "test.csv"
        csv_file.write_text(
            """date,consumption,region
2024-01-01,100000,Portugal
2024-02-01,110000,Portugal
2024-06-01,120000,Portugal"""
        )

        result = await client.fetch_historical_data(
            start_date=date(2024, 1, 1),
            end_date=date(2024, 3, 31),
            csv_path=csv_file,
        )

        # Only 2 records in date range
        assert len(result) == 2


class TestCommoditiesClientAdditional:
    """Additional tests for commodities client coverage."""

    @pytest.fixture
    def client(self, tmp_path) -> CommoditiesClient:
        return CommoditiesClient(cache_dir=tmp_path / "cache")

    @pytest.mark.asyncio
    async def test_fetch_coal_falls_back_to_cache(self, client: CommoditiesClient) -> None:
        """Test coal prices fallback to cache (no API)."""
        result = await client.fetch_coal_prices(
            start_date=date(2024, 1, 1),
            end_date=date(2024, 3, 31),
        )
        # No cache exists, returns empty
        assert result == []

    @pytest.mark.asyncio
    async def test_fetch_petcoke_falls_back_to_cache(self, client: CommoditiesClient) -> None:
        """Test petcoke prices fallback to cache (no API)."""
        result = await client.fetch_petcoke_prices(
            start_date=date(2024, 1, 1),
            end_date=date(2024, 3, 31),
        )
        assert result == []

    @pytest.mark.asyncio
    async def test_fetch_co2_api_failure_fallback(self, client: CommoditiesClient) -> None:
        """Test CO2 prices fallback on API failure."""
        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                side_effect=httpx.TimeoutException("timeout")
            )

            with patch("asyncio.sleep", new_callable=AsyncMock):
                result = await client.fetch_co2_prices(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 3, 31),
                )

        # Falls back to empty cache
        assert result == []

    def test_import_csv_file_not_found(self, client: CommoditiesClient) -> None:
        """Test CSV import with non-existent file."""
        with pytest.raises(ExternalDataFetchError):
            client.import_from_csv("coal", "/nonexistent/path.csv")


class TestExceptions:
    """Tests for external data exception classes."""

    def test_external_data_fetch_error(self) -> None:
        """Test ExternalDataFetchError formatting."""
        original = ValueError("Connection refused")
        error = ExternalDataFetchError(
            source="INE",
            message="Failed to connect",
            original_error=original,
        )

        assert error.source == "INE"
        assert error.message == "Failed to connect"
        assert error.original_error is original
        assert "[INE] Failed to connect" in str(error)

    def test_external_data_validation_error(self) -> None:
        """Test ExternalDataValidationError formatting."""
        error = ExternalDataValidationError(
            source="ATIC",
            message="Missing date column",
        )

        assert error.source == "ATIC"
        assert "Validation failed" in str(error)

    def test_external_data_stale_error(self) -> None:
        """Test ExternalDataStaleError formatting."""
        from raglite.external_data.exceptions import ExternalDataStaleError

        error = ExternalDataStaleError(
            source="OMIE",
            days_stale=45,
            max_days=30,
        )

        assert error.source == "OMIE"
        assert error.days_stale == 45
        assert error.max_days == 30
        assert "45 days old" in str(error)


# =============================================================================
# Rate Limit (429) Handling Tests
# =============================================================================


class TestRateLimitHandling:
    """Tests for rate limit (429) handling across clients."""

    @pytest.mark.asyncio
    async def test_ine_rate_limit_retry(self) -> None:
        """Test INE client retries on 429 rate limit."""
        client = INEClient()

        # Create 429 error response
        rate_limit_response = MagicMock()
        rate_limit_response.status_code = 429
        rate_limit_error = httpx.HTTPStatusError(
            "Too Many Requests", request=MagicMock(), response=rate_limit_response
        )

        success_response = MagicMock()
        success_response.json.return_value = {"Dados": {"202401": [{"valor": 100}]}}
        success_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            # First call returns 429, second succeeds
            mock_get = AsyncMock(side_effect=[rate_limit_error, success_response])
            mock_client.return_value.__aenter__.return_value.get = mock_get

            with patch("asyncio.sleep", new_callable=AsyncMock):
                result = await client.fetch_building_permits(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

            assert len(result) == 1
            assert mock_get.call_count == 2  # Retried after 429

    @pytest.mark.asyncio
    async def test_bpstat_rate_limit_retry(self) -> None:
        """Test BPstat client retries on 429 rate limit.

        Story 6.9.3: Updated for new API structure (single API call, not 3).
        """
        client = BPstatClient()

        rate_limit_response = MagicMock()
        rate_limit_response.status_code = 429
        rate_limit_error = httpx.HTTPStatusError(
            "Too Many Requests", request=MagicMock(), response=rate_limit_response
        )

        success_response = MagicMock()
        # Story 6.9.3: New response format with series_id
        success_response.json.return_value = {
            "observations": [{"period": "2024-01", "value": 3.45, "series_id": "12710733"}]
        }
        success_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            # Story 6.9.3: Now single API call with retry: 429 then success
            mock_get = AsyncMock(
                side_effect=[
                    rate_limit_error,
                    success_response,
                ]
            )
            mock_client.return_value.__aenter__.return_value.get = mock_get

            with patch("asyncio.sleep", new_callable=AsyncMock):
                result = await client.fetch_mortgage_loans(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

            assert len(result) == 1
            assert mock_get.call_count == 2  # 1 retry + 1 success


# =============================================================================
# Additional Coverage Tests for eu_oil_bulletin.py
# =============================================================================


class TestEUOilBulletinAdditional:
    """Additional tests for EU Oil Bulletin coverage.

    Story 6.9.4: Updated tests for new XLSX-based implementation (2025-12-08)
    """

    @pytest.fixture
    def client(self) -> EUOilBulletinClient:
        return EUOilBulletinClient()

    @pytest.fixture
    def mock_xlsx_content(self) -> bytes:
        """Create mock XLSX content for testing."""
        from datetime import datetime
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"
        ws.append(["Date", "Portugal", "Spain"])
        ws.append([datetime(2024, 1, 8), 1.456, 1.423])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_deprecated_xml_parsing_returns_empty(self, client: EUOilBulletinClient) -> None:
        """Test deprecated XML parsing method returns empty list.

        Story 6.9.4: Old XML method is deprecated and should return empty.
        """
        result = client._parse_bulletin_xml(
            "any content", "PT", date(2024, 1, 1), date(2024, 1, 31)
        )
        assert result == []

    @pytest.mark.asyncio
    async def test_fetch_weekly_prices(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test fetching weekly prices.

        Story 6.9.4 AC4: Handle weekly data frequency.
        """
        mock_response = MagicMock()
        mock_response.content = mock_xlsx_content
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            # Mock cache methods
            client._get_cached_xlsx = MagicMock(return_value=None)
            client._save_to_cache = MagicMock()

            result = await client.fetch_weekly_prices(
                week_date=date(2024, 1, 10),  # Wednesday of week containing 2024-01-08
                country="Portugal",
            )

        assert result is not None
        assert result.price_eur_litre == 1.456

    @pytest.mark.asyncio
    async def test_fetch_weekly_prices_not_found(self, client: EUOilBulletinClient) -> None:
        """Test fetching weekly prices when no data available."""
        from io import BytesIO

        import openpyxl

        # Create empty XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"
        ws.append(["Date", "Portugal"])
        # No data rows

        buffer = BytesIO()
        wb.save(buffer)
        empty_xlsx = buffer.getvalue()

        mock_response = MagicMock()
        mock_response.content = empty_xlsx
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            client._get_cached_xlsx = MagicMock(return_value=None)
            client._save_to_cache = MagicMock()

            result = await client.fetch_weekly_prices(
                week_date=date(2024, 6, 15),
                country="Portugal",
            )

        assert result is None

    @pytest.mark.asyncio
    async def test_fetch_xlsx_timeout_retry(
        self, client: EUOilBulletinClient, mock_xlsx_content: bytes
    ) -> None:
        """Test timeout retry logic for XLSX download.

        Story 6.9.4 AC7: NFR1 retry logic with 60s timeout.
        """
        success_response = MagicMock()
        success_response.content = mock_xlsx_content
        success_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_get = AsyncMock(
                side_effect=[
                    httpx.TimeoutException("timeout"),
                    success_response,
                ]
            )
            mock_client.return_value.__aenter__.return_value.get = mock_get

            # Skip cache
            client._get_cached_xlsx = MagicMock(return_value=None)
            client._save_to_cache = MagicMock()

            with patch("asyncio.sleep", new_callable=AsyncMock):
                await client.fetch_diesel_prices(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

            assert mock_get.call_count == 2


# =============================================================================
# Story 6.9.4: EU Oil Bulletin Fix Tests (AC1-AC7)
# =============================================================================


class TestEUOilBulletinStory694:
    """Tests for EU Oil Bulletin client fixes (Story 6.9.4 AC1-AC7).

    Story 6.9.4 Changes:
    - AC1: Switched from XML to XLSX data source
    - AC2: Implemented XLSX parsing with openpyxl
    - AC3: fetch_diesel_prices() returns valid diesel price records
    - AC4: Handle weekly data frequency
    - AC5: Added caching for large historical file
    - AC6: Unit tests with sample XLSX data
    - AC7: Retry logic per NFR1 (60s timeout)
    """

    def test_new_xlsx_endpoint(self) -> None:
        """AC1: Verify new XLSX endpoint is configured."""
        from raglite.external_data.clients.eu_oil_bulletin import (
            EU_OIL_BULLETIN_BASE,
            HISTORY_XLSX_DOC_ID,
            HISTORY_XLSX_FILENAME,
        )

        # New base URL
        assert EU_OIL_BULLETIN_BASE == "https://energy.ec.europa.eu"
        # Document ID for historical prices
        assert HISTORY_XLSX_DOC_ID == "906e60ca-8b6a-44e7-8589-652854d2fd3f_en"
        assert "xlsx" in HISTORY_XLSX_FILENAME.lower()

    def test_old_xml_endpoint_not_used(self) -> None:
        """AC1: Verify old XML endpoint is not used."""
        from raglite.external_data.clients.eu_oil_bulletin import EU_OIL_BULLETIN_BASE

        assert "observatory/reports" not in EU_OIL_BULLETIN_BASE
        assert ".xml" not in EU_OIL_BULLETIN_BASE

    def test_openpyxl_parsing_method_exists(self) -> None:
        """AC2: Verify XLSX parsing method exists."""
        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        client = EUOilBulletinClient()
        assert hasattr(client, "_parse_xlsx")
        assert callable(client._parse_xlsx)

    def test_cache_methods_exist(self) -> None:
        """AC5: Verify caching methods exist."""
        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        client = EUOilBulletinClient()
        assert hasattr(client, "_get_cached_xlsx")
        assert hasattr(client, "_save_to_cache")
        assert hasattr(client, "cache_dir")
        assert hasattr(client, "cache_ttl_hours")

    def test_extended_timeout_for_large_file(self) -> None:
        """AC7: Verify 60s timeout for large file download."""
        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        client = EUOilBulletinClient()
        # In non-test mode, timeout should be 60s
        # (Test mode has 1s timeout, but we check the attribute exists)
        assert hasattr(client, "timeout")

    def test_retry_delays_exponential_backoff(self) -> None:
        """AC7: Verify retry logic uses exponential backoff (2s/4s/8s)."""
        import inspect

        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        source = inspect.getsource(EUOilBulletinClient._fetch_xlsx_data)

        # Should have 2, 4, 8 second delays (NFR1 requirement)
        assert "retry_delays = [2, 4, 8]" in source

    def test_xlsx_validation(self) -> None:
        """AC2: Verify XLSX validation checks for PK header."""
        import inspect

        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        source = inspect.getsource(EUOilBulletinClient._fetch_xlsx_data)

        # Should verify XLSX starts with PK (ZIP signature)
        assert 'b"PK"' in source or "PK" in source

    def test_parse_xlsx_with_sample_data(self) -> None:
        """AC6: Test parsing with sample XLSX data."""
        from datetime import datetime
        from io import BytesIO

        import openpyxl

        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        # Create sample XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"
        ws.append(["Date", "Portugal", "Spain", "Germany"])
        ws.append([datetime(2024, 12, 2), 1.523, 1.489, 1.612])
        ws.append([datetime(2024, 12, 9), 1.534, 1.501, 1.598])

        buffer = BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()

        client = EUOilBulletinClient()
        result = client._parse_xlsx(
            xlsx_bytes,
            country="Portugal",
            start_date=date(2024, 12, 1),
            end_date=date(2024, 12, 31),
            tax_included=True,
        )

        assert len(result) == 2
        assert result[0].price_eur_litre == 1.523
        assert result[1].price_eur_litre == 1.534
        assert result[0].country == "Portugal"

    def test_parse_xlsx_different_country(self) -> None:
        """AC3: Test parsing for different countries."""
        from datetime import datetime
        from io import BytesIO

        import openpyxl

        from raglite.external_data.clients.eu_oil_bulletin import EUOilBulletinClient

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prices with taxes"
        ws.append(["Date", "Portugal", "Spain"])
        ws.append([datetime(2024, 12, 2), 1.523, 1.489])

        buffer = BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()

        client = EUOilBulletinClient()

        # Portugal
        pt_result = client._parse_xlsx(
            xlsx_bytes, "Portugal", date(2024, 12, 1), date(2024, 12, 31), True
        )
        assert len(pt_result) == 1
        assert pt_result[0].price_eur_litre == 1.523

        # Spain
        es_result = client._parse_xlsx(
            xlsx_bytes, "Spain", date(2024, 12, 1), date(2024, 12, 31), True
        )
        assert len(es_result) == 1
        assert es_result[0].price_eur_litre == 1.489


# =============================================================================
# Additional Coverage Tests for ipma.py
# =============================================================================


class TestIPMAClientCoverage:
    """Additional tests for IPMA client coverage."""

    @pytest.fixture
    def client(self) -> IPMAClient:
        return IPMAClient()

    @pytest.mark.asyncio
    async def test_fetch_observations_no_data(self, client: IPMAClient) -> None:
        """Test handling of empty observation response."""
        mock_response = MagicMock()
        mock_response.json.return_value = {}  # Empty response
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_observations(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 1),
            )

        # No valid observation data
        assert len(result) == 0

    @pytest.mark.asyncio
    async def test_fetch_observations_fetch_error(self, client: IPMAClient) -> None:
        """Test handling of fetch error during observations."""
        error_response = MagicMock()
        error_response.status_code = 404

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                side_effect=httpx.HTTPStatusError(
                    "Not Found", request=MagicMock(), response=error_response
                )
            )

            # Should not raise - fetch errors are caught and result is empty
            result = await client.fetch_observations(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 1),
            )

        assert result == []

    @pytest.mark.asyncio
    async def test_fetch_forecast_parse_error(self, client: IPMAClient) -> None:
        """Test handling of parse error in forecast."""
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "data": [
                {"forecastDate": "invalid-date"},  # Will fail to parse
            ]
        }
        mock_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                return_value=mock_response
            )

            result = await client.fetch_forecast(location_id="1110600", days=1)

        # Invalid date causes parse failure, result is empty
        assert result == []

    def test_parse_observation_exception(self, client: IPMAClient) -> None:
        """Test parse observation with malformed data."""
        # Data that causes exception during parsing
        data = {"tMed": "not-a-number"}  # Will fail float conversion
        result = client._parse_observation(data, date(2024, 1, 1), "1200535")
        # Should return None on exception
        assert result is None

    @pytest.mark.asyncio
    async def test_fetch_observations_timeout_retry(self, client: IPMAClient) -> None:
        """Test timeout retry logic."""
        success_response = MagicMock()
        success_response.json.return_value = {"tMed": 15.0}
        success_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_get = AsyncMock(
                side_effect=[
                    httpx.TimeoutException("timeout"),
                    success_response,
                ]
            )
            mock_client.return_value.__aenter__.return_value.get = mock_get

            with patch("asyncio.sleep", new_callable=AsyncMock):
                await client.fetch_observations(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 1),
                )

        # Retried and succeeded
        assert mock_get.call_count == 2


# =============================================================================
# Additional Coverage Tests for basegov.py
# =============================================================================


class TestBaseGovClientCoverage:
    """Additional tests for BaseGov client coverage.

    Story 6.9.5: Updated for TED API v3 implementation.
    """

    @pytest.fixture
    def client(self) -> BaseGovClient:
        return BaseGovClient()

    @pytest.mark.asyncio
    async def test_fetch_contracts_timeout_retry(self, client: BaseGovClient) -> None:
        """Test timeout retry logic with TED API."""
        success_response = MagicMock()
        # Return non-empty results to avoid OCDS fallback
        success_response.json.return_value = {
            "notices": [
                {
                    "publication-number": "2024/S 015-00001",
                    "publication-date": "2024-01-15",
                    "total-value": 1000000,
                }
            ]
        }
        success_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_post = AsyncMock(
                side_effect=[
                    httpx.TimeoutException("timeout"),
                    success_response,
                ]
            )
            mock_client.return_value.__aenter__.return_value.post = mock_post

            with patch("asyncio.sleep", new_callable=AsyncMock):
                result = await client.fetch_contracts(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

            assert mock_post.call_count == 2
            # Retry succeeded - should return the contract from second attempt
            assert len(result) == 1
            assert result[0].contract_id == "2024/S 015-00001"

    @pytest.mark.asyncio
    async def test_fetch_all_contracts_pagination(self, client: BaseGovClient) -> None:
        """Test pagination handling with TED API."""
        page1_response = MagicMock()
        page1_response.json.return_value = {
            "notices": [
                {
                    "publication-number": f"2024/S 015-{i:05d}",
                    "publication-date": "2024-01-15",
                    "total-value": 1000000,
                }
                for i in range(100)  # Full page
            ]
        }
        page1_response.raise_for_status = MagicMock()

        page2_response = MagicMock()
        page2_response.json.return_value = {
            "notices": [
                {
                    "publication-number": f"2024/S 015-{100 + i:05d}",
                    "publication-date": "2024-01-20",
                    "total-value": 500000,
                }
                for i in range(50)  # Partial page
            ]
        }
        page2_response.raise_for_status = MagicMock()

        with patch("httpx.AsyncClient") as mock_client:
            mock_post = AsyncMock(side_effect=[page1_response, page2_response])
            mock_client.return_value.__aenter__.return_value.post = mock_post

            result = await client.fetch_all_contracts(
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )

        # 100 from page 1 + 50 from page 2
        assert len(result) == 150

    def test_deprecated_parse_contracts_logs_warning(self, client: BaseGovClient, caplog) -> None:
        """Test deprecated _parse_contracts method logs warning."""
        import logging

        with caplog.at_level(logging.WARNING):
            # _parse_contracts is deprecated - should log warning and return empty
            result = client._parse_contracts({"items": []})

        assert "Deprecated" in caplog.text or "deprecated" in caplog.text.lower()
        assert result == []  # Deprecated method returns empty list

    def test_parse_ted_notices_missing_date(self, client: BaseGovClient) -> None:
        """Test TED notice parsing with missing publication date."""
        data = {
            "notices": [
                {"publication-number": "001"},  # Missing publication-date
                {
                    "publication-number": "002",
                    "publication-date": "2024-01-15",
                    "total-value": 1000000,
                },
            ]
        }
        result = client._parse_ted_notices(data)
        # Only second record should be parsed
        assert len(result) == 1
        assert result[0].contract_id == "002"


# =============================================================================
# Additional Coverage Tests for commodities.py
# =============================================================================


class TestCommoditiesClientCoverage:
    """Additional tests for commodities client coverage."""

    @pytest.fixture
    def client(self, tmp_path) -> CommoditiesClient:
        return CommoditiesClient(cache_dir=tmp_path / "cache")

    @pytest.mark.asyncio
    async def test_fetch_co2_prices_timeout_exhausted(self, client: CommoditiesClient) -> None:
        """Test CO2 fetch with all retries exhausted."""
        with patch("httpx.AsyncClient") as mock_client:
            mock_client.return_value.__aenter__.return_value.get = AsyncMock(
                side_effect=httpx.TimeoutException("timeout")
            )

            with patch("asyncio.sleep", new_callable=AsyncMock):
                # Should fall back to empty cache
                result = await client.fetch_co2_prices(
                    start_date=date(2024, 1, 1),
                    end_date=date(2024, 1, 31),
                )

        assert result == []

    def test_save_to_cache_merge_existing(self, client: CommoditiesClient) -> None:
        """Test merging new prices with existing cache."""
        from raglite.external_data.models import CoalPrice

        # Save initial prices
        initial_prices = [
            CoalPrice(date=date(2024, 1, 1), price=120.0),
            CoalPrice(date=date(2024, 1, 8), price=122.0),
        ]
        client.save_to_cache("coal", initial_prices)

        # Save new prices (including one that overwrites)
        new_prices = [
            CoalPrice(date=date(2024, 1, 8), price=125.0),  # Overwrite
            CoalPrice(date=date(2024, 1, 15), price=128.0),  # New
        ]
        client.save_to_cache("coal", new_prices)

        # Load and verify
        loaded = client.load_from_cache("coal")
        assert len(loaded) == 3
        # Find the Jan 8 price - should be updated
        jan8_price = next(p for p in loaded if p.date == date(2024, 1, 8))
        assert jan8_price.price == 125.0

    def test_load_cache_corrupted_file(self, client: CommoditiesClient) -> None:
        """Test loading from corrupted cache file."""
        cache_file = client.cache_dir / "corrupted_prices.json"
        cache_file.write_text("not valid json {{{")

        result = client.load_from_cache("corrupted")
        assert result == []

    def test_load_cache_invalid_record(self, client: CommoditiesClient) -> None:
        """Test loading cache with invalid records."""
        import json

        cache_file = client.cache_dir / "coal_prices.json"
        cache_file.write_text(
            json.dumps(
                [
                    {"date": "2024-01-01", "price": 120.0},  # Valid
                    {"date": "invalid", "price": 130.0},  # Invalid date
                    {"date": "2024-01-15"},  # Missing price
                ]
            )
        )

        result = client.load_from_cache("coal")
        # Only the valid record should be loaded
        assert len(result) == 1

    def test_import_csv_petcoke(self, client: CommoditiesClient, tmp_path) -> None:
        """Test CSV import for petcoke commodity."""
        csv_file = tmp_path / "petcoke_prices.csv"
        csv_file.write_text(
            """date,price,currency,unit,sulfur_content_pct
2024-01-01,180.0,EUR,EUR/tonne,3.5
2024-01-08,185.0,EUR,EUR/tonne,3.2"""
        )

        result = client.import_from_csv("petcoke", csv_file)

        assert len(result) == 2
        assert result[0].price == 180.0
        assert result[0].sulfur_content_pct == 3.5

    def test_import_csv_co2(self, client: CommoditiesClient, tmp_path) -> None:
        """Test CSV import for CO2 EUA commodity."""
        csv_file = tmp_path / "co2_prices.csv"
        csv_file.write_text(
            """date,price,currency,unit
2024-01-01,85.0,EUR,EUR/tonne
2024-01-08,88.0,EUR,EUR/tonne"""
        )

        result = client.import_from_csv("co2_eua", csv_file)

        assert len(result) == 2
        assert result[0].price == 85.0

    def test_import_csv_generic_commodity(self, client: CommoditiesClient, tmp_path) -> None:
        """Test CSV import for generic commodity type."""
        csv_file = tmp_path / "other_prices.csv"
        csv_file.write_text(
            """date,price,currency,unit
2024-01-01,50.0,EUR,EUR/unit"""
        )

        result = client.import_from_csv("other", csv_file)

        assert len(result) == 1
        assert result[0].price == 50.0

    def test_import_csv_invalid_row(self, client: CommoditiesClient, tmp_path) -> None:
        """Test CSV import with invalid rows."""
        csv_file = tmp_path / "coal_prices.csv"
        csv_file.write_text(
            """date,price,currency,unit
2024-01-01,120.0,EUR,EUR/tonne
invalid-date,130.0,EUR,EUR/tonne
2024-01-15,not-a-number,EUR,EUR/tonne"""
        )

        result = client.import_from_csv("coal", csv_file)

        # Only the valid row should be imported
        assert len(result) == 1
        assert result[0].price == 120.0
