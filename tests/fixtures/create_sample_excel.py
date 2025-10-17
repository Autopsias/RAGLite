"""Generate sample Excel file for integration testing.

Creates a multi-sheet financial workbook with realistic data for testing
the Excel extraction pipeline.
"""

from pathlib import Path

import openpyxl


def create_sample_financial_excel() -> Path:
    """Create sample financial Excel file with 3 sheets."""
    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Revenue Data
    ws1 = wb.create_sheet("Revenue_Analysis")
    ws1.append(["Quarter", "Revenue", "Cost", "Profit", "Margin"])
    ws1.append(["Q1 2024", "$1,250,000", "$800,000", "$450,000", "36%"])
    ws1.append(["Q2 2024", "$1,500,000", "$900,000", "$600,000", "40%"])
    ws1.append(["Q3 2024", "$1,750,000", "$1,000,000", "$750,000", "43%"])
    ws1.append(["Q4 2024", "$2,000,000", "$1,100,000", "$900,000", "45%"])

    # Sheet 2: Balance Sheet
    ws2 = wb.create_sheet("Balance_Sheet")
    ws2.append(["Account", "2024", "2023", "Change"])
    ws2.append(["Assets", "$10,500,000", "$9,200,000", "14.1%"])
    ws2.append(["Liabilities", "$6,200,000", "$5,500,000", "12.7%"])
    ws2.append(["Equity", "$4,300,000", "$3,700,000", "16.2%"])

    # Sheet 3: Metrics
    ws3 = wb.create_sheet("Key_Metrics")
    ws3.append(["Metric", "Value", "Target", "Status"])
    ws3.append(["EBITDA", "$2,700,000", "$2,500,000", "Above"])
    ws3.append(["Operating Margin", "42%", "40%", "Above"])
    ws3.append(["ROE", "20.9%", "18%", "Above"])
    ws3.append(["Debt-to-Equity", "1.44", "1.50", "Within Range"])

    # Save workbook
    output_path = Path(__file__).parent / "sample_financial_data.xlsx"
    wb.save(output_path)
    print(f"Created sample Excel file: {output_path}")
    return output_path


if __name__ == "__main__":
    create_sample_financial_excel()
