"""Document ingestion for PDF and Excel files.

Extracts text, tables, and page numbers from financial documents.
"""

from __future__ import annotations

import asyncio
import base64
import os
import sys
import tempfile
import time
from collections.abc import Generator
from contextlib import contextmanager
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    pass

import openpyxl
import pandas as pd
from qdrant_client.models import Distance, SparseIndexParams, SparseVectorParams, VectorParams

from raglite.ingestion.chunking_strategy import chunk_by_docling_items, chunk_document
from raglite.ingestion.embedding_generation import extract_chunk_metadata, generate_embeddings
from raglite.ingestion.storage_operations import (
    store_metadata_in_postgresql,
    store_tables_in_postgresql,
    store_vectors_in_qdrant,
)
from raglite.ingestion.table_extraction import TableExtractor
from raglite.shared.clients import get_mistral_client, get_qdrant_client
from raglite.shared.config import settings
from raglite.shared.logging import get_logger
from raglite.shared.models import BatchIngestionResult, Chunk, DocumentMetadata, ExtractedMetadata
from raglite.shared.safety import SafetyGuard

logger = get_logger(__name__)

# Story 4.0.7: Maximum base64 content size (25MB encoded ≈ 18MB decoded)
MAX_BASE64_CONTENT_SIZE_BYTES = 25 * 1024 * 1024  # 25MB

# Story 4.0.8: Maximum URL download size (50MB - larger than base64 since no encoding overhead)
MAX_URL_DOWNLOAD_SIZE_BYTES = 50 * 1024 * 1024  # 50MB

# Story 4.0.8: URL download timeout (30 seconds for connection, 300 seconds total for large files)
URL_DOWNLOAD_TIMEOUT_CONNECT = 30
URL_DOWNLOAD_TIMEOUT_TOTAL = 300

# Story 4.0.7: Supported file extensions for base64 ingestion
SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls"}

# Story 4.0.8: Allowed URL schemes for security
ALLOWED_URL_SCHEMES = {"http", "https"}

# Story 4.0.8: Domain allowlist for URL downloads (empty = all domains allowed)
# Can be configured via environment variable URL_DOMAIN_ALLOWLIST (comma-separated)
URL_DOMAIN_ALLOWLIST: set[str] = (
    set()
)  # e.g., {"drive.google.com", "dropbox.com", "s3.amazonaws.com"}


@contextmanager
def temp_file_from_base64(content_b64: str, filename: str) -> Generator[str, None, None]:
    """Create temporary file from base64 content with automatic cleanup.

    Story 4.0.7 AC3/AC4: Context manager for safe temporary file handling.
    Decodes base64 content, writes to temp file, and ensures cleanup on exit.

    Args:
        content_b64: Base64-encoded file content (max 25MB encoded).
        filename: Original filename with extension (e.g., "report.pdf").
                  Used for extension detection and validation.

    Yields:
        str: Absolute path to temporary file with correct extension.

    Raises:
        ValueError: If base64 content is invalid, extension unsupported,
                    or size exceeds 25MB limit.

    Example:
        >>> with temp_file_from_base64(pdf_b64, "report.pdf") as tmp_path:
        ...     metadata = await ingest_document(tmp_path)
        >>> # tmp_path is automatically deleted after context exits
    """
    # AC5: Size check (before decoding to fail fast)
    if len(content_b64) > MAX_BASE64_CONTENT_SIZE_BYTES:
        size_mb = len(content_b64) / (1024 * 1024)
        raise ValueError(
            f"File content ({size_mb:.1f}MB encoded) exceeds 25MB limit. "
            "For larger files, save to filesystem and use doc_path parameter."
        )

    # AC3: Decode base64
    try:
        file_bytes = base64.b64decode(content_b64)
    except Exception as e:
        raise ValueError(f"Invalid base64 content: {e}") from e

    # AC6: Extension validation
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Unsupported file type: {suffix}. Supported extensions: {supported}")

    # Create temp file with correct extension (required for format detection)
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        logger.info(
            "Created temp file from base64 content",
            extra={
                "original_filename": filename,
                "extension": suffix,
                "size_bytes": len(file_bytes),
                "temp_path": tmp_path,
            },
        )

        yield tmp_path

    finally:
        # AC4: Guaranteed cleanup on success or failure
        if tmp_path:
            try:
                Path(tmp_path).unlink(missing_ok=True)
                logger.debug(
                    "Cleaned up temp file",
                    extra={"temp_path": tmp_path},
                )
            except Exception as e:
                logger.warning(
                    "Failed to clean up temp file",
                    extra={"temp_path": tmp_path, "error": str(e)},
                )


@contextmanager
def temp_file_from_url(url: str) -> Generator[tuple[str, str], None, None]:
    """Download file from URL to temporary file with automatic cleanup.

    Story 4.0.8: URL-based ingestion for Claude.ai and Claude Desktop compatibility.
    Downloads file from HTTP/HTTPS URL, validates, and provides temp file path.

    This solves the MCP file transfer limitation where:
    - Claude.ai cannot access uploaded files (sandboxed at /mnt/user-data/uploads/)
    - Claude Desktop uploads are also sandboxed, not accessible to MCP servers
    - URL-based ingestion works universally across all Claude clients

    Args:
        url: HTTP or HTTPS URL to download file from.
             Supports direct download links from:
             - Google Drive (use export links)
             - Dropbox (use dl=1 parameter)
             - S3 presigned URLs
             - Any direct file URL

    Yields:
        tuple[str, str]: (temp_file_path, detected_filename)

    Raises:
        ValueError: If URL scheme is not allowed, domain not in allowlist,
                    file too large, or extension not supported.
        RuntimeError: If download fails (network error, 404, etc.)

    Example:
        >>> with temp_file_from_url("https://example.com/report.pdf") as (path, name):
        ...     metadata = await ingest_document(path)
        >>> # temp file automatically cleaned up
    """
    import urllib.parse
    import urllib.request
    from urllib.error import HTTPError, URLError

    # Parse and validate URL
    parsed = urllib.parse.urlparse(url)

    # AC1: Scheme validation (security)
    if parsed.scheme.lower() not in ALLOWED_URL_SCHEMES:
        raise ValueError(
            f"URL scheme '{parsed.scheme}' not allowed. "
            f"Supported schemes: {', '.join(sorted(ALLOWED_URL_SCHEMES))}"
        )

    # AC2: Domain allowlist check (if configured)
    if URL_DOMAIN_ALLOWLIST and parsed.netloc.lower() not in URL_DOMAIN_ALLOWLIST:
        raise ValueError(
            f"Domain '{parsed.netloc}' not in allowlist. "
            "Contact administrator to add trusted domains."
        )

    # Extract filename from URL path or Content-Disposition header
    url_path = urllib.parse.unquote(parsed.path)
    filename_from_url = Path(url_path).name if url_path else ""

    # Validate extension from URL (preliminary check)
    if filename_from_url:
        suffix = Path(filename_from_url).suffix.lower()
        if suffix and suffix not in SUPPORTED_EXTENSIONS:
            supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
            raise ValueError(
                f"Unsupported file type from URL: {suffix}. Supported extensions: {supported}"
            )

    logger.info(
        "Starting URL download",
        extra={
            "url_domain": parsed.netloc,
            "url_path": url_path[:100],  # Truncate for logging
            "detected_filename": filename_from_url,
        },
    )

    tmp_path = None
    try:
        # Create request with timeout and headers
        request = urllib.request.Request(
            url,
            headers={
                "User-Agent": "RAGLite/1.0 (Financial Document Ingestion)",
                "Accept": "application/pdf, application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/vnd.ms-excel, */*",
            },
        )

        # Download with streaming to handle large files
        with urllib.request.urlopen(request, timeout=URL_DOWNLOAD_TIMEOUT_TOTAL) as response:  # nosec B310 - URL scheme validated above
            # Check Content-Length if available
            content_length = response.headers.get("Content-Length")
            if content_length and int(content_length) > MAX_URL_DOWNLOAD_SIZE_BYTES:
                size_mb = int(content_length) / (1024 * 1024)
                raise ValueError(
                    f"File too large ({size_mb:.1f}MB). Maximum allowed: "
                    f"{MAX_URL_DOWNLOAD_SIZE_BYTES / (1024 * 1024):.0f}MB"
                )

            # Try to get filename from Content-Disposition header
            content_disposition = response.headers.get("Content-Disposition", "")
            if "filename=" in content_disposition:
                # Extract filename from header
                import re

                match = re.search(r'filename[*]?=["\']?([^"\';\n]+)', content_disposition)
                if match:
                    filename_from_url = match.group(1).strip()

            # Determine file extension
            suffix = ""
            if filename_from_url:
                suffix = Path(filename_from_url).suffix.lower()

            # If no extension from URL/headers, try Content-Type
            if not suffix:
                content_type = response.headers.get("Content-Type", "")
                if "pdf" in content_type:
                    suffix = ".pdf"
                    filename_from_url = "downloaded_document.pdf"
                elif "spreadsheet" in content_type or "excel" in content_type:
                    suffix = ".xlsx"
                    filename_from_url = "downloaded_document.xlsx"
                elif not filename_from_url:
                    raise ValueError(
                        "Cannot determine file type from URL. "
                        "Ensure URL ends with .pdf, .xlsx, or .xls, "
                        "or server provides Content-Type header."
                    )

            # Final extension validation
            if suffix not in SUPPORTED_EXTENSIONS:
                supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
                raise ValueError(f"Unsupported file type: {suffix}. Supported: {supported}")

            # Create temp file and download content
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp_path = tmp.name
                downloaded_size = 0
                chunk_size = 8192  # 8KB chunks

                while True:
                    chunk = response.read(chunk_size)
                    if not chunk:
                        break
                    downloaded_size += len(chunk)

                    # Check size limit during download
                    if downloaded_size > MAX_URL_DOWNLOAD_SIZE_BYTES:
                        raise ValueError(
                            f"Download exceeded size limit during transfer. "
                            f"Maximum: {MAX_URL_DOWNLOAD_SIZE_BYTES / (1024 * 1024):.0f}MB"
                        )

                    tmp.write(chunk)

            logger.info(
                "URL download complete",
                extra={
                    "url_domain": parsed.netloc,
                    "doc_filename": filename_from_url,
                    "size_bytes": downloaded_size,
                    "temp_path": tmp_path,
                },
            )

            yield tmp_path, filename_from_url

    except HTTPError as e:
        logger.error(
            "HTTP error during URL download",
            extra={
                "url_domain": parsed.netloc,
                "status_code": e.code,
                "reason": e.reason,
            },
        )
        raise RuntimeError(f"Failed to download from URL: HTTP {e.code} {e.reason}") from e

    except URLError as e:
        logger.error(
            "Network error during URL download",
            extra={"url_domain": parsed.netloc, "error": str(e.reason)},
        )
        raise RuntimeError(f"Failed to download from URL: {e.reason}") from e

    except TimeoutError:
        logger.error(
            "Timeout during URL download",
            extra={
                "url_domain": parsed.netloc,
                "timeout_seconds": URL_DOWNLOAD_TIMEOUT_TOTAL,
            },
        )
        raise RuntimeError(
            f"Download timed out after {URL_DOWNLOAD_TIMEOUT_TOTAL} seconds. "
            "Try a faster connection or smaller file."
        ) from None

    finally:
        # Cleanup temp file
        if tmp_path:
            try:
                Path(tmp_path).unlink(missing_ok=True)
                logger.debug(
                    "Cleaned up temp file from URL download",
                    extra={"temp_path": tmp_path},
                )
            except Exception as e:
                logger.warning(
                    "Failed to clean up temp file from URL download",
                    extra={"temp_path": tmp_path, "error": str(e)},
                )


async def ingest_document(
    file_path: str, unit_cache: dict[str, str] | None = None
) -> DocumentMetadata:
    """Ingest financial document (PDF or Excel) with automatic format detection.

    Routes documents to appropriate extraction handler based on file extension.
    Supports PDF (.pdf) and Excel (.xlsx, .xls) formats.

    Story 5.0.6 AC3: Supports cross-document unit cache for 30% additional API reduction.

    Args:
        file_path: Path to document file (relative or absolute)
        unit_cache: Optional shared cache for cross-document unit inference (AC3).
                   If None, creates local cache. If provided (from parallel ingestion),
                   enables metric unit reuse across documents in a batch.

    Returns:
        DocumentMetadata with extraction results

    Raises:
        FileNotFoundError: If document file doesn't exist
        RuntimeError: If parsing fails or format is unsupported
        ValueError: If file extension is not supported

    Example:
        >>> metadata = await ingest_document("reports/Q4_2024.pdf")
        >>> metadata = await ingest_document("data/financials.xlsx")

        >>> # Batch ingestion with shared cache (Story 5.0.6 AC3)
        >>> cache = {}
        >>> meta1 = await ingest_document("report1.pdf", unit_cache=cache)
        >>> meta2 = await ingest_document("report2.pdf", unit_cache=cache)  # Reuses cache
    """
    # Resolve file path to check extension
    doc_path = Path(file_path).resolve()

    if not doc_path.exists():
        error_msg = f"Document file not found: {file_path}"
        logger.error(
            "Document ingestion failed - file not found",
            extra={"path": str(doc_path), "error": error_msg},
        )
        raise FileNotFoundError(error_msg)

    # Route based on file extension
    extension = doc_path.suffix.lower()

    if extension == ".pdf":
        return await ingest_pdf(str(doc_path), unit_cache=unit_cache)
    elif extension in [".xlsx", ".xls"]:
        return await extract_excel(str(doc_path))
    else:
        error_msg = f"Unsupported file format: {extension}. Supported formats: .pdf, .xlsx, .xls"
        logger.error(
            "Unsupported document format",
            extra={"path": str(doc_path), "extension": extension},
        )
        raise ValueError(error_msg)


async def ingest_pdf(
    file_path: str,
    clear_existing: bool = False,
    skip_metadata: bool = False,
    force_production: bool = False,
    unit_cache: dict[str, str] | None = None,
) -> DocumentMetadata:
    """Ingest financial PDF and extract text, tables, and structure with page numbers.

    Uses Docling library for high-accuracy extraction (97.9% table accuracy).
    Extracts page numbers from element provenance metadata.

    Story 4.0.6: Changed default from clear_collection=True to clear_existing=False
    to prevent accidental data loss. Production operations require explicit override.

    Story 5.0.6 AC3: Supports cross-document unit cache for 30% additional API reduction.

    Args:
        file_path: Path to PDF file (relative or absolute)
        clear_existing: If True, clears existing Qdrant collection and PostgreSQL
                       tables before ingestion. Default False to preserve data.
        skip_metadata: If True, skips LLM metadata extraction (Story 2.4) to avoid API issues.
                       Default False. Use when Mistral API is unavailable.
        force_production: If True, allows clear_existing on production database.
                         Required for intentional production data replacement.
        unit_cache: Optional shared cache for cross-document unit inference (AC3).
                   If None, creates local cache. If provided (from parallel ingestion),
                   enables metric unit reuse across documents in a batch.

    Returns:
        DocumentMetadata with extraction results including page_count and ingestion timestamp

    Raises:
        FileNotFoundError: If PDF file doesn't exist at specified path
        RuntimeError: If Docling parsing fails or PDF is corrupted
        ProductionProtectionError: If clear_existing=True on production without force_production

    Example:
        >>> metadata = await ingest_pdf("docs/sample pdf/report.pdf")
        >>> print(f"Ingested {metadata.page_count} pages")

        >>> # Clear existing data (safe in test environment)
        >>> metadata = await ingest_pdf("report.pdf", clear_existing=True)

        >>> # Clear production data (requires explicit override)
        >>> metadata = await ingest_pdf("report.pdf", clear_existing=True, force_production=True)

        >>> # Skip metadata extraction to avoid API errors
        >>> metadata = await ingest_pdf("report.pdf", skip_metadata=True)

        >>> # Batch ingestion with shared cache (Story 5.0.6 AC3)
        >>> cache = {}
        >>> meta1 = await ingest_pdf("report1.pdf", unit_cache=cache)
        >>> meta2 = await ingest_pdf("report2.pdf", unit_cache=cache)  # Reuses cache
    """
    # Lazy import Docling: Avoid hanging on import when this module loads
    # Docling initializes PyTorch/CUDA on import which can hang without GPU
    # Only load when actually ingesting PDFs
    from docling.datamodel.accelerator_options import AcceleratorOptions
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions, TableFormerMode
    from docling.document_converter import DocumentConverter, PdfFormatOption

    start_time = time.time()

    # Checkpoint: Start of function
    print(f"CHECKPOINT: ingest_pdf started for {file_path}", file=sys.stderr, flush=True)

    # Resolve file path
    pdf_path = Path(file_path).resolve()
    print(f"CHECKPOINT: PDF path resolved to {pdf_path}", file=sys.stderr, flush=True)

    if not pdf_path.exists():
        error_msg = f"PDF file not found: {file_path}"
        logger.error(
            "PDF ingestion failed - file not found",
            extra={"path": str(pdf_path), "error": error_msg},
        )
        raise FileNotFoundError(error_msg)

    # Story 4.0.6: SafetyGuard protection for destructive operations
    guard = SafetyGuard()

    # Clear Qdrant collection if requested (Story 2.2 fix - prevent data contamination)
    # Story 4.0.6: Default changed to False, requires explicit clear_existing=True
    if clear_existing:
        # AC1/AC2: Check environment before destructive operation
        guard.check_environment("clear_collection", force_production=force_production)

        # AC2: Require confirmation in interactive mode for production
        if guard.is_production and not force_production:
            if not guard.require_confirmation("About to DELETE ALL DATA in production database"):
                raise SystemExit("Operation cancelled by user")

        client = get_qdrant_client()
        try:
            client.delete_collection(settings.qdrant_collection_name)
            logger.info(
                "Cleared existing collection",
                extra={
                    "collection": settings.qdrant_collection_name,
                    "environment": "PRODUCTION" if guard.is_production else "TEST",
                },
            )
        except Exception:
            logger.info(
                "Collection doesn't exist, will create new",
                extra={"collection": settings.qdrant_collection_name},
            )

        # CRITICAL FIX: Also clear PostgreSQL to maintain symmetric data lifecycle
        # This prevents mixed document IDs from accumulating across ingestion runs
        try:
            import psycopg2

            conn_str = f"postgresql://{settings.postgres_user}:{settings.postgres_password}@{settings.postgres_host}:{settings.postgres_port}/{settings.postgres_db}"
            conn = psycopg2.connect(conn_str)
            cursor = conn.cursor()

            # Delete all data from both PostgreSQL tables
            cursor.execute("DELETE FROM financial_chunks")
            chunks_deleted = cursor.rowcount
            cursor.execute("DELETE FROM financial_tables")
            tables_deleted = cursor.rowcount

            conn.commit()
            cursor.close()
            conn.close()

            logger.info(
                "Cleared PostgreSQL tables",
                extra={
                    "financial_chunks_deleted": chunks_deleted,
                    "financial_tables_deleted": tables_deleted,
                    "environment": "PRODUCTION" if guard.is_production else "TEST",
                },
            )
        except Exception as e:
            logger.warning(
                "Failed to clear PostgreSQL tables (might not exist yet)",
                extra={"error": str(e)},
            )

        # Recreate collection with proper config (named vectors + sparse for BM25)
        try:
            client.create_collection(
                collection_name=settings.qdrant_collection_name,
                vectors_config={
                    "text-dense": VectorParams(size=1024, distance=Distance.COSINE),
                },
                sparse_vectors_config={
                    "text-sparse": SparseVectorParams(
                        index=SparseIndexParams(on_disk=False),
                    )
                },
            )
            logger.info(
                "Created fresh collection",
                extra={
                    "collection": settings.qdrant_collection_name,
                    "vector_size": 1024,
                },
            )
        except Exception as e:
            logger.warning(
                "Collection may already exist",
                extra={"collection": settings.qdrant_collection_name, "error": str(e)},
            )

    logger.info(
        "Starting PDF ingestion",
        extra={
            "path": str(pdf_path),
            "doc_filename": pdf_path.name,
            "size_mb": round(pdf_path.stat().st_size / (1024 * 1024), 2),
            "clear_existing": clear_existing,
            "environment": "PRODUCTION" if guard.is_production else "TEST",
        },
    )

    # Initialize Docling converter with table extraction enabled (Story 1.15 fix)
    # Configure table structure recognition to extract table cell data
    # Story 2.1: Use pypdfium backend for 50-60% memory reduction
    try:
        # Story 2.2: Configure parallel page processing
        # Thread count configurable via PDF_PROCESSING_THREADS env var (default: 8)
        # NOTE: Default AcceleratorOptions is 4 threads - we use 8 for 1.55x speedup
        thread_count = settings.pdf_processing_threads

        # Story 2.3 Fix: Add document_timeout to prevent indefinite hangs on large PDFs
        # Timeout set to 1500s (25 minutes) for 160-page PDFs
        # Based on: 40-page = 3m51s, 160-page expected = ~15-18min, buffer = 25min
        pipeline_options = PdfPipelineOptions(
            do_table_structure=True,
            do_ocr=False,  # Disable OCR for 50% speedup - financial PDFs have embedded text
            accelerator_options=AcceleratorOptions(
                num_threads=thread_count,
                device="cpu",  # CRITICAL: Force CPU-only mode to prevent CUDA hang on CI runners
            ),
            document_timeout=1500,  # 25 minutes max per document
        )
        pipeline_options.table_structure_options.mode = TableFormerMode.ACCURATE

        # Story 2.1: PyPdfium backend (optimized)
        from docling.backend.pypdfium2_backend import PyPdfiumDocumentBackend

        print("CHECKPOINT: Creating DocumentConverter...", file=sys.stderr, flush=True)
        converter = DocumentConverter(
            format_options={
                InputFormat.PDF: PdfFormatOption(
                    pipeline_options=pipeline_options, backend=PyPdfiumDocumentBackend
                )
            }
        )
        print(
            "CHECKPOINT: DocumentConverter created successfully",
            file=sys.stderr,
            flush=True,
        )
        logger.info(
            "Docling converter initialized with pypdfium backend and table extraction",
            extra={
                "table_mode": "ACCURATE",
                "backend": "pypdfium",
                "num_threads": thread_count,
                "path": str(pdf_path),
            },
        )
    except Exception as e:
        error_msg = f"Failed to initialize Docling converter: {e}"
        logger.error(
            "Docling initialization failed",
            extra={"path": str(pdf_path), "error": str(e)},
            exc_info=True,
        )
        raise RuntimeError(error_msg) from e

    # Convert PDF with Docling
    try:
        print(
            f"CHECKPOINT: Starting Docling conversion of {pdf_path.name}...",
            file=sys.stderr,
            flush=True,
        )
        result = converter.convert(str(pdf_path))
        print(
            f"CHECKPOINT: Docling conversion complete - {result.document.num_pages()} pages",
            file=sys.stderr,
            flush=True,
        )
    except Exception as e:
        error_msg = f"Docling parsing failed for {pdf_path.name}: {e}"
        logger.error(
            "PDF parsing failed",
            extra={
                "path": str(pdf_path),
                "doc_filename": pdf_path.name,
                "error": str(e),
            },
            exc_info=True,
        )
        raise RuntimeError(error_msg) from e

    # Story 2.13 AC1: Extract tables to PostgreSQL (avoid double-conversion)
    # Extract tables from Docling result before chunking to reuse conversion
    print("CHECKPOINT: Starting table extraction...", file=sys.stderr, flush=True)
    logger.info(
        "Extracting tables for SQL storage",
        extra={"doc_filename": pdf_path.name},
    )

    # Skip table extraction in CI if env var set (for debugging hangs)
    skip_table_extraction = os.getenv("SKIP_TABLE_EXTRACTION", "false").lower() == "true"

    try:
        if not skip_table_extraction:
            # Reuse existing converter to avoid duplicate initialization (more efficient)
            # and to enable test mocking (tests can mock converter once, used by both stages)
            extractor = TableExtractor(converter=converter)
            # Milestone 1: Async table extraction with 10x speedup (62 min → 6 min)
            # Story 5.0.6 AC3: Pass unit_cache for cross-document reuse
            # FIX: Use pdf_path.name (with extension) for consistency with Qdrant document IDs
            table_rows = await extractor.extract_tables_from_result(
                result, pdf_path.name, unit_cache=unit_cache
            )
            print(
                f"CHECKPOINT: Table extraction complete - {len(table_rows)} rows",
                file=sys.stderr,
                flush=True,
            )
        else:
            table_rows = []
            print("CHECKPOINT: Table extraction SKIPPED", file=sys.stderr, flush=True)

        if table_rows:
            logger.info(
                "Tables extracted from document",
                extra={
                    "doc_filename": pdf_path.name,
                    "table_count": len({row["table_index"] for row in table_rows}),
                    "row_count": len(table_rows),
                },
            )

            # Store tables in PostgreSQL
            rows_stored, rows_skipped = await store_tables_in_postgresql(table_rows)
            logger.info(
                "Table storage complete",
                extra={
                    "doc_filename": pdf_path.name,
                    "rows_stored": rows_stored,
                    "rows_skipped": rows_skipped,
                },
            )
        else:
            logger.info(
                "No tables found in document",
                extra={"doc_filename": pdf_path.name},
            )
    except Exception as e:
        # Don't fail ingestion if table extraction fails - log and continue
        # Story: Fix PostgreSQL Data Synchronization Gap - Make failures visible
        logger.warning(
            "Table extraction failed - document will have vectors but no structured tables in PostgreSQL. "
            "Run 'python scripts/backfill-postgresql-tables.py' to fix after resolving the error.",
            extra={
                "doc_filename": pdf_path.name,
                "error": str(e),
                "error_type": type(e).__name__,
                "action_required": "backfill_tables",
            },
            exc_info=True,
        )

    # Extract page count from DoclingDocument
    # Use num_pages() method which returns total page count
    page_count = result.document.num_pages()

    # Count elements with provenance data for metrics
    total_elements = 0
    elements_with_pages = 0

    for item, _ in result.document.iterate_items():
        total_elements += 1
        if hasattr(item, "prov") and item.prov:
            elements_with_pages += 1

    # Validate page extraction
    if page_count == 0:
        logger.warning(
            "No page numbers extracted - verify PDF structure",
            extra={"path": str(pdf_path), "total_elements": total_elements},
        )

    # Extract full text from Docling result
    # Use export_to_markdown() to get structured text with tables
    try:
        result.document.export_to_markdown()
    except Exception as e:
        logger.warning(
            "Failed to export markdown - falling back to plain text",
            extra={"path": str(pdf_path), "error": str(e)},
        )
        # Fallback: concatenate all text from elements
        "\n".join(item.text for item, _ in result.document.iterate_items() if hasattr(item, "text"))

    # Create initial metadata for chunking
    metadata = DocumentMetadata(
        filename=pdf_path.name,
        doc_type="PDF",
        ingestion_timestamp=datetime.now(UTC).isoformat(),
        page_count=page_count,
        source_path=str(pdf_path),
        chunk_count=0,  # Will be updated after chunking
    )

    # Chunk the document using Docling items with provenance (Story 1.13 fix)
    # This extracts actual page numbers from Docling metadata instead of estimating
    print("CHECKPOINT: Starting chunking...", file=sys.stderr, flush=True)
    chunks = await chunk_by_docling_items(result, metadata)
    print(
        f"CHECKPOINT: Chunking complete - {len(chunks)} chunks created",
        file=sys.stderr,
        flush=True,
    )

    # Story 2.4 AC1 (REVISED): Extract business context metadata PER CHUNK using Mistral Small
    # ARCHITECTURAL CHANGE: Per-chunk extraction avoids reasoning token overflow and provides
    # more accurate metadata for each chunk (chunks are ~512 tokens, perfect size for Mistral Small)
    # RE-ENABLED for Story 2.6 PostgreSQL data migration
    # NOTE: Performance bugs (sync client, per-request client, no timeout) will be fixed in Task 6 (AC6)
    if settings.mistral_api_key and not skip_metadata:
        logger.info(
            "Starting per-chunk metadata extraction with Mistral Small",
            extra={
                "doc_filename": pdf_path.name,
                "chunk_count": len(chunks),
                "model": settings.metadata_extraction_model,
                "expected_time_sec": len(chunks) * 2,  # ~2 sec per chunk estimate
            },
        )

        # Story 2.6 AC6 FIX: Create single Mistral client for reuse (client pooling)
        # This eliminates per-request connection overhead (10-15x speedup)
        # Now uses shared client factory with timeout configuration
        mistral_client = get_mistral_client()

        # Extract metadata for each chunk with rate limiting (Story 2.4 FIX + Story 2.5 OPTIMIZATION)
        # Semaphore limits concurrent API calls to respect Mistral rate limits
        # RATE LIMIT FIX: Reduced from 20 to 5 concurrent requests to avoid 429 errors
        # Mistral Free API has stricter rate limits than initially tested
        semaphore = asyncio.Semaphore(5)  # Max 5 concurrent requests to Mistral API

        async def extract_for_chunk(
            chunk: Chunk,
        ) -> tuple[Chunk, ExtractedMetadata | None]:
            """Extract metadata for a single chunk with error handling and rate limiting."""
            async with semaphore:  # Limit concurrent requests
                try:
                    # Story 2.6 AC6 FIX: Pass shared client instance to enable connection pooling
                    extracted = await extract_chunk_metadata(
                        text=chunk.content,
                        chunk_id=chunk.chunk_id,
                        client=mistral_client,
                    )
                    return (chunk, extracted)
                except Exception as e:
                    # Graceful degradation - continue without metadata for this chunk
                    logger.debug(
                        "Chunk metadata extraction failed (graceful degradation)",
                        extra={"chunk_id": chunk.chunk_id, "error": str(e)},
                    )
                    return (chunk, None)

        # Process chunks with rate-limited concurrency
        results = await asyncio.gather(*[extract_for_chunk(chunk) for chunk in chunks])

        # Inject extracted metadata into chunks (15 RICH SCHEMA fields)
        successful_extractions = 0
        for chunk, extracted_metadata in results:
            if extracted_metadata:
                # Document-Level (7 fields)
                chunk.document_type = extracted_metadata.document_type
                chunk.reporting_period = extracted_metadata.reporting_period
                chunk.time_granularity = extracted_metadata.time_granularity
                chunk.company_name = extracted_metadata.company_name
                chunk.geographic_jurisdiction = extracted_metadata.geographic_jurisdiction
                chunk.data_source_type = extracted_metadata.data_source_type
                chunk.version_date = extracted_metadata.version_date
                # Section-Level (5 fields)
                chunk.section_type = extracted_metadata.section_type
                chunk.metric_category = extracted_metadata.metric_category
                chunk.units = extracted_metadata.units
                chunk.department_scope = extracted_metadata.department_scope
                # Table-Specific (3 fields)
                chunk.table_context = extracted_metadata.table_context
                chunk.table_name = extracted_metadata.table_name
                chunk.statistical_summary = extracted_metadata.statistical_summary
                successful_extractions += 1

        logger.info(
            "Per-chunk metadata extraction complete",
            extra={
                "doc_filename": pdf_path.name,
                "total_chunks": len(chunks),
                "successful_extractions": successful_extractions,
                "success_rate": f"{successful_extractions / len(chunks) * 100:.1f}%",
            },
        )
    else:
        skip_reason = "skip_metadata=True" if skip_metadata else "MISTRAL_API_KEY not configured"
        logger.info(
            f"Metadata extraction skipped - {skip_reason}",
            extra={"doc_filename": pdf_path.name, "skip_metadata": skip_metadata},
        )

    # Generate embeddings for chunks (Story 1.5)
    chunks_with_embeddings = await generate_embeddings(chunks)

    # Store vectors in Qdrant (Story 1.6)
    if chunks_with_embeddings:
        points_stored = await store_vectors_in_qdrant(
            chunks_with_embeddings, collection_name=settings.qdrant_collection_name
        )
        logger.info(
            "Vectors stored in Qdrant",
            extra={
                "doc_filename": pdf_path.name,
                "points_stored": points_stored,
                "collection": settings.qdrant_collection_name,
            },
        )

        # Story 2.6 AC4: Store metadata in PostgreSQL for structured filtering
        # Only attempts storage if chunks have extracted metadata (company_name, metric_category, etc.)
        records_stored, records_skipped = await store_metadata_in_postgresql(chunks_with_embeddings)
        logger.info(
            "Metadata stored in PostgreSQL",
            extra={
                "doc_filename": pdf_path.name,
                "records_stored": records_stored,
                "records_skipped": records_skipped,
            },
        )

    # Update metadata with chunk count
    metadata.chunk_count = len(chunks_with_embeddings)

    # Calculate ingestion metrics
    duration_ms = int((time.time() - start_time) * 1000)

    logger.info(
        "PDF ingested successfully",
        extra={
            "doc_filename": pdf_path.name,
            "page_count": page_count,
            "chunk_count": len(chunks_with_embeddings),
            "total_elements": total_elements,
            "elements_with_pages": elements_with_pages,
            "duration_ms": duration_ms,
            "pages_per_second": (
                round(page_count / (duration_ms / 1000), 2) if duration_ms > 0 else 0
            ),
        },
    )

    return metadata


async def extract_excel(file_path: str) -> DocumentMetadata:
    """Extract financial data from Excel spreadsheet with multi-sheet support.

    Uses openpyxl for Excel parsing and pandas for data manipulation.
    Extracts all sheets preserving numeric formatting and sheet numbers for citations.

    Note: This function is marked async for consistency with the ingestion pipeline
    pattern (ingest_pdf, future chunking/embedding operations). While openpyxl and
    pandas operations are currently synchronous, this allows for future async
    enhancements like streaming large files or parallel sheet processing.

    Args:
        file_path: Path to Excel file (relative or absolute, .xlsx or .xls)

    Returns:
        DocumentMetadata with extraction results including sheet_count as page_count

    Raises:
        FileNotFoundError: If Excel file doesn't exist at specified path
        RuntimeError: If Excel parsing fails, file is password-protected, or corrupted

    Example:
        >>> metadata = await extract_excel("data/financial_report.xlsx")
        >>> print(f"Extracted {metadata.page_count} sheets")
    """
    start_time = time.time()

    # Resolve file path
    excel_path = Path(file_path).resolve()

    if not excel_path.exists():
        error_msg = f"Excel file not found: {file_path}"
        logger.error(
            "Excel extraction failed - file not found",
            extra={"path": str(excel_path), "error": error_msg},
        )
        raise FileNotFoundError(error_msg)

    logger.info(
        "Starting Excel extraction",
        extra={
            "path": str(excel_path),
            "doc_filename": excel_path.name,
            "size_mb": round(excel_path.stat().st_size / (1024 * 1024), 2),
        },
    )

    # Load Excel workbook
    try:
        # data_only=True: Load computed values instead of formulas
        workbook = openpyxl.load_workbook(str(excel_path), data_only=True)
    except openpyxl.utils.exceptions.InvalidFileException as e:
        error_msg = (
            f"Excel parsing failed for {excel_path.name}: Invalid or password-protected file"
        )
        logger.error(
            "Excel file is invalid or password-protected",
            extra={
                "path": str(excel_path),
                "doc_filename": excel_path.name,
                "error": str(e),
            },
            exc_info=True,
        )
        raise RuntimeError(error_msg) from e
    except Exception as e:
        error_msg = f"Unexpected error loading Excel file {excel_path.name}: {e}"
        logger.error(
            "Excel loading failed",
            extra={
                "path": str(excel_path),
                "doc_filename": excel_path.name,
                "error": str(e),
            },
            exc_info=True,
        )
        raise RuntimeError(error_msg) from e

    # Check for empty workbook
    if not workbook.sheetnames:
        logger.warning(
            "Empty Excel workbook - no sheets found",
            extra={"path": str(excel_path), "doc_filename": excel_path.name},
        )
        # Return metadata with zero sheets for empty workbook
        metadata = DocumentMetadata(
            filename=excel_path.name,
            doc_type="Excel",
            ingestion_timestamp=datetime.now(UTC).isoformat(),
            page_count=0,
            source_path=str(excel_path),
        )
        return metadata

    # Extract all sheets with sheet numbers
    sheets_data = []
    total_rows = 0
    skipped_sheets = 0

    try:
        for sheet_number, sheet_name in enumerate(workbook.sheetnames, start=1):
            sheet = workbook[sheet_name]

            # Convert sheet to pandas DataFrame
            # Get all cell values from the sheet
            data = list(sheet.values)

            if not data:
                # Empty sheet - skip but log
                skipped_sheets += 1
                logger.info(
                    "Empty sheet skipped",
                    extra={"sheet_name": sheet_name, "sheet_number": sheet_number},
                )
                continue

            # First row as column headers
            headers = data[0] if data else []
            rows = data[1:] if len(data) > 1 else []

            # Create DataFrame with proper headers
            df = pd.DataFrame(rows, columns=headers)

            # Convert to markdown table format (preserves numeric formatting)
            # to_markdown() preserves numbers, dates, currencies as-is
            sheet_markdown = f"## Sheet {sheet_number}: {sheet_name}\n\n"
            sheet_markdown += df.to_markdown(index=False)

            sheets_data.append(
                {
                    "sheet_name": sheet_name,
                    "sheet_number": sheet_number,
                    "content": sheet_markdown,
                    "row_count": len(df),
                }
            )

            total_rows += len(df)

    except Exception as e:
        error_msg = f"Failed to extract data from sheets in {excel_path.name}: {e}"
        logger.error(
            "Sheet extraction failed",
            extra={
                "path": str(excel_path),
                "doc_filename": excel_path.name,
                "error": str(e),
            },
            exc_info=True,
        )
        raise RuntimeError(error_msg) from e

    # Calculate extraction metrics
    sheet_count = len(sheets_data)

    # Validate sheet extraction
    if sheet_count == 0:
        logger.warning(
            "No sheets extracted - verify Excel file structure",
            extra={"path": str(excel_path), "total_sheets": len(workbook.sheetnames)},
        )

    # Concatenate all sheet markdown for chunking
    full_text = "\n\n".join(sheet["content"] for sheet in sheets_data)

    # Create initial metadata for chunking (use sheet_count as page_count)
    metadata = DocumentMetadata(
        filename=excel_path.name,
        doc_type="Excel",
        ingestion_timestamp=datetime.now(UTC).isoformat(),
        page_count=sheet_count,
        source_path=str(excel_path),
        chunk_count=0,  # Will be updated after chunking
    )

    # Chunk the document if there's content
    chunks = []
    if full_text.strip():
        chunks = await chunk_document(full_text, metadata)

    # Generate embeddings for chunks (Story 1.5)
    chunks_with_embeddings = []
    if chunks:
        chunks_with_embeddings = await generate_embeddings(chunks)

    # Store vectors in Qdrant (Story 1.6)
    if chunks_with_embeddings:
        points_stored = await store_vectors_in_qdrant(
            chunks_with_embeddings, collection_name=settings.qdrant_collection_name
        )
        logger.info(
            "Vectors stored in Qdrant",
            extra={
                "doc_filename": excel_path.name,
                "points_stored": points_stored,
                "collection": settings.qdrant_collection_name,
            },
        )

        # Store metadata in PostgreSQL for structured filtering (Story 2.6 AC4)
        # Maintains parity with PDF ingestion path
        records_stored, records_skipped = await store_metadata_in_postgresql(chunks_with_embeddings)
        logger.info(
            "Metadata stored in PostgreSQL",
            extra={
                "doc_filename": excel_path.name,
                "records_stored": records_stored,
                "records_skipped": records_skipped,
            },
        )

    # Update metadata with chunk count
    metadata.chunk_count = len(chunks_with_embeddings)

    # Calculate final metrics
    duration_ms = int((time.time() - start_time) * 1000)

    logger.info(
        "Excel extracted successfully",
        extra={
            "doc_filename": excel_path.name,
            "sheet_count": sheet_count,
            "chunk_count": len(chunks_with_embeddings),
            "total_rows": total_rows,
            "skipped_sheets": skipped_sheets,
            "duration_ms": duration_ms,
            "sheets_per_second": (
                round(sheet_count / (duration_ms / 1000), 2) if duration_ms > 0 else 0
            ),
        },
    )

    return metadata


async def ingest_documents_parallel(
    file_paths: list[str],
    max_concurrent: int | None = None,
) -> BatchIngestionResult:
    """Ingest multiple documents in parallel with concurrency control.

    Story 5.0.6 AC1: Parallel document ingestion with memory-safe concurrency limits.
    Uses asyncio.Semaphore to limit concurrent ingestions and prevent memory exhaustion.

    Args:
        file_paths: List of document file paths to ingest (PDF, Excel)
        max_concurrent: Maximum concurrent documents (default: settings.ingestion_parallel_docs).
                       Set to 1 for sequential processing, 2-4 for parallel.

    Returns:
        BatchIngestionResult with success/failure counts and per-document results

    Raises:
        ValueError: If file_paths is empty or max_concurrent < 1

    Example:
        >>> paths = ["report1.pdf", "report2.pdf", "data.xlsx"]
        >>> result = await ingest_documents_parallel(paths, max_concurrent=2)
        >>> print(f"Success: {result.successful}/{result.total_documents}")
        Success: 3/3
        >>> print(f"Duration: {result.duration_seconds:.1f}s")
        Duration: 45.2s

    Performance (Story 5.0.6 AC8):
        - Sequential (max_concurrent=1): ~4-5 hours for 10 PDFs (40 pages each)
        - Parallel (max_concurrent=2): ~45 minutes for 10 PDFs (6-10x speedup target)
        - Memory usage: ~4GB per concurrent document, max 8GB with default limit
    """
    # AC1: Validation
    if not file_paths:
        raise ValueError("file_paths cannot be empty")

    # AC1: Use config default if not specified
    if max_concurrent is None:
        max_concurrent = settings.ingestion_parallel_docs

    if max_concurrent < 1:
        raise ValueError(f"max_concurrent must be >= 1, got {max_concurrent}")

    # Start timing for batch
    batch_start = time.time()
    total_docs = len(file_paths)

    logger.info(
        "Starting parallel document ingestion",
        extra={
            "total_documents": total_docs,
            "max_concurrent": max_concurrent,
            "batch_size": total_docs,
        },
    )

    # AC1: Semaphore for concurrency control (max 2 by default to stay within 8GB memory)
    semaphore = asyncio.Semaphore(max_concurrent)

    # AC3: Create shared unit cache for cross-document inference (30% API reduction)
    # Cache persists across all documents in the batch, enabling metric unit reuse
    shared_unit_cache: dict[str, str] = {}

    # Results tracking
    successful_results: list[DocumentMetadata] = []
    error_details: list[dict[str, str]] = []
    completed_count = 0

    async def process_document(file_path: str, doc_index: int) -> None:
        """Process single document with semaphore control and error handling."""
        nonlocal completed_count

        async with semaphore:
            try:
                # AC1/AC8: Progress logging (before processing)
                logger.info(
                    "Processing document",
                    extra={
                        "doc_index": doc_index + 1,
                        "total_documents": total_docs,
                        "file_path": file_path,
                        "concurrent_slots": max_concurrent,
                    },
                )

                # AC3: Ingest document with shared cache for cross-document unit reuse
                metadata = await ingest_document(file_path, unit_cache=shared_unit_cache)

                successful_results.append(metadata)
                completed_count += 1

                # AC1/AC8: Progress logging (after success)
                logger.info(
                    "Document ingested successfully",
                    extra={
                        "doc_index": doc_index + 1,
                        "total_documents": total_docs,
                        "completed": completed_count,
                        "file_path": file_path,
                        "chunk_count": metadata.chunk_count,
                        "page_count": metadata.page_count,
                    },
                )

            except Exception as e:
                # AC1: Error tracking (don't fail entire batch on single document error)
                error_msg = str(e)
                error_details.append({"filename": str(file_path), "error": error_msg})
                completed_count += 1

                logger.error(
                    "Document ingestion failed",
                    extra={
                        "doc_index": doc_index + 1,
                        "total_documents": total_docs,
                        "completed": completed_count,
                        "file_path": file_path,
                        "error": error_msg,
                    },
                )

    # AC1: Launch all ingestion tasks (asyncio schedules them with semaphore control)
    tasks = [process_document(path, idx) for idx, path in enumerate(file_paths)]
    await asyncio.gather(*tasks)

    # Calculate batch duration
    batch_duration = time.time() - batch_start

    # AC1/AC8: Final batch summary logging
    success_count = len(successful_results)
    fail_count = len(error_details)
    total_chunks = sum(m.chunk_count for m in successful_results)
    total_pages = sum(m.page_count for m in successful_results)

    logger.info(
        "Parallel batch ingestion complete",
        extra={
            "total_documents": total_docs,
            "successful": success_count,
            "failed": fail_count,
            "duration_seconds": round(batch_duration, 1),
            "duration_minutes": round(batch_duration / 60, 1),
            "docs_per_minute": round(success_count / (batch_duration / 60), 2)
            if batch_duration > 0
            else 0,
            "total_chunks": total_chunks,
            "total_pages": total_pages,
            "max_concurrent": max_concurrent,
        },
    )

    # AC1: Return BatchIngestionResult
    return BatchIngestionResult(
        total_documents=total_docs,
        successful=success_count,
        failed=fail_count,
        duration_seconds=round(batch_duration, 2),
        results=successful_results,
        errors=error_details,
    )
