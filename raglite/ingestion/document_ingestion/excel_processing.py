"""Excel document processing and ingestion.

Handles Excel-specific extraction using openpyxl and pandas.
"""

from __future__ import annotations

import time
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pandas as pd

from raglite.ingestion.chunking_strategy import chunk_document
from raglite.ingestion.embedding_generation import generate_embeddings
from raglite.ingestion.storage import (
    store_metadata_in_postgresql,
    store_vectors_in_qdrant,
)
from raglite.shared.config import settings
from raglite.shared.logging import get_logger
from raglite.shared.models import DocumentMetadata

logger = get_logger(__name__)


def _validate_excel_file(excel_path: Path) -> None:
    """Validate Excel file exists and log metadata.

    Args:
        excel_path: Resolved path to Excel file

    Raises:
        FileNotFoundError: If Excel file doesn't exist
    """
    if not excel_path.exists():
        error_msg = f"Excel file not found: {excel_path}"
        logger.error(
            "Excel extraction failed - file not found",
            extra={"path": str(excel_path), "error": error_msg},
        )
        raise FileNotFoundError(error_msg)

    logger.info(
        "Starting Excel extraction",
        extra={
            "path": str(excel_path),
            "doc_filename": excel_path.name,
            "size_mb": round(excel_path.stat().st_size / (1024 * 1024), 2),
        },
    )


def _load_excel_workbook(excel_path: Path) -> openpyxl.Workbook:
    """Load Excel workbook with error handling.

    Args:
        excel_path: Path to Excel file

    Returns:
        Loaded openpyxl Workbook object

    Raises:
        RuntimeError: If Excel parsing fails, file is password-protected, or corrupted
    """
    try:
        # data_only=True: Load computed values instead of formulas
        workbook = openpyxl.load_workbook(str(excel_path), data_only=True)
        return workbook
    except openpyxl.utils.exceptions.InvalidFileException as e:
        error_msg = (
            f"Excel parsing failed for {excel_path.name}: Invalid or password-protected file"
        )
        logger.error(
            "Excel file is invalid or password-protected",
            extra={
                "path": str(excel_path),
                "doc_filename": excel_path.name,
                "error": str(e),
            },
            exc_info=True,
        )
        raise RuntimeError(error_msg) from e
    except Exception as e:
        error_msg = f"Unexpected error loading Excel file {excel_path.name}: {e}"
        logger.error(
            "Excel loading failed",
            extra={
                "path": str(excel_path),
                "doc_filename": excel_path.name,
                "error": str(e),
            },
            exc_info=True,
        )
        raise RuntimeError(error_msg) from e


def _create_empty_metadata(excel_path: Path) -> DocumentMetadata:
    """Create DocumentMetadata for empty workbook.

    Args:
        excel_path: Path to Excel file

    Returns:
        DocumentMetadata with zero sheet count
    """
    logger.warning(
        "Empty Excel workbook - no sheets found",
        extra={"path": str(excel_path), "doc_filename": excel_path.name},
    )
    return DocumentMetadata(
        filename=excel_path.name,
        doc_type="Excel",
        ingestion_timestamp=datetime.now(UTC).isoformat(),
        page_count=0,
        source_path=str(excel_path),
    )


def _extract_sheet_to_markdown(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    sheet_number: int,
    sheet_name: str,
) -> dict | None:
    """Convert single Excel sheet to markdown format.

    Args:
        sheet: openpyxl Worksheet object
        sheet_number: Sheet index (1-based)
        sheet_name: Sheet name

    Returns:
        Dictionary with sheet data or None if sheet is empty
    """
    # Get all cell values from the sheet
    data = list(sheet.values)

    if not data:
        # Empty sheet
        logger.info(
            "Empty sheet skipped",
            extra={"sheet_name": sheet_name, "sheet_number": sheet_number},
        )
        return None

    # First row as column headers
    headers = data[0] if data else []
    rows = data[1:] if len(data) > 1 else []

    # Create DataFrame with proper headers
    df = pd.DataFrame(rows, columns=headers)

    # Convert to markdown table format (preserves numeric formatting)
    # to_markdown() preserves numbers, dates, currencies as-is
    sheet_markdown = f"## Sheet {sheet_number}: {sheet_name}\n\n"
    sheet_markdown += df.to_markdown(index=False)

    return {
        "sheet_name": sheet_name,
        "sheet_number": sheet_number,
        "content": sheet_markdown,
        "row_count": len(df),
    }


def _extract_all_sheets(workbook: openpyxl.Workbook, excel_path: Path) -> tuple[list[dict], int]:
    """Extract all sheets from workbook to markdown format.

    Args:
        workbook: Loaded openpyxl Workbook
        excel_path: Path to Excel file (for logging)

    Returns:
        Tuple of (list of sheet data dicts, skipped sheets count)

    Raises:
        RuntimeError: If sheet extraction fails
    """
    sheets_data = []
    skipped_sheets = 0

    try:
        for sheet_number, sheet_name in enumerate(workbook.sheetnames, start=1):
            sheet = workbook[sheet_name]

            sheet_data = _extract_sheet_to_markdown(sheet, sheet_number, sheet_name)

            if sheet_data is None:
                skipped_sheets += 1
            else:
                sheets_data.append(sheet_data)

    except Exception as e:
        error_msg = f"Failed to extract data from sheets in {excel_path.name}: {e}"
        logger.error(
            "Sheet extraction failed",
            extra={
                "path": str(excel_path),
                "doc_filename": excel_path.name,
                "error": str(e),
            },
            exc_info=True,
        )
        raise RuntimeError(error_msg) from e

    return sheets_data, skipped_sheets


async def _process_excel_chunks(
    full_text: str,
    sheets_data: list[dict],
    excel_path: Path,
) -> list:
    """Chunk, embed, and store Excel data.

    Args:
        full_text: Concatenated markdown from all sheets
        sheets_data: List of sheet data dictionaries
        excel_path: Path to Excel file (for logging)

    Returns:
        List of chunks with embeddings
    """
    sheet_count = len(sheets_data)

    # Create initial metadata for chunking (use sheet_count as page_count)
    metadata = DocumentMetadata(
        filename=excel_path.name,
        doc_type="Excel",
        ingestion_timestamp=datetime.now(UTC).isoformat(),
        page_count=sheet_count,
        source_path=str(excel_path),
        chunk_count=0,  # Will be updated after chunking
    )

    # Chunk the document if there's content
    chunks = []
    if full_text.strip():
        chunks = await chunk_document(full_text, metadata)

    # Generate embeddings for chunks (Story 1.5)
    chunks_with_embeddings = []
    if chunks:
        chunks_with_embeddings = await generate_embeddings(chunks)

    # Store vectors in Qdrant (Story 1.6)
    if chunks_with_embeddings:
        points_stored = await store_vectors_in_qdrant(
            chunks_with_embeddings, collection_name=settings.qdrant_collection_name
        )
        logger.info(
            "Vectors stored in Qdrant",
            extra={
                "doc_filename": excel_path.name,
                "points_stored": points_stored,
                "collection": settings.qdrant_collection_name,
            },
        )

        # Store metadata in PostgreSQL for structured filtering (Story 2.6 AC4)
        # Maintains parity with PDF ingestion path
        records_stored, records_skipped = await store_metadata_in_postgresql(chunks_with_embeddings)
        logger.info(
            "Metadata stored in PostgreSQL",
            extra={
                "doc_filename": excel_path.name,
                "records_stored": records_stored,
                "records_skipped": records_skipped,
            },
        )

    return chunks_with_embeddings


async def extract_excel(file_path: str) -> DocumentMetadata:
    """Extract financial data from Excel spreadsheet with multi-sheet support.

    Uses openpyxl for Excel parsing and pandas for data manipulation.
    Extracts all sheets preserving numeric formatting and sheet numbers for citations.

    Note: This function is marked async for consistency with the ingestion pipeline
    pattern (ingest_pdf, future chunking/embedding operations). While openpyxl and
    pandas operations are currently synchronous, this allows for future async
    enhancements like streaming large files or parallel sheet processing.

    Args:
        file_path: Path to Excel file (relative or absolute, .xlsx or .xls)

    Returns:
        DocumentMetadata with extraction results including sheet_count as page_count

    Raises:
        FileNotFoundError: If Excel file doesn't exist at specified path
        RuntimeError: If Excel parsing fails, file is password-protected, or corrupted

    Example:
        >>> metadata = await extract_excel("data/financial_report.xlsx")
        >>> print(f"Extracted {metadata.page_count} sheets")
    """
    start_time = time.time()

    # Resolve file path
    excel_path = Path(file_path).resolve()

    # Validate file exists
    _validate_excel_file(excel_path)

    # Load Excel workbook
    workbook = _load_excel_workbook(excel_path)

    # Check for empty workbook
    if not workbook.sheetnames:
        return _create_empty_metadata(excel_path)

    # Extract all sheets with sheet numbers
    sheets_data, skipped_sheets = _extract_all_sheets(workbook, excel_path)

    # Calculate extraction metrics
    sheet_count = len(sheets_data)
    total_rows = sum(sheet["row_count"] for sheet in sheets_data)

    # Validate sheet extraction
    if sheet_count == 0:
        logger.warning(
            "No sheets extracted - verify Excel file structure",
            extra={"path": str(excel_path), "total_sheets": len(workbook.sheetnames)},
        )

    # Concatenate all sheet markdown for chunking
    full_text = "\n\n".join(sheet["content"] for sheet in sheets_data)

    # Process chunks, embeddings, and storage
    chunks_with_embeddings = await _process_excel_chunks(full_text, sheets_data, excel_path)

    # Create metadata with final chunk count
    metadata = DocumentMetadata(
        filename=excel_path.name,
        doc_type="Excel",
        ingestion_timestamp=datetime.now(UTC).isoformat(),
        page_count=sheet_count,
        source_path=str(excel_path),
        chunk_count=len(chunks_with_embeddings),
    )

    # Calculate final metrics
    duration_ms = int((time.time() - start_time) * 1000)

    logger.info(
        "Excel extracted successfully",
        extra={
            "doc_filename": excel_path.name,
            "sheet_count": sheet_count,
            "chunk_count": len(chunks_with_embeddings),
            "total_rows": total_rows,
            "skipped_sheets": skipped_sheets,
            "duration_ms": duration_ms,
            "sheets_per_second": (
                round(sheet_count / (duration_ms / 1000), 2) if duration_ms > 0 else 0
            ),
        },
    )

    return metadata
