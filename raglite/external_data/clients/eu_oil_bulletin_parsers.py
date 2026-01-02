"""EU Oil Bulletin XLSX parsing utilities.

Story 6.9.4: Parse XLSX format for historical diesel prices.
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from raglite.external_data.exceptions import ExternalDataFetchError
from raglite.external_data.models import EUDieselPrice
from raglite.shared.logging import get_logger

logger = get_logger(__name__)


def parse_xlsx(
    content: bytes,
    country: str,
    start_date: date,
    end_date: date,
    country_codes: dict[str, str],
    tax_included: bool = True,
) -> list[EUDieselPrice]:
    """Parse EU Oil Bulletin XLSX data.

    Story 6.9.4 AC2: XLSX parsing with openpyxl

    The XLSX file has multiple sheets:
    - "Prices with taxes" - consumer prices including taxes
    - "Prices wo taxes, per CTR" - prices without taxes by country

    Each sheet has dates in column A and country prices in subsequent columns.

    Args:
        content: XLSX file bytes
        country: Country name
        start_date: Filter start date
        end_date: Filter end date
        country_codes: Mapping of country names to 2-letter codes
        tax_included: Which sheet to use

    Returns:
        List of diesel price records
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ExternalDataFetchError(
            source="EU_Oil_Bulletin",
            message="openpyxl not installed - required for XLSX parsing",
            original_error=e,
        ) from e

    results: list[EUDieselPrice] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        logger.warning(
            "Failed to open EU Oil Bulletin XLSX",
            extra={"error": str(e)},
        )
        return results

    # Try to find the sheet (name may vary slightly)
    ws = None
    for name in wb.sheetnames:
        if "taxes" in name.lower():
            if tax_included and "with" in name.lower():
                ws = wb[name]
                break
            elif not tax_included and ("wo" in name.lower() or "without" in name.lower()):
                ws = wb[name]
                break

    if ws is None:
        # Fallback to first sheet
        ws = wb.active
        if ws is None:
            logger.warning("No active sheet in EU Oil Bulletin XLSX")
            return results

    # Find country diesel column
    country_col_idx = _find_country_column(ws, country, country_codes)
    if country_col_idx is None:
        wb.close()
        return results

    # Detect data start row
    data_start_row = _detect_data_start_row(ws)

    # Parse data rows
    results = _parse_data_rows(
        ws,
        data_start_row,
        country_col_idx,
        country,
        start_date,
        end_date,
        tax_included,
    )

    wb.close()

    # Sort by date (oldest first)
    results.sort(key=lambda x: x.date)

    return results


def _find_country_column(
    ws: Any,
    country: str,
    country_codes: dict[str, str],
) -> int | None:
    """Find the column index for the country diesel prices.

    Story 6.9.5: Header format is {CC}_price_with_tax_diesel where CC is 2-letter country code.

    Args:
        ws: Worksheet object
        country: Country name
        country_codes: Mapping of country names to 2-letter codes

    Returns:
        Column index or None if not found
    """
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Get country code
    country_code = country_codes.get(country, "")
    target_pattern = f"{country_code}_price_with_tax_diesel".lower() if country_code else None

    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        cell_str = str(cell_value).strip().lower()

        # Priority 1: Exact match on country code diesel column (e.g., "pt_price_with_tax_diesel")
        if target_pattern and target_pattern in cell_str:
            return idx

        # Priority 2: Partial match on country name with diesel
        if country.lower() in cell_str and "diesel" in cell_str:
            return idx

        # Priority 3: Exact country name match (for simple headers like "Portugal")
        if cell_str == country.lower():
            return idx

    logger.warning(
        "Country not found in EU Oil Bulletin XLSX headers",
        extra={"country": country, "headers": header_row[:15]},
    )
    return None


def _detect_data_start_row(ws: Any) -> int:
    """Detect which row the data starts on.

    Production files have 3 header rows (start at row 4).
    Test mocks may have 1 header row (start at row 2).

    Args:
        ws: Worksheet object

    Returns:
        Row number to start parsing from (1-indexed)
    """
    data_start_row = 4  # Default for production files
    try:
        row_2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        first_cell = row_2[0] if row_2 else None
        if isinstance(first_cell, (datetime, date)):
            data_start_row = 2  # Test mock format with single header
    except (IndexError, StopIteration):
        pass
    return data_start_row


def _parse_data_rows(
    ws: Any,
    data_start_row: int,
    country_col_idx: int,
    country: str,
    start_date: date,
    end_date: date,
    tax_included: bool,
) -> list[EUDieselPrice]:
    """Parse data rows from worksheet.

    Args:
        ws: Worksheet object
        data_start_row: Row to start parsing from
        country_col_idx: Column index for country prices
        country: Country name
        start_date: Filter start date
        end_date: Filter end date
        tax_included: Tax inclusion flag

    Returns:
        List of diesel price records
    """
    results: list[EUDieselPrice] = []

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        try:
            # First column is date
            date_cell = row[0]
            if date_cell is None:
                continue

            # Parse date
            record_date = _parse_date_cell(date_cell)
            if record_date is None:
                continue

            # Filter by date range
            if not (start_date <= record_date <= end_date):
                continue

            # Get price value
            if country_col_idx >= len(row):
                continue
            price_cell = row[country_col_idx]
            if price_cell is None:
                continue

            price = float(price_cell)

            # Story 6.9.5: Prices in XLSX are in cents (e.g., 1604 = €1.604/L)
            # Convert to EUR per litre
            if price > 100:  # Clearly in cents
                price = price / 1000.0

            # Skip invalid prices (reasonable range for EUR/litre is 0.5 - 3.0)
            if price <= 0.3 or price > 5.0:
                continue

            results.append(
                EUDieselPrice(
                    date=record_date,
                    price_eur_litre=price,
                    country=country,
                    tax_included=tax_included,
                )
            )

        except (ValueError, TypeError, IndexError) as e:
            logger.warning(
                "Failed to parse EU Oil Bulletin row",
                extra={"error": str(e)},
            )
            continue

    return results


def _parse_date_cell(date_cell: Any) -> date | None:
    """Parse a date cell value.

    Args:
        date_cell: Cell value (datetime, date, or string)

    Returns:
        Parsed date or None if invalid
    """
    if isinstance(date_cell, datetime):
        return date_cell.date()
    elif isinstance(date_cell, date):
        return date_cell
    elif isinstance(date_cell, str):
        # Try common date formats
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
            try:
                return datetime.strptime(date_cell, fmt).date()
            except ValueError:
                continue
    return None
