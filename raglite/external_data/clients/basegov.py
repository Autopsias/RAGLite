"""Base.gov.pt (Portuguese Public Procurement) client.

Story 6.1: Tier 1 External Data Source Integration
Story 6.9.5: BaseGov Public Procurement Fix

Fetches Portuguese public works contract data:
- Contract values
- Contracting entities
- Contractors
- CPV codes

Data Sources (Story 6.9.5 - verified 2025-12-08, updated after user feedback):

PRIMARY: dados.gov.pt IMPIC XLSX Dataset
- URL: https://dados.gov.pt/pt/datasets/contratos-publicos-portal-base-impic-contratos-de-2012-a-2025/
- Format: Yearly XLSX files (contratos2012.xlsx ... contratos2025.xlsx)
- Coverage: ALL Portuguese public contracts (not just EU-threshold)
- Updated: Daily/Weekly by IMPIC
- Last verified: 2025-12-08 (28 resources, updated 2025-12-07)

FALLBACK: TED API v3
- URL: https://tedweb.api.ted.europa.eu/v3
- Coverage: Only contracts above EU thresholds (~EUR 140K+ services, ~EUR 5.4M works)
- Used when: IMPIC dataset unavailable

DEPRECATED:
- dados.gov.pt OCDS dataset - Empty resources array as of 2025-12-08
- Base.gov.pt JSON API - Does NOT exist (HTML only)
"""

from __future__ import annotations

import asyncio
import os
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

import httpx
import openpyxl

from raglite.external_data.exceptions import ExternalDataFetchError
from raglite.external_data.models import BaseGovContract
from raglite.shared.config import settings
from raglite.shared.logging import get_logger

logger = get_logger(__name__)

# API Configuration
# Story 6.9.5: Multiple data sources with fallback

# PRIMARY: dados.gov.pt IMPIC XLSX dataset (ALL Portuguese contracts)
DADOS_GOV_API_BASE = "https://dados.gov.pt/api/1"
IMPIC_CONTRACTS_DATASET = "contratos-publicos-portal-base-impic-contratos-de-2012-a-2025"

# FALLBACK: TED API v3 (EU public procurement, Portugal contracts above EU thresholds)
TED_API_BASE = "https://tedweb.api.ted.europa.eu/v3"

# DEPRECATED: dados.gov.pt OCDS dataset (empty resources as of 2025-12-08)
OCDS_DATASET_ID = "ocds-portal-base-www-base-gov-pt"

# Deprecated: Base.gov.pt (NO public API - HTML only)
BASEGOV_API_BASE = "https://www.base.gov.pt/Base4/pt/pesquisa"  # Does NOT work


class BaseGovClient:
    """Client for Portuguese public procurement data.

    Story 6.9.5: Updated to use dados.gov.pt IMPIC XLSX dataset as primary source.

    Data Sources (priority order):
    1. dados.gov.pt IMPIC XLSX - ALL Portuguese contracts (2012-2025)
    2. TED API v3 - Fallback for EU-threshold contracts only

    The IMPIC dataset provides comprehensive coverage of all Portuguese public
    contracts, including those below EU thresholds. Files are organized by year
    (contratos2012.xlsx through contratos2025.xlsx).

    Example:
        >>> client = BaseGovClient()
        >>> contracts = await client.fetch_contracts(
        ...     start_date=date(2024, 1, 1),
        ...     end_date=date(2024, 3, 31),
        ...     cpv_code="45000000"  # Construction works
        ... )
    """

    # CPV codes for construction-related contracts
    CPV_CONSTRUCTION = "45000000"  # Construction works
    CPV_BUILDING = "45210000"  # Building construction
    CPV_CIVIL_ENGINEERING = "45220000"  # Civil engineering
    CPV_ROAD = "45233000"  # Highway construction

    # EU procurement thresholds (approximate, as of 2024)
    # Contracts below these thresholds are NOT in TED (but ARE in IMPIC dataset)
    EU_THRESHOLD_WORKS = 5_382_000  # EUR for works
    EU_THRESHOLD_SUPPLIES = 221_000  # EUR for supplies/services (central govt)
    EU_THRESHOLD_SERVICES = 221_000  # EUR for services

    # Cache configuration
    CACHE_DIR = Path(".cache/external_data")
    CACHE_TTL_HOURS = 24  # XLSX files are updated daily/weekly

    def __init__(self) -> None:
        self.ted_api_base = TED_API_BASE
        self.dados_gov_base = DADOS_GOV_API_BASE
        self.impic_dataset = IMPIC_CONTRACTS_DATASET
        self.ocds_dataset_id = OCDS_DATASET_ID

        # Story 6.9.5 AC8: Test-aware timeout
        is_test = os.getenv("PYTEST_CURRENT_TEST") is not None
        self.timeout = 1.0 if is_test else float(settings.external_data_timeout)

        # Cache directory
        self.cache_dir = self.CACHE_DIR
        if not is_test:
            self.cache_dir.mkdir(parents=True, exist_ok=True)

    async def _fetch_ted_notices(
        self,
        start_date: date,
        end_date: date,
        cpv_code: str | None = None,
        page: int = 1,
        limit: int = 100,
    ) -> dict:
        """Fetch notices from TED API v3.

        Story 6.9.5 AC2/AC6: TED API for EU-threshold contracts

        Args:
            start_date: Start of date range
            end_date: End of date range
            cpv_code: CPV code filter
            page: Page number
            limit: Results per page

        Returns:
            TED API response dict

        Raises:
            ExternalDataFetchError: If all retries fail
        """
        max_retries = settings.external_data_retry_attempts
        # Story 6.9.5 AC8: NFR1 exponential backoff at 2s/4s/8s intervals
        retry_delays = [2, 4, 8]

        url = f"{self.ted_api_base}/notices/search"

        # Build TED query
        # TED uses its own query language
        query_parts = [
            "(place-of-performance = PT)",  # Portugal
            f"(publication-date >= {start_date.isoformat()})",
            f"(publication-date <= {end_date.isoformat()})",
        ]

        if cpv_code:
            query_parts.append(f"(cpv = {cpv_code}*)")

        query = " AND ".join(query_parts)

        payload = {
            "query": query,
            "fields": [
                "publication-number",
                "publication-date",
                "notice-title",
                "buyer-name",
                "winner-name",
                "total-value",
                "cpv",
                "place-of-performance",
                "contract-nature",
            ],
            "page": page,
            "limit": limit,
            "scope": "ALL",  # Include historical and active
            "onlyLatestVersions": True,
        }

        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
        }

        async with httpx.AsyncClient(timeout=self.timeout) as client:
            for attempt in range(max_retries):
                try:
                    logger.info(
                        "Fetching TED notices",
                        extra={
                            "query": query[:100],
                            "page": page,
                            "attempt": attempt + 1,
                        },
                    )

                    response = await client.post(
                        url,
                        json=payload,
                        headers=headers,
                    )
                    response.raise_for_status()
                    return dict(response.json())  # type: ignore[no-any-return]

                except httpx.TimeoutException as e:
                    if attempt < max_retries - 1:
                        delay = retry_delays[attempt]
                        logger.warning(
                            "TED API timeout, retrying",
                            extra={"attempt": attempt + 1, "delay": delay},
                        )
                        await asyncio.sleep(delay)
                    else:
                        raise ExternalDataFetchError(
                            source="BaseGov_TED",
                            message=f"TED API timeout after {max_retries} attempts",
                            original_error=e,
                        ) from e

                except httpx.HTTPStatusError as e:
                    # Retry on server errors (5xx) or rate limit (429)
                    should_retry = e.response.status_code >= 500 or e.response.status_code == 429
                    if attempt < max_retries - 1 and should_retry:
                        delay = retry_delays[attempt]
                        logger.warning(
                            "TED API error, retrying",
                            extra={
                                "attempt": attempt + 1,
                                "status": e.response.status_code,
                            },
                        )
                        await asyncio.sleep(delay)
                    else:
                        raise ExternalDataFetchError(
                            source="BaseGov_TED",
                            message=f"TED API HTTP {e.response.status_code}",
                            original_error=e,
                        ) from e

        raise ExternalDataFetchError(
            source="BaseGov_TED",
            message="Unexpected retry loop exit",
        )

    async def _get_impic_resource_urls(self) -> dict[int, str]:
        """Get URLs for IMPIC yearly XLSX files.

        Story 6.9.5: Primary data source - dados.gov.pt IMPIC dataset

        Returns:
            Dict mapping year to XLSX URL, e.g. {2024: "https://...contratos2024.xlsx"}
        """
        url = f"{self.dados_gov_base}/datasets/{self.impic_dataset}/"

        try:
            async with httpx.AsyncClient(timeout=self.timeout) as client:
                response = await client.get(url)
                response.raise_for_status()
                data = response.json()

                year_urls: dict[int, str] = {}
                for resource in data.get("resources", []):
                    resource_url = resource.get("url", "")
                    title = resource.get("title", "")

                    # Extract year from filename like "contratos2024.xlsx"
                    if title.startswith("contratos") and title.endswith(".xlsx"):
                        try:
                            year = int(title.replace("contratos", "").replace(".xlsx", ""))
                            year_urls[year] = resource_url
                        except ValueError:
                            continue

                logger.info(
                    "Found IMPIC XLSX resources",
                    extra={"years": sorted(year_urls.keys()), "count": len(year_urls)},
                )
                return year_urls

        except (httpx.HTTPError, httpx.TimeoutException) as e:
            logger.warning(
                "Failed to get IMPIC resource URLs",
                extra={"error": str(e)},
            )
            return {}

    def _get_cached_impic_xlsx(self, year: int) -> bytes | None:
        """Get cached IMPIC XLSX file if valid.

        Args:
            year: Contract year

        Returns:
            XLSX bytes if cached and valid, None otherwise
        """
        cache_file = self.cache_dir / f"impic_contratos{year}.xlsx"

        if not cache_file.exists():
            return None

        # Check cache age
        age = datetime.now() - datetime.fromtimestamp(cache_file.stat().st_mtime)
        if age > timedelta(hours=self.CACHE_TTL_HOURS):
            logger.debug(f"IMPIC cache expired for year {year}")
            return None

        return cache_file.read_bytes()

    def _save_impic_xlsx_cache(self, year: int, content: bytes) -> None:
        """Save IMPIC XLSX to cache.

        Args:
            year: Contract year
            content: XLSX bytes
        """
        try:
            cache_file = self.cache_dir / f"impic_contratos{year}.xlsx"
            cache_file.write_bytes(content)
            logger.debug(f"Cached IMPIC XLSX for year {year}")
        except OSError as e:
            logger.warning(f"Failed to cache IMPIC XLSX: {e}")

    async def _fetch_impic_xlsx(self, year: int, url: str) -> bytes | None:
        """Download IMPIC XLSX file for a specific year.

        Story 6.9.5: Fetches yearly contract data from dados.gov.pt

        Args:
            year: Contract year
            url: URL to XLSX file

        Returns:
            XLSX bytes or None if fetch failed
        """
        # Check cache first
        cached = self._get_cached_impic_xlsx(year)
        if cached:
            logger.info(f"Using cached IMPIC XLSX for {year}")
            return cached

        max_retries = settings.external_data_retry_attempts
        retry_delays = [2, 4, 8]

        async with httpx.AsyncClient(timeout=60.0) as client:  # Larger timeout for XLSX
            for attempt in range(max_retries):
                try:
                    logger.info(
                        f"Downloading IMPIC XLSX for {year}",
                        extra={"url": url[:80], "attempt": attempt + 1},
                    )

                    response = await client.get(url, follow_redirects=True)
                    response.raise_for_status()

                    content = response.content
                    self._save_impic_xlsx_cache(year, content)
                    return content

                except httpx.TimeoutException:
                    if attempt < max_retries - 1:
                        await asyncio.sleep(retry_delays[attempt])
                    else:
                        logger.warning(f"IMPIC XLSX download timeout for {year}")
                        return None

                except httpx.HTTPStatusError as e:
                    if attempt < max_retries - 1 and e.response.status_code >= 500:
                        await asyncio.sleep(retry_delays[attempt])
                    else:
                        logger.warning(
                            f"IMPIC XLSX download failed for {year}: HTTP {e.response.status_code}"
                        )
                        return None

        return None

    def _parse_impic_xlsx(
        self,
        content: bytes,
        start_date: date,
        end_date: date,
        cpv_code: str | None = None,
    ) -> list[BaseGovContract]:
        """Parse IMPIC XLSX content to BaseGovContract models.

        Story 6.9.5: Parse dados.gov.pt IMPIC format

        XLSX columns (verified 2025-12-08):
        - idcontrato: Contract ID
        - tipoContrato: Contract type
        - tipoprocedimento: Procedure type
        - objectoContrato: Contract object/title
        - descContrato: Description
        - adjudicante: Contracting entity (NIF - Name)
        - adjudicatarios: Contractor/winner (NIF - Name)
        - dataPublicacao: Publication date
        - dataCelebracaoContrato: Contract celebration date
        - precoContratual: Contract value
        - CPV: CPV codes

        Args:
            content: XLSX bytes
            start_date: Filter start date
            end_date: Filter end date
            cpv_code: CPV code filter (prefix match)

        Returns:
            List of contracts matching filters
        """
        results: list[BaseGovContract] = []

        try:
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # Find column indices
            col_map = {h: i for i, h in enumerate(headers) if h}

            id_col = col_map.get("idcontrato", 0)
            obj_col = col_map.get("objectoContrato")
            desc_col = col_map.get("descContrato")
            buyer_col = col_map.get("adjudicante")
            winner_col = col_map.get("adjudicatarios")
            pub_date_col = col_map.get("dataPublicacao")
            contract_date_col = col_map.get("dataCelebracaoContrato")
            value_col = col_map.get("precoContratual")
            cpv_col = col_map.get("CPV")

            # Parse data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    # Get publication date
                    pub_date_val = row[pub_date_col] if pub_date_col is not None else None
                    if not pub_date_val:
                        pub_date_val = (
                            row[contract_date_col] if contract_date_col is not None else None
                        )

                    if not pub_date_val:
                        continue

                    # Parse date
                    if isinstance(pub_date_val, datetime):
                        pub_date = pub_date_val.date()
                    elif isinstance(pub_date_val, date):
                        pub_date = pub_date_val
                    elif isinstance(pub_date_val, str):
                        pub_date = date.fromisoformat(pub_date_val[:10])
                    else:
                        continue

                    # Filter by date range
                    if not (start_date <= pub_date <= end_date):
                        continue

                    # Get CPV code
                    item_cpv = str(row[cpv_col]) if cpv_col is not None and row[cpv_col] else ""

                    # Filter by CPV if specified (prefix match)
                    if cpv_code and item_cpv:
                        # Extract main CPV code (before dash)
                        main_cpv = item_cpv.split("-")[0].split()[0] if item_cpv else ""
                        if not main_cpv.startswith(cpv_code[:2]):
                            continue

                    # Get contract value
                    value = 0.0
                    if value_col is not None and row[value_col]:
                        try:
                            val = row[value_col]
                            if isinstance(val, (int, float)):
                                value = float(val)
                            elif isinstance(val, str):
                                value = float(val.replace(",", ".").replace(" ", ""))
                        except (ValueError, TypeError):
                            value = 0.0

                    # Get description
                    description = ""
                    if obj_col is not None and row[obj_col]:
                        description = str(row[obj_col])[:500]
                    elif desc_col is not None and row[desc_col]:
                        description = str(row[desc_col])[:500]

                    # Parse buyer/winner (format: "NIF - Name")
                    buyer = str(row[buyer_col]) if buyer_col is not None and row[buyer_col] else ""
                    winner = (
                        str(row[winner_col]) if winner_col is not None and row[winner_col] else ""
                    )

                    results.append(
                        BaseGovContract(
                            publication_date=pub_date,
                            contract_id=str(row[id_col]) if row[id_col] else "",
                            description=description,
                            contract_value_eur=value,
                            contracting_entity=buyer,
                            contractor=winner,
                            cpv_code=item_cpv.split("\n")[0] if item_cpv else "",  # First CPV only
                            execution_location="Portugal",
                        )
                    )

                except (IndexError, TypeError, ValueError):
                    # Skip malformed rows
                    continue

            wb.close()

        except Exception as e:
            logger.warning(f"Failed to parse IMPIC XLSX: {e}")
            return []

        return results

    async def _fetch_impic_contracts(
        self,
        start_date: date,
        end_date: date,
        cpv_code: str | None = None,
    ) -> list[BaseGovContract]:
        """Fetch contracts from dados.gov.pt IMPIC XLSX dataset.

        Story 6.9.5: Primary data source for Portuguese public contracts

        Args:
            start_date: Start of date range
            end_date: End of date range
            cpv_code: CPV code filter

        Returns:
            List of contracts from IMPIC dataset
        """
        # Get available yearly files
        year_urls = await self._get_impic_resource_urls()

        if not year_urls:
            logger.warning("IMPIC dataset not available")
            return []

        # Determine which years to fetch
        years_needed = set(range(start_date.year, end_date.year + 1))
        years_available = set(year_urls.keys())
        years_to_fetch = years_needed & years_available

        if not years_to_fetch:
            logger.warning(
                "No IMPIC data for requested period",
                extra={
                    "years_needed": list(years_needed),
                    "years_available": list(years_available),
                },
            )
            return []

        results: list[BaseGovContract] = []

        for year in sorted(years_to_fetch):
            url = year_urls[year]
            content = await self._fetch_impic_xlsx(year, url)

            if content:
                parsed = self._parse_impic_xlsx(content, start_date, end_date, cpv_code)
                results.extend(parsed)
                logger.info(f"Parsed {len(parsed)} contracts from IMPIC {year}")

        return results

    async def _check_ocds_availability(self) -> dict | None:
        """Check if dados.gov.pt OCDS dataset has resources.

        Story 6.9.5 AC1/AC2: Check OCDS dataset availability

        Returns:
            Dataset metadata if resources available, None otherwise
        """
        url = f"{self.dados_gov_base}/datasets/{self.ocds_dataset_id}/"

        try:
            async with httpx.AsyncClient(timeout=self.timeout) as client:
                response = await client.get(url)
                response.raise_for_status()
                data = response.json()

                resources = data.get("resources", [])
                if not resources:
                    logger.warning(
                        "dados.gov.pt OCDS dataset has no resources",
                        extra={"dataset_id": self.ocds_dataset_id},
                    )
                    return None

                return data

        except (httpx.HTTPError, httpx.TimeoutException) as e:
            logger.warning(
                "Failed to check dados.gov.pt OCDS dataset",
                extra={"error": str(e)},
            )
            return None

    async def _fetch_ocds_data(
        self,
        start_date: date,
        end_date: date,
        cpv_code: str | None = None,
    ) -> list[BaseGovContract]:
        """Fetch contracts from dados.gov.pt OCDS dataset.

        Story 6.9.5 AC2/AC3: OCDS data fetching (when available)

        Currently returns empty list as dataset has no resources.
        Implementation ready for when IMPIC restores the data.

        Args:
            start_date: Start of date range
            end_date: End of date range
            cpv_code: CPV code filter

        Returns:
            List of contracts (empty if dataset unavailable)
        """
        dataset = await self._check_ocds_availability()

        if dataset is None:
            return []

        results: list[BaseGovContract] = []

        # When dataset becomes available, fetch and parse OCDS JSON/CSV
        for resource in dataset.get("resources", []):
            resource_format = resource.get("format", "").lower()
            resource_url = resource.get("url")

            if not resource_url:
                continue

            # Prefer JSON format
            if resource_format == "json":
                try:
                    async with httpx.AsyncClient(timeout=60.0) as client:
                        response = await client.get(resource_url)
                        response.raise_for_status()
                        ocds_data = response.json()

                        parsed = self._parse_ocds_data(ocds_data, start_date, end_date, cpv_code)
                        results.extend(parsed)
                        break  # Got data from JSON, no need for CSV

                except (httpx.HTTPError, ValueError) as e:
                    logger.warning(
                        "Failed to fetch OCDS JSON",
                        extra={"url": resource_url, "error": str(e)},
                    )

        return results

    def _parse_ocds_data(
        self,
        ocds_data: dict | list,
        start_date: date,
        end_date: date,
        cpv_code: str | None = None,
    ) -> list[BaseGovContract]:
        """Parse OCDS format data to BaseGovContract models.

        Story 6.9.5 AC3: OCDS parsing

        OCDS structure:
        - releases[] or records[] array
        - Each has tender, awards, contracts sections
        - Uses ocid as unique identifier

        Args:
            ocds_data: OCDS JSON data
            start_date: Filter start date
            end_date: Filter end date
            cpv_code: CPV code filter

        Returns:
            List of BaseGovContract records
        """
        results: list[BaseGovContract] = []

        # Handle both releases and records format
        items = []
        if isinstance(ocds_data, dict):
            items = ocds_data.get("releases", ocds_data.get("records", []))
        elif isinstance(ocds_data, list):
            items = ocds_data

        for item in items:
            try:
                # Get contract info from awards/contracts section
                awards = item.get("awards", [])
                contracts_list = item.get("contracts", [])
                tender = item.get("tender", {})

                # Extract date from published date or tender period
                pub_date_str = item.get("date", item.get("publishedDate"))
                if not pub_date_str and tender:
                    pub_date_str = tender.get("tenderPeriod", {}).get("endDate")

                if not pub_date_str:
                    continue

                # Parse date
                pub_date = date.fromisoformat(pub_date_str[:10])

                # Filter by date range
                if not (start_date <= pub_date <= end_date):
                    continue

                # Get CPV codes
                item_cpv = None
                tender_items = tender.get("items", [])
                if tender_items:
                    classification = tender_items[0].get("classification", {})
                    item_cpv = classification.get("id", "")

                # Filter by CPV if specified
                if cpv_code and item_cpv and not item_cpv.startswith(cpv_code[:2]):
                    continue

                # Get contract value
                value = 0.0
                if contracts_list:
                    value = contracts_list[0].get("value", {}).get("amount", 0)
                elif awards:
                    value = awards[0].get("value", {}).get("amount", 0)
                elif tender:
                    value = tender.get("value", {}).get("amount", 0)

                # Get parties (buyer and supplier)
                parties = item.get("parties", [])
                buyer = ""
                supplier = ""
                for party in parties:
                    roles = party.get("roles", [])
                    if "buyer" in roles:
                        buyer = party.get("name", "")
                    if "supplier" in roles or "tenderer" in roles:
                        supplier = party.get("name", "")

                results.append(
                    BaseGovContract(
                        publication_date=pub_date,
                        contract_id=item.get("ocid", item.get("id", "")),
                        description=tender.get("title", tender.get("description", "")),
                        contract_value_eur=float(value),
                        contracting_entity=buyer,
                        contractor=supplier,
                        cpv_code=item_cpv,
                        execution_location="Portugal",
                    )
                )

            except (ValueError, KeyError, TypeError) as e:
                logger.warning(
                    "Failed to parse OCDS item",
                    extra={"error": str(e)},
                )
                continue

        return results

    def _parse_ted_notices(self, data: dict) -> list[BaseGovContract]:
        """Parse TED API response to BaseGovContract models.

        Story 6.9.5 AC6: TED API parsing

        Args:
            data: TED API response

        Returns:
            List of contracts
        """
        results: list[BaseGovContract] = []

        notices = data.get("notices", data.get("results", []))

        for notice in notices:
            try:
                # Get publication date
                pub_date_str = notice.get("publication-date", notice.get("publicationDate"))
                if not pub_date_str:
                    continue

                pub_date = date.fromisoformat(pub_date_str[:10])

                # Get value (may be in different formats)
                value = notice.get("total-value", notice.get("totalValue", 0))
                if isinstance(value, dict):
                    value = value.get("amount", 0)
                if isinstance(value, str):
                    value = float(value.replace(",", ".").replace(" ", ""))

                # Get CPV code
                cpv = notice.get("cpv", notice.get("cpvCode", ""))
                if isinstance(cpv, list):
                    cpv = cpv[0] if cpv else ""

                results.append(
                    BaseGovContract(
                        publication_date=pub_date,
                        contract_id=str(notice.get("publication-number", notice.get("id", ""))),
                        description=notice.get("notice-title", notice.get("title", "")),
                        contract_value_eur=float(value) if value else 0.0,
                        contracting_entity=notice.get("buyer-name", notice.get("buyer", "")),
                        contractor=notice.get("winner-name", notice.get("winner", "")),
                        cpv_code=cpv,
                        execution_location="Portugal",
                    )
                )

            except (ValueError, KeyError, TypeError) as e:
                logger.warning(
                    "Failed to parse TED notice",
                    extra={"error": str(e)},
                )
                continue

        return results

    async def fetch_contracts(
        self,
        start_date: date,
        end_date: date,
        cpv_code: str | None = None,
        min_value: float | None = None,
        max_value: float | None = None,
        page: int = 1,
        page_size: int = 100,
    ) -> list[BaseGovContract]:
        """Fetch public works contracts.

        Story 6.9.5 AC4: Updated with IMPIC XLSX as primary source

        Data sources (in priority order):
        1. dados.gov.pt IMPIC XLSX - ALL Portuguese contracts (2012-2025)
        2. TED API v3 - Fallback for EU-threshold contracts only

        The IMPIC dataset includes ALL Portuguese public contracts, not just
        those above EU thresholds. This provides much better coverage than TED.

        Args:
            start_date: Start of date range (publication date)
            end_date: End of date range
            cpv_code: CPV code filter (default: construction works)
            min_value: Minimum contract value in EUR
            max_value: Maximum contract value in EUR
            page: Page number for pagination (used for TED fallback)
            page_size: Results per page (used for TED fallback)

        Returns:
            List of contract records
        """
        if cpv_code is None:
            cpv_code = self.CPV_CONSTRUCTION

        logger.info(
            "Fetching public procurement contracts",
            extra={
                "start": str(start_date),
                "end": str(end_date),
                "cpv": cpv_code,
            },
        )

        results: list[BaseGovContract] = []

        # Try IMPIC XLSX first (primary source - ALL Portuguese contracts)
        try:
            impic_results = await self._fetch_impic_contracts(
                start_date=start_date,
                end_date=end_date,
                cpv_code=cpv_code,
            )
            results.extend(impic_results)
            logger.info(
                "Fetched contracts from IMPIC XLSX",
                extra={"count": len(results)},
            )
        except Exception as e:
            logger.warning(
                "IMPIC dataset unavailable, trying TED API fallback",
                extra={"error": str(e)},
            )

        # Try TED API as fallback if IMPIC returned no results
        if not results:
            try:
                ted_data = await self._fetch_ted_notices(
                    start_date=start_date,
                    end_date=end_date,
                    cpv_code=cpv_code,
                    page=page,
                    limit=page_size,
                )
                results.extend(self._parse_ted_notices(ted_data))
                logger.info(
                    "Fetched contracts from TED API (fallback)",
                    extra={"count": len(results)},
                )
            except ExternalDataFetchError as e:
                logger.warning(
                    "TED API also unavailable",
                    extra={"error": str(e)},
                )

        # Apply value filters if specified
        if min_value is not None:
            results = [r for r in results if r.contract_value_eur >= min_value]
        if max_value is not None:
            results = [r for r in results if r.contract_value_eur <= max_value]

        # Story 6.9.5 AC7: Document limitations
        if not results:
            logger.warning(
                "No contracts found - Check date range and CPV code. "
                "IMPIC dataset covers 2012-2025, TED only includes EU-threshold contracts.",
                extra={
                    "start": str(start_date),
                    "end": str(end_date),
                    "cpv": cpv_code,
                },
            )

        logger.info(
            "Fetched public procurement contracts",
            extra={"record_count": len(results), "source": "IMPIC/TED"},
        )
        return results

    async def fetch_all_contracts(
        self,
        start_date: date,
        end_date: date,
        cpv_code: str | None = None,
    ) -> list[BaseGovContract]:
        """Fetch all contracts in date range (handles pagination).

        Args:
            start_date: Start of date range
            end_date: End of date range
            cpv_code: CPV code filter

        Returns:
            List of all contract records
        """
        all_results: list[BaseGovContract] = []
        page = 1
        page_size = 100

        while True:
            results = await self.fetch_contracts(
                start_date=start_date,
                end_date=end_date,
                cpv_code=cpv_code,
                page=page,
                page_size=page_size,
            )

            if not results:
                break

            all_results.extend(results)

            if len(results) < page_size:
                break

            page += 1

            # Safety limit
            if page > 100:
                logger.warning("Pagination limit reached (100 pages)")
                break

        return all_results

    async def fetch_construction_contracts_summary(
        self,
        year: int,
        month: int | None = None,
    ) -> dict:
        """Fetch summary of construction contracts for a period.

        Args:
            year: Year
            month: Month (optional, for monthly summary)

        Returns:
            Summary dict with total_contracts, total_value, avg_value
        """
        if month:
            start = date(year, month, 1)
            if month == 12:
                end = date(year + 1, 1, 1)
            else:
                end = date(year, month + 1, 1)
            end = end.replace(day=1) - timedelta(days=1)
        else:
            start = date(year, 1, 1)
            end = date(year, 12, 31)

        contracts = await self.fetch_all_contracts(
            start_date=start,
            end_date=end,
            cpv_code=self.CPV_CONSTRUCTION,
        )

        total_value = sum(c.contract_value_eur for c in contracts)
        avg_value = total_value / len(contracts) if contracts else 0

        return {
            "period_start": start,
            "period_end": end,
            "total_contracts": len(contracts),
            "total_value_eur": total_value,
            "avg_value_eur": avg_value,
            "data_source": "dados.gov.pt IMPIC (ALL Portuguese contracts)",
            "note": "Includes all contract values, not just EU-threshold",
        }

    # Legacy method - kept for backward compatibility
    async def _fetch_with_retry(
        self,
        params: dict,
    ) -> dict:
        """Deprecated: Fetch from Base.gov.pt API.

        Story 6.9.5: This method is deprecated. Base.gov.pt does NOT have a public API.
        Kept for backward compatibility - always returns empty response.

        Args:
            params: Query parameters (ignored)

        Returns:
            Empty dict (API does not exist)
        """
        logger.warning(
            "Deprecated _fetch_with_retry called - "
            "Base.gov.pt does NOT have a public API, use fetch_contracts() instead"
        )
        return {"items": []}

    def _parse_contracts(self, data: dict) -> list[BaseGovContract]:
        """Deprecated: Parse Base.gov.pt response.

        Story 6.9.5: This method is deprecated. Base.gov.pt does NOT have a public API.
        Kept for backward compatibility.

        Returns:
            Empty list
        """
        logger.warning(
            "Deprecated _parse_contracts called - "
            "use _parse_ted_notices or _parse_ocds_data instead"
        )
        return []
