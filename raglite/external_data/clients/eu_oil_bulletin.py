"""EU Oil Bulletin client.

Story 6.1: Tier 1 External Data Source Integration
Story 6.9.4: EU Oil Bulletin Fix

Fetches EU fuel prices:
- Diesel prices (weekly, by country)
- Gasoline prices

Data Source: https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en

IMPORTANT: Story 6.9.4 - Data source changed on 2025-12-08
Old XML endpoint (ec.europa.eu/energy/observatory/reports/) redirects to homepage (302).
New endpoint serves XLSX files from energy.ec.europa.eu.

Historical data XLSX (~4MB):
https://energy.ec.europa.eu/document/download/906e60ca-8b6a-44e7-8589-652854d2fd3f_en
"""

from __future__ import annotations

import asyncio
import os
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

import httpx

from raglite.external_data.exceptions import ExternalDataFetchError
from raglite.external_data.models import EUDieselPrice
from raglite.shared.config import settings
from raglite.shared.logging import get_logger

logger = get_logger(__name__)

# EU Oil Bulletin API Configuration
# Story 6.9.4 AC1: Updated from old XML endpoint to new XLSX endpoint
EU_OIL_BULLETIN_BASE = "https://energy.ec.europa.eu"

# XLSX document IDs (from energy.ec.europa.eu document download URLs)
# Story 6.9.4 AC1: Historical prices file (~4MB, 2005 onwards)
HISTORY_XLSX_DOC_ID = "906e60ca-8b6a-44e7-8589-652854d2fd3f_en"
HISTORY_XLSX_FILENAME = "Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx"

# Cache settings for large historical file
# Story 6.9.4 AC5: Avoid re-downloading ~4MB file on every call
DEFAULT_CACHE_DIR = Path(".cache/external_data")
DEFAULT_CACHE_TTL_HOURS = 24


class EUOilBulletinClient:
    """Client for EU Oil Bulletin fuel prices.

    Story 6.9.4: Updated to use XLSX format instead of deprecated XML endpoint.

    The EU Oil Bulletin provides weekly fuel prices for all EU member states.
    Data is now served as XLSX files from energy.ec.europa.eu.

    Example:
        >>> client = EUOilBulletinClient()
        >>> prices = await client.fetch_diesel_prices(
        ...     start_date=date(2024, 1, 1),
        ...     end_date=date(2024, 3, 31),
        ...     country="Portugal"
        ... )
    """

    # Country codes and names for EU Oil Bulletin
    # The XLSX file uses full country names in column headers
    COUNTRY_CODES = {
        "Portugal": "PT",
        "Spain": "ES",
        "France": "FR",
        "Germany": "DE",
        "Italy": "IT",
        "Netherlands": "NL",
        "Belgium": "BE",
        "Austria": "AT",
        "Greece": "EL",
        "Ireland": "IE",
        "Poland": "PL",
        "Czech Republic": "CZ",
        "Hungary": "HU",
        "Romania": "RO",
        "Bulgaria": "BG",
    }

    def __init__(
        self,
        cache_dir: Path | None = None,
        cache_ttl_hours: int = DEFAULT_CACHE_TTL_HOURS,
    ) -> None:
        """Initialize EU Oil Bulletin client.

        Args:
            cache_dir: Directory for caching XLSX file (default: .cache/external_data)
            cache_ttl_hours: Cache TTL in hours (default: 24)
        """
        self.base_url = EU_OIL_BULLETIN_BASE
        self.cache_dir = cache_dir or DEFAULT_CACHE_DIR
        self.cache_ttl_hours = cache_ttl_hours

        # Story 6.9.4 AC7: Extended timeout for ~4MB file download (NFR2: 60s)
        is_test = os.getenv("PYTEST_CURRENT_TEST") is not None
        self.timeout = 1.0 if is_test else 60.0  # 60s for large file

        # Lazy-loaded workbook to avoid re-parsing
        self._cached_workbook: Any = None
        self._cached_workbook_time: datetime | None = None

    async def _fetch_xlsx_data(self) -> bytes:
        """Fetch historical oil bulletin XLSX file.

        Story 6.9.4 AC1/AC7: New XLSX download with retry logic

        Returns:
            XLSX file content as bytes

        Raises:
            ExternalDataFetchError: If fetch fails
        """
        max_retries = settings.external_data_retry_attempts
        # Story 6.9.4 AC7: NFR1 exponential backoff at 2s/4s/8s intervals
        retry_delays = [2, 4, 8]

        # Story 6.9.4 AC1: New XLSX download URL
        url = f"{self.base_url}/document/download/{HISTORY_XLSX_DOC_ID}"
        params = {"filename": HISTORY_XLSX_FILENAME}

        async with httpx.AsyncClient(timeout=self.timeout) as client:
            for attempt in range(max_retries):
                try:
                    logger.info(
                        "Downloading EU Oil Bulletin XLSX",
                        extra={"url": url, "attempt": attempt + 1},
                    )
                    response = await client.get(url, params=params)
                    response.raise_for_status()

                    content = response.content
                    # Verify it's actually an XLSX file (starts with PK - ZIP signature)
                    if not content[:2] == b"PK":
                        raise ExternalDataFetchError(
                            source="EU_Oil_Bulletin",
                            message="Downloaded file is not a valid XLSX",
                        )

                    logger.info(
                        "Downloaded EU Oil Bulletin XLSX",
                        extra={"size_bytes": len(content)},
                    )
                    return content

                except httpx.TimeoutException as e:
                    if attempt < max_retries - 1:
                        delay = retry_delays[attempt]
                        logger.warning(
                            "EU Oil Bulletin timeout, retrying",
                            extra={"attempt": attempt + 1, "delay": delay},
                        )
                        await asyncio.sleep(delay)
                    else:
                        raise ExternalDataFetchError(
                            source="EU_Oil_Bulletin",
                            message=f"Timeout after {max_retries} attempts",
                            original_error=e,
                        ) from e

                except httpx.HTTPStatusError as e:
                    # Retry on server errors (5xx) or rate limit (429)
                    should_retry = e.response.status_code >= 500 or e.response.status_code == 429
                    if attempt < max_retries - 1 and should_retry:
                        delay = retry_delays[attempt]
                        logger.warning(
                            "EU Oil Bulletin HTTP error, retrying",
                            extra={"attempt": attempt + 1, "status": e.response.status_code},
                        )
                        await asyncio.sleep(delay)
                    else:
                        raise ExternalDataFetchError(
                            source="EU_Oil_Bulletin",
                            message=f"HTTP {e.response.status_code}",
                            original_error=e,
                        ) from e

        raise ExternalDataFetchError(
            source="EU_Oil_Bulletin",
            message="Unexpected retry loop exit",
        )

    def _get_cached_xlsx(self) -> bytes | None:
        """Get cached XLSX file if valid.

        Story 6.9.4 AC5: Caching for large historical file

        Returns:
            Cached XLSX bytes or None if cache invalid/expired
        """
        cache_file = self.cache_dir / "eu_oil_bulletin_history.xlsx"

        if not cache_file.exists():
            return None

        # Check cache age
        mtime = datetime.fromtimestamp(cache_file.stat().st_mtime)
        age = datetime.now() - mtime

        if age > timedelta(hours=self.cache_ttl_hours):
            logger.info(
                "EU Oil Bulletin cache expired",
                extra={"age_hours": age.total_seconds() / 3600},
            )
            return None

        logger.info(
            "Using cached EU Oil Bulletin XLSX",
            extra={"cache_age_hours": age.total_seconds() / 3600},
        )
        return cache_file.read_bytes()

    def _save_to_cache(self, content: bytes) -> None:
        """Save XLSX content to cache.

        Story 6.9.4 AC5: Caching for large historical file
        """
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        cache_file = self.cache_dir / "eu_oil_bulletin_history.xlsx"
        cache_file.write_bytes(content)
        logger.info("Saved EU Oil Bulletin XLSX to cache")

    async def fetch_diesel_prices(
        self,
        start_date: date,
        end_date: date,
        country: str = "Portugal",
        tax_included: bool = True,
    ) -> list[EUDieselPrice]:
        """Fetch diesel prices for a country.

        Story 6.9.4 AC3: Updated to parse XLSX format

        Args:
            start_date: Start of date range
            end_date: End of date range
            country: Country name (default: Portugal)
            tax_included: If True, fetch prices with taxes (default)

        Returns:
            List of diesel price records
        """
        logger.info(
            "Fetching EU Oil Bulletin diesel prices",
            extra={
                "start": str(start_date),
                "end": str(end_date),
                "country": country,
            },
        )

        # Try cache first (Story 6.9.4 AC5)
        xlsx_content = self._get_cached_xlsx()

        if xlsx_content is None:
            xlsx_content = await self._fetch_xlsx_data()
            self._save_to_cache(xlsx_content)

        # Parse XLSX content
        results = self._parse_xlsx(xlsx_content, country, start_date, end_date, tax_included)

        logger.info(
            "Fetched EU Oil Bulletin diesel prices",
            extra={"record_count": len(results), "country": country},
        )
        return results

    def _parse_xlsx(
        self,
        content: bytes,
        country: str,
        start_date: date,
        end_date: date,
        tax_included: bool = True,
    ) -> list[EUDieselPrice]:
        """Parse EU Oil Bulletin XLSX data.

        Story 6.9.4 AC2: XLSX parsing with openpyxl

        The XLSX file has multiple sheets:
        - "Prices with taxes" - consumer prices including taxes
        - "Prices wo taxes, per CTR" - prices without taxes by country

        Each sheet has dates in column A and country prices in subsequent columns.

        Args:
            content: XLSX file bytes
            country: Country name
            start_date: Filter start date
            end_date: Filter end date
            tax_included: Which sheet to use

        Returns:
            List of diesel price records
        """
        try:
            import openpyxl
        except ImportError as e:
            raise ExternalDataFetchError(
                source="EU_Oil_Bulletin",
                message="openpyxl not installed - required for XLSX parsing",
                original_error=e,
            ) from e

        results: list[EUDieselPrice] = []

        try:
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            logger.warning(
                "Failed to open EU Oil Bulletin XLSX",
                extra={"error": str(e)},
            )
            return results

        # Try to find the sheet (name may vary slightly)
        ws = None
        for name in wb.sheetnames:
            if "taxes" in name.lower():
                if tax_included and "with" in name.lower():
                    ws = wb[name]
                    break
                elif not tax_included and ("wo" in name.lower() or "without" in name.lower()):
                    ws = wb[name]
                    break

        if ws is None:
            # Fallback to first sheet
            ws = wb.active
            if ws is None:
                logger.warning("No active sheet in EU Oil Bulletin XLSX")
                return results

        # Find country column
        # First row contains headers with country names
        country_col_idx: int | None = None
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_str = str(cell_value).strip()
            # Match by country name or code
            if country.lower() in cell_str.lower():
                country_col_idx = idx
                break
            # Also check country code
            country_code = self.COUNTRY_CODES.get(country, "")
            if country_code and country_code.lower() in cell_str.lower():
                country_col_idx = idx
                break

        if country_col_idx is None:
            logger.warning(
                "Country not found in EU Oil Bulletin XLSX headers",
                extra={"country": country, "headers": header_row[:15]},
            )
            return results

        # Parse data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                # First column is date
                date_cell = row[0]
                if date_cell is None:
                    continue

                # Parse date (may be datetime or string)
                if isinstance(date_cell, datetime):
                    record_date = date_cell.date()
                elif isinstance(date_cell, date):
                    record_date = date_cell
                elif isinstance(date_cell, str):
                    # Try common date formats
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
                        try:
                            record_date = datetime.strptime(date_cell, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        continue
                else:
                    continue

                # Filter by date range
                if not (start_date <= record_date <= end_date):
                    continue

                # Get price value
                if country_col_idx >= len(row):
                    continue
                price_cell = row[country_col_idx]
                if price_cell is None:
                    continue

                price = float(price_cell)

                # Skip invalid prices
                if price <= 0 or price > 10:  # Reasonable range for EUR/litre
                    continue

                results.append(
                    EUDieselPrice(
                        date=record_date,
                        price_eur_litre=price,
                        country=country,
                        tax_included=tax_included,
                    )
                )

            except (ValueError, TypeError, IndexError) as e:
                logger.warning(
                    "Failed to parse EU Oil Bulletin row",
                    extra={"error": str(e)},
                )
                continue

        wb.close()

        # Sort by date (oldest first)
        results.sort(key=lambda x: x.date)

        return results

    def _code_to_country(self, code: str) -> str:
        """Convert country code to name."""
        for name, c in self.COUNTRY_CODES.items():
            if c == code:
                return name
        return code

    async def fetch_weekly_prices(
        self,
        week_date: date,
        country: str = "Portugal",
    ) -> EUDieselPrice | None:
        """Fetch diesel price for a specific week.

        Story 6.9.4 AC4: Handle weekly data frequency

        Args:
            week_date: Any date within the target week
            country: Country name

        Returns:
            Diesel price for that week or None
        """
        # EU Oil Bulletin publishes on Mondays
        # Find the Monday of the week containing week_date
        days_since_monday = week_date.weekday()
        monday = week_date - timedelta(days=days_since_monday)

        # Fetch a week range to find the Monday record
        prices = await self.fetch_diesel_prices(
            start_date=monday,
            end_date=monday + timedelta(days=6),
            country=country,
        )

        return prices[0] if prices else None

    # Legacy method for backward compatibility
    async def _fetch_bulletin_data(self, year: int) -> str:
        """Deprecated: Fetch annual oil bulletin data.

        Story 6.9.4: This method is deprecated. Old XML endpoint no longer works.
        Kept for backward compatibility - always returns empty string.

        Args:
            year: Year to fetch (ignored)

        Returns:
            Empty string (XML endpoint deprecated)
        """
        logger.warning(
            "Deprecated _fetch_bulletin_data called - "
            "old XML endpoint no longer works, use fetch_diesel_prices instead"
        )
        return ""

    def _parse_bulletin_xml(
        self,
        content: str,
        country_code: str,
        start_date: date,
        end_date: date,
    ) -> list[EUDieselPrice]:
        """Deprecated: Parse EU Oil Bulletin XML data.

        Story 6.9.4: This method is deprecated. Use _parse_xlsx instead.

        Returns:
            Empty list
        """
        logger.warning("Deprecated _parse_bulletin_xml called - use _parse_xlsx instead")
        return []
