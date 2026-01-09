"""Data parsers for BaseGov client.

Story 8.2 Task 4: Extract parsing logic from basegov.py
Handles IMPIC XLSX, OCDS JSON, and TED API response parsing.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import openpyxl

from raglite.external_data.clients.basegov.parsers_utils import _process_impic_row
from raglite.external_data.models import BaseGovContract
from raglite.shared.logging import get_logger

logger = get_logger(__name__)


def parse_impic_xlsx(
    content: bytes,
    start_date: date,
    end_date: date,
    cpv_code: str | None = None,
) -> list[BaseGovContract]:
    """Parse IMPIC XLSX content to BaseGovContract models.

    Story 6.9.5: Parse dados.gov.pt IMPIC format

    XLSX columns (verified 2025-12-08):
    - idcontrato: Contract ID
    - tipoContrato: Contract type
    - tipoprocedimento: Procedure type
    - objectoContrato: Contract object/title
    - descContrato: Description
    - adjudicante: Contracting entity (NIF - Name)
    - adjudicatarios: Contractor/winner (NIF - Name)
    - dataPublicacao: Publication date
    - dataCelebracaoContrato: Contract celebration date
    - precoContratual: Contract value
    - CPV: CPV codes

    Args:
        content: XLSX bytes
        start_date: Filter start date
        end_date: Filter end date
        cpv_code: CPV code filter (prefix match)

    Returns:
        List of contracts matching filters
    """
    results: list[BaseGovContract] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Find column indices
        col_map = {h: i for i, h in enumerate(headers) if h}

        id_col = col_map.get("idcontrato", 0)
        obj_col = col_map.get("objectoContrato")
        desc_col = col_map.get("descContrato")
        buyer_col = col_map.get("adjudicante")
        winner_col = col_map.get("adjudicatarios")
        pub_date_col = col_map.get("dataPublicacao")
        contract_date_col = col_map.get("dataCelebracaoContrato")
        value_col = col_map.get("precoContratual")
        cpv_col = col_map.get("CPV")

        # Parse data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                contract = _process_impic_row(
                    row=row,
                    start_date=start_date,
                    end_date=end_date,
                    cpv_code=cpv_code,
                    id_col=id_col,
                    obj_col=obj_col,
                    desc_col=desc_col,
                    buyer_col=buyer_col,
                    winner_col=winner_col,
                    pub_date_col=pub_date_col,
                    contract_date_col=contract_date_col,
                    value_col=value_col,
                    cpv_col=cpv_col,
                )
                if contract:
                    results.append(contract)
            except (IndexError, TypeError, ValueError):
                # Skip malformed rows
                continue

        wb.close()

    except Exception as e:
        logger.warning(f"Failed to parse IMPIC XLSX: {e}")
        return []

    return results


def _extract_ocds_items(ocds_data: dict | list) -> list:
    """Extract items array from OCDS data.

    Args:
        ocds_data: OCDS JSON data (dict or list)

    Returns:
        List of OCDS items (releases or records)
    """
    if isinstance(ocds_data, dict):
        return ocds_data.get("releases", ocds_data.get("records", []))
    return ocds_data if isinstance(ocds_data, list) else []


def _extract_ocds_date(item: dict, tender: dict) -> date | None:
    """Extract and parse publication date from OCDS item.

    Args:
        item: OCDS item
        tender: Tender section from item

    Returns:
        Parsed date or None
    """
    pub_date_str = item.get("date", item.get("publishedDate"))
    if not pub_date_str and tender:
        pub_date_str = tender.get("tenderPeriod", {}).get("endDate")

    if not pub_date_str:
        return None

    return date.fromisoformat(pub_date_str[:10])


def _extract_ocds_cpv(tender: dict) -> str | None:
    """Extract CPV code from OCDS tender section.

    Args:
        tender: Tender section from OCDS item

    Returns:
        CPV code or None
    """
    tender_items = tender.get("items", [])
    if tender_items:
        classification = tender_items[0].get("classification", {})
        return classification.get("id", "")
    return None


def _extract_ocds_value(contracts_list: list, awards: list, tender: dict) -> float:
    """Extract contract value from OCDS item.

    Args:
        contracts_list: Contracts section
        awards: Awards section
        tender: Tender section

    Returns:
        Contract value in EUR
    """
    if contracts_list:
        return contracts_list[0].get("value", {}).get("amount", 0)
    if awards:
        return awards[0].get("value", {}).get("amount", 0)
    return tender.get("value", {}).get("amount", 0)


def _extract_ocds_parties(parties: list) -> tuple[str, str]:
    """Extract buyer and supplier from OCDS parties.

    Args:
        parties: Parties section from OCDS item

    Returns:
        Tuple of (buyer, supplier)
    """
    buyer = ""
    supplier = ""
    for party in parties:
        roles = party.get("roles", [])
        if "buyer" in roles:
            buyer = party.get("name", "")
        if "supplier" in roles or "tenderer" in roles:
            supplier = party.get("name", "")
    return buyer, supplier


def parse_ocds_data(
    ocds_data: dict | list,
    start_date: date,
    end_date: date,
    cpv_code: str | None = None,
) -> list[BaseGovContract]:
    """Parse OCDS format data to BaseGovContract models.

    Story 6.9.5 AC3: OCDS parsing

    OCDS structure:
    - releases[] or records[] array
    - Each has tender, awards, contracts sections
    - Uses ocid as unique identifier

    Args:
        ocds_data: OCDS JSON data
        start_date: Filter start date
        end_date: Filter end date
        cpv_code: CPV code filter

    Returns:
        List of BaseGovContract records
    """
    results: list[BaseGovContract] = []
    items = _extract_ocds_items(ocds_data)

    for item in items:
        try:
            tender = item.get("tender", {})
            awards = item.get("awards", [])
            contracts_list = item.get("contracts", [])

            # Extract and validate date
            pub_date = _extract_ocds_date(item, tender)
            if not pub_date:
                continue

            # Filter by date range
            if not (start_date <= pub_date <= end_date):
                continue

            # Extract and filter by CPV
            item_cpv = _extract_ocds_cpv(tender)
            if cpv_code and item_cpv and not item_cpv.startswith(cpv_code[:2]):
                continue

            # Extract value and parties
            value = _extract_ocds_value(contracts_list, awards, tender)
            buyer, supplier = _extract_ocds_parties(item.get("parties", []))

            results.append(
                BaseGovContract(
                    publication_date=pub_date,
                    contract_id=item.get("ocid", item.get("id", "")),
                    description=tender.get("title", tender.get("description", "")),
                    contract_value_eur=float(value),
                    contracting_entity=buyer,
                    contractor=supplier,
                    cpv_code=item_cpv,
                    execution_location="Portugal",
                )
            )

        except (ValueError, KeyError, TypeError) as e:
            logger.warning(
                "Failed to parse OCDS item",
                extra={"error": str(e)},
            )
            continue

    return results


def parse_ted_notices(data: dict) -> list[BaseGovContract]:
    """Parse TED API response to BaseGovContract models.

    Story 6.9.5 AC6: TED API parsing

    Args:
        data: TED API response

    Returns:
        List of contracts
    """
    results: list[BaseGovContract] = []

    notices = data.get("notices", data.get("results", []))

    for notice in notices:
        try:
            # Get publication date
            pub_date_str = notice.get("publication-date", notice.get("publicationDate"))
            if not pub_date_str:
                continue

            pub_date = date.fromisoformat(pub_date_str[:10])

            # Get value (may be in different formats)
            value = notice.get("total-value", notice.get("totalValue", 0))
            if isinstance(value, dict):
                value = value.get("amount", 0)
            if isinstance(value, str):
                value = float(value.replace(",", ".").replace(" ", ""))

            # Get CPV code
            cpv = notice.get("cpv", notice.get("cpvCode", ""))
            if isinstance(cpv, list):
                cpv = cpv[0] if cpv else ""

            results.append(
                BaseGovContract(
                    publication_date=pub_date,
                    contract_id=str(notice.get("publication-number", notice.get("id", ""))),
                    description=notice.get("notice-title", notice.get("title", "")),
                    contract_value_eur=float(value) if value else 0.0,
                    contracting_entity=notice.get("buyer-name", notice.get("buyer", "")),
                    contractor=notice.get("winner-name", notice.get("winner", "")),
                    cpv_code=cpv,
                    execution_location="Portugal",
                )
            )

        except (ValueError, KeyError, TypeError) as e:
            logger.warning(
                "Failed to parse TED notice",
                extra={"error": str(e)},
            )
            continue

    return results
